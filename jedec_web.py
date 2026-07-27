#!/usr/bin/env python3
"""
Package Reliability Qualification Suite
JEDEC reliability test calculator and qualification report generator.

Run:  python3 jedec_web.py
Open: http://localhost:5000

Requires: tornado (stdlib-only otherwise — reportlab for PDF, openpyxl for Excel export)
"""

from __future__ import annotations
import os, sys, math, io, uuid, webbrowser, base64, hashlib

# ── Auto-install optional dependencies if missing ─────────────────────────────
def _ensure_pkg(import_name: str, pip_name: str) -> None:
    try:
        __import__(import_name)
    except ImportError:
        import subprocess
        print(f"[startup] Installing {pip_name}…", flush=True)
        # Try system-wide first; fall back to --user if permission is denied.
        for extra in ([], ["--user"]):
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", pip_name, "-q"] + extra,
                capture_output=True,
            )
            if result.returncode == 0:
                break
        else:
            print(f"[startup] WARNING: could not install {pip_name}. "
                  f"Run:  pip install --user {pip_name}", flush=True)

_ensure_pkg("openpyxl", "openpyxl")
from datetime import datetime

# ── Compatibility patch: Python 3.8 + macOS OpenSSL rejects usedforsecurity ──
_orig_md5 = hashlib.md5
def _patched_md5(*args, **kwargs):
    kwargs.pop("usedforsecurity", None)
    return _orig_md5(*args, **kwargs)
hashlib.md5 = _patched_md5

import tornado.ioloop
import tornado.web

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from jedec_calc import (TESTS, PRECOND, PART_TYPE_LABELS,
                        min_sample_size, demonstrated_reliability, pass_fail as _pf,
                        min_sample_size_ltpd, TABLE_A, TABLE_A_LTPD)
import jedec_db as _db

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable, KeepTogether, PageBreak,
                                Image as RLImage)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Constants ──────────────────────────────────────────────────────────────────
RTD_SENSORS    = [f"T{i}" for i in range(1, 17)]
BOND_TYPES     = ["Cu TCB", "Ag Sinter", "Ag TCB"]
CSAM_THRESHOLD = 95.0

# ── JEDEC spec document URLs — served locally from /specs/ ────────────────────
# Each entry is a list of (label, url) pairs so tests with multiple standards
# can link each document individually.
# Files marked TODO are not yet in the specs/ folder — add them to enable the link.
SPEC_URLS: dict[str, list[tuple[str, str]]] = {
    "uhast":       [("JESD22-A118B",  "/specs/JESD22-A118B.pdf")],
    "tc":          [("JESD22-A104F",  "/specs/JESD22-A104F.pdf")],
    "tshock":      [("JESD22-A106B",  "/specs/JESD22-A106B.pdf")],
    "mshock":      [("JESD22-B110B",  "/specs/JESD22-B110B.pdf")],
    "vib":         [("JESD22-B103B",  "/specs/JESD22-B103B.pdf")],
    "pc":          [("JESD22-A122",   "/specs/JESD22-A122.pdf")],
    "ptc":         [("JESD22-A105D",  "/specs/JESD22-A105D.pdf")],
    "hts":         [("JESD22-A103D",  "/specs/JESD22-A103D.pdf")],
    "shadow_moire":[("JESD22-B112C",  "/specs/JESD22-B112C.pdf")],
    "htol":        [("JESD22-A108G",  "/specs/JESD22-A108G.pdf"),
                    ("JESD85",        "/specs/JESD85.pdf")],
    "elfr":        [("JESD22-A108G",  "/specs/JESD22-A108G.pdf"),
                    ("JESD74A",       "/specs/JESD74A.pdf")],
    "thb":         [("JESD22-A110",   "/specs/JESD22-A110.pdf")],
    "esd_cdm":     [("JS-002",        "/specs/JS-002.pdf")],
    "esd_hbm":     [("JS-001",        "/specs/JS-001.pdf")],
    "latchup":     [("JESD78F",       "/specs/JESD78F.pdf")],
    "jesd47":      [("JESD47I",       "/specs/JESD47I.pdf")],
    "precond":     [("JESD22-A113I",  "/specs/JESD22-A113I.pdf"),
                    ("J-STD-020F",    "/specs/J-STD-020F.pdf")],
}

# ── TC condition table (JESD22-A104F Table 1 & Table 3) ───────────────────────
# keys: condition letter → tmin, tmax (°C), typical cycles/hr, applicable soak modes
# Soak mode min times: 1→1 min, 2→5 min, 3→10 min, 4→15 min
# t4: Table 4 (solder interconnect) cycle rates keyed by soak mode number (str)
# Conditions not in Table 4 have t4={}; display falls back to Table 3 "cycles" value.
TC_CONDITIONS = {
    "A": {"tmin": -55, "tmax":  85, "cycles": "2–3",   "soak": [1, 2, 3],    "t4": {}},
    "B": {"tmin": -55, "tmax": 125, "cycles": "2–3",   "soak": [1, 2],       "t4": {}},
    "C": {"tmin": -65, "tmax": 150, "cycles": "2",     "soak": [1, 2],       "t4": {}},
    "G": {"tmin": -40, "tmax": 125, "cycles": "<1–2",  "soak": [1, 2, 3, 4], "t4": {"2":"2","3":"\u22642","4":"<1"}},
    "H": {"tmin": -55, "tmax": 150, "cycles": "2",     "soak": [1, 2],       "t4": {}},
    "I": {"tmin": -40, "tmax": 115, "cycles": "1–2",   "soak": [1, 2, 3, 4], "t4": {"2":"2","3":"\u22642","4":"<1"}},
    "J": {"tmin":   0, "tmax": 100, "cycles": "1–3",   "soak": [1, 2, 3, 4], "t4": {"2":"2","3":"\u22642","4":"<1"}},
    "K": {"tmin":   0, "tmax": 125, "cycles": "1–3",   "soak": [1, 2, 3, 4], "t4": {"2":"2","3":"\u22642","4":"<1"}},
    "L": {"tmin": -55, "tmax": 110, "cycles": "1–3",   "soak": [1, 2, 3, 4], "t4": {"2":"2","3":"\u22642","4":"<1"}},
    "M": {"tmin": -40, "tmax": 150, "cycles": "1–3",   "soak": [1, 2, 3, 4], "t4": {}},
    "N": {"tmin": -40, "tmax":  85, "cycles": "1–3",   "soak": [1, 2, 3],    "t4": {}},
    "R": {"tmin": -25, "tmax": 125, "cycles": "1–2",   "soak": [1, 2],       "t4": {"1":"2"}},
    "T": {"tmin": -40, "tmax": 100, "cycles": "1–2",   "soak": [3, 4],       "t4": {"3":"\u22642","4":"<1"}},
}
TC_SOAK_TIMES = {1: 1, 2: 5, 3: 10, 4: 15}  # mode → min soak time (minutes)

# ── T-Shock condition table (JESD22-A106B Table 1) ────────────────────────────
# Step 1 = hot bath, Step 2 = cold bath; tolerances ±10/0 & 0/−10 respectively
# fluid_s1: recommended fluid for Step 1 (hot); fluid_s2: always Perfluorocarbon
TSHOCK_CONDITIONS = {
    "A": {"hot":  85, "cold": -40, "fluid_s1": "Water or Perfluorocarbon", "fluid_s2": "Perfluorocarbon"},
    "B": {"hot": 100, "cold":   0, "fluid_s1": "Perfluorocarbon",          "fluid_s2": "Perfluorocarbon"},
    "C": {"hot": 125, "cold": -55, "fluid_s1": "Perfluorocarbon",          "fluid_s2": "Perfluorocarbon"},
    "D": {"hot": 150, "cold": -65, "fluid_s1": "Perfluorocarbon",          "fluid_s2": "Perfluorocarbon"},
}

# ── UHAST condition table (JESD22-A118B Table 1) ──────────────────────────────
# temp_db: dry-bulb °C (±2); rh: relative humidity % (±5)
# temp_wb: wet-bulb °C; vp_kpa / vp_psia: vapor pressure
# duration: typical test duration per condition
UHAST_CONDITIONS = {
    "A": {"temp_db": 130, "rh": 85, "temp_wb": 124.7, "vp_kpa": 230,  "vp_psia": 33.3, "duration": "96 hours (−0, +2)"},
    "B": {"temp_db": 110, "rh": 85, "temp_wb": 105.2, "vp_kpa": 122,  "vp_psia": 17.7, "duration": "264 hours (−0, +2)"},
}

# ── THB condition table (JESD22-A110 §3.1 Table 1) ───────────────────────────
# Same temperature/humidity extremes as UHAST; difference is bias (Vdd) applied.
# tmin_tol: dry-bulb tolerance ±2°C; rh_tol: ±5%
THB_CONDITIONS = {
    "A": {"temp_db": 130, "rh": 85, "temp_wb": 124.7, "vp_kpa": 230, "vp_psia": 33.3, "duration": "96 hours (−0, +2)"},
    "B": {"temp_db": 110, "rh": 85, "temp_wb": 105.2, "vp_kpa": 122, "vp_psia": 17.7, "duration": "264 hours (−0, +2)"},
}

# ── M-Shock condition table (JESD22-B110B Table 1 — free state test levels) ───
# accel_g: acceleration peak (g); pulse_ms: half-sine pulse duration (ms)
# vel_cms / vel_ins: velocity change; drop_cm / drop_in: equivalent drop height
MSHOCK_CONDITIONS = {
    "H": {"accel_g": 2900, "pulse_ms": 0.3, "vel_cms": 543,  "vel_ins": 214,  "drop_cm": 150,  "drop_in": 59},
    "G": {"accel_g": 2000, "pulse_ms": 0.4, "vel_cms": 499,  "vel_ins": 197,  "drop_cm": 127,  "drop_in": 50},
    "B": {"accel_g": 1500, "pulse_ms": 0.5, "vel_cms": 468,  "vel_ins": 184,  "drop_cm": 112,  "drop_in": 44},
    "F": {"accel_g":  900, "pulse_ms": 0.7, "vel_cms": 393,  "vel_ins": 155,  "drop_cm":  78.9,"drop_in": 31},
    "A": {"accel_g":  500, "pulse_ms": 1.0, "vel_cms": 312,  "vel_ins": 123,  "drop_cm":  49.7,"drop_in": 20},
    "E": {"accel_g":  340, "pulse_ms": 1.2, "vel_cms": 255,  "vel_ins": 100,  "drop_cm":  33.1,"drop_in": 13},
    "D": {"accel_g":  200, "pulse_ms": 1.5, "vel_cms": 187,  "vel_ins":  73.7,"drop_cm":  17.9,"drop_in":  7},
    "C": {"accel_g":  100, "pulse_ms": 2.0, "vel_cms": 125,  "vel_ins":  49.2,"drop_cm":   7.9,"drop_in":  3},
}

# ── Vibration condition tables (JESD22-B103B) ─────────────────────────────────
# Table 1 — Sinusoidal component test levels (conditions 1–8)
VIB_SIN_CONDITIONS = {
    "1": {"accel_g": 20,    "disp_in": 0.060,   "disp_mm": 1.5,    "xover_hz": 80,  "fmin_hz": 20, "fmax_hz": 2000},
    "2": {"accel_g": 10,    "disp_in": 0.040,   "disp_mm": 1.0,    "xover_hz": 70,  "fmin_hz": 10, "fmax_hz": 1000},
    "3": {"accel_g":  3,    "disp_in": 0.030,   "disp_mm": 0.75,   "xover_hz": 45,  "fmin_hz":  5, "fmax_hz":  500},
    "4": {"accel_g":  1,    "disp_in": 0.020,   "disp_mm": 0.5,    "xover_hz": 31,  "fmin_hz":  5, "fmax_hz":  500},
    "5": {"accel_g":  0.3,  "disp_in": 0.010,   "disp_mm": 0.25,   "xover_hz": 24,  "fmin_hz":  5, "fmax_hz":  500},
    "6": {"accel_g":  0.1,  "disp_in": 0.005,   "disp_mm": 0.125,  "xover_hz": 20,  "fmin_hz":  5, "fmax_hz":  500},
    "7": {"accel_g":  0.01, "disp_in": 0.001,   "disp_mm": 0.039,  "xover_hz": 14,  "fmin_hz":  5, "fmax_hz":  500},
    "8": {"accel_g":  0.001,"disp_in": 0.0005,  "disp_mm": 0.0127, "xover_hz":  6.2,"fmin_hz":  5, "fmax_hz":  500},
}
# Table 2 — Random vibration overall test levels (conditions A–I)
# rms_g: RMS acceleration; vel_ins: RMS velocity; disp_in: RMS displacement; sigma_in: 6×RMS displacement (3σ pk-pk)
VIB_RAN_CONDITIONS = {
    "A": {"rms_g": 6.27,   "vel_ins": 29.0,  "disp_in": 0.926,   "sigma_in": 5.55},
    "B": {"rms_g": 3.10,   "vel_ins": 13.2,  "disp_in": 0.426,   "sigma_in": 2.56},
    "C": {"rms_g": 1.24,   "vel_ins":  5.22, "disp_in": 0.178,   "sigma_in": 1.07},
    "D": {"rms_g": 1.11,   "vel_ins":  1.64, "disp_in": 0.0310,  "sigma_in": 0.186},
    "E": {"rms_g": 0.686,  "vel_ins":  0.703,"disp_in": 0.00543, "sigma_in": 0.0326},
    "F": {"rms_g": 0.416,  "vel_ins":  0.425,"disp_in": 0.00355, "sigma_in": 0.0213},
    "G": {"rms_g": 0.246,  "vel_ins":  0.215,"disp_in": 0.00171, "sigma_in": 0.0102},
    "H": {"rms_g": 0.123,  "vel_ins":  0.113,"disp_in": 0.000832,"sigma_in": 0.00499},
    "I": {"rms_g": 0.0626, "vel_ins":  0.0589,"disp_in":0.000395,"sigma_in": 0.002237},
}

# ── Power Cycling condition table (JESD22-A122 Table 2) ───────────────────────
# tmin / tmax: Tcycle(min/max) in °C; tolerance ±5°C on both
# delta_t: nominal ΔT = tmax − tmin
# Typical cycle rate: 2–6 cycles/hr (NOTE 2)
PC_CONDITIONS = {
    "A": {"tmin":  25, "tmax": 100, "delta_t":  75},
    "B": {"tmin":  25, "tmax": 125, "delta_t": 100},
    "C": {"tmin":  10, "tmax": 100, "delta_t":  90},
    "D": {"tmin":  10, "tmax": 125, "delta_t": 115},
    "E": {"tmin":  40, "tmax": 100, "delta_t":  60},
}
# Table 1: method combinations (power × cooling)
PC_METHODS = [
    ("Constant Power / Constant Cooling",  "Easiest to implement; requires careful test detail"),
    ("Constant Power / Variable Cooling",  "Best for functional devices & tightly controlled heater resistance; allows Tj variation modeling"),
    ("Variable Power / Constant Cooling",  "Closely matches actual product profile; requires surge control"),
    ("Variable Power / Variable Cooling",  "Most flexible but most complex; best for devices with complex chip power maps"),
]

# ── PTC condition table (JESD22-A105D Table 1) ───────────────────────────────
# tmin/tmax: temperature extremes °C (tolerances: tmin +0/−10, tmax +10/−0)
# trans_min: max transition time between extremes (minutes)
# dwell_min: min dwell time at each extreme (minutes)
PTC_CONDITIONS = {
    "A": {"tmin": -40, "tmax":  85, "tmin_tol": "(+0, −10)", "tmax_tol": "(+10, −0)", "trans_min": 20, "dwell_min": 10},
    "B": {"tmin": -40, "tmax": 125, "tmin_tol": "(+0, −10)", "tmax_tol": "(+10, −0)", "trans_min": 30, "dwell_min": 10},
}

# ── HTS condition table (JESD22-A103D Table 1) ────────────────────────────────
# temp_c: storage temperature °C; tolerance −0/+10°C
# JESD47 default duration: 1000 hours at Condition B
HTS_CONDITIONS = {
    "A": {"temp_c": 125},
    "B": {"temp_c": 150},
    "C": {"temp_c": 175},
    "D": {"temp_c": 200},
    "E": {"temp_c": 250},
    "F": {"temp_c": 300},
}

# Known JEDEC durations in hours per (test_key, condition_key).
# Tests not listed here use GANTT task duration (weeks × 168 h) as fallback.
_CONDITION_HOURS: dict[str, dict[str, float]] = {
    "uhast": {"A":  96.0, "B": 264.0},
    "thb":   {"A":  96.0, "B": 264.0},
    "hts":   {k: 1000.0 for k in ["A","B","C","D","E","F"]},
    # TC: approximate hours (500–900 cycles × ~1 h/cycle)
    "tc":    {"A":500,"B":500,"C":500,"G":700,"H":500,"I":700,
              "J":600,"K":600,"L":600,"M":600,"N":700,"R":700,"T":900},
}

# ── Per-test selectable conditions (for Planner + Schedule tracker) ───────────
# Maps test_key → list of (condition_key, display_label) tuples
# Duration appended where a fixed hour count is known.
_TEST_CONDITION_OPTIONS: dict[str, list[tuple[str, str]]] = {
    "uhast":  [(k, f"Cond {k} — {v['temp_db']}°C / {v['rh']}% RH / {v['duration']}")
               for k, v in UHAST_CONDITIONS.items()],
    "tc":     [(k, (f"Cond {k} — {v['tmin']:+d}°C to {v['tmax']:+d}°C"
                    + (f" / {int(_CONDITION_HOURS['tc'][k])}h"
                       if k in _CONDITION_HOURS['tc'] else "")))
               for k, v in TC_CONDITIONS.items()],
    "tshock": [(k, f"Cond {k} — {v['cold']:+d}°C / {v['hot']:+d}°C")
               for k, v in TSHOCK_CONDITIONS.items()],
    "mshock": [(k, f"SC {k} — {v['accel_g']}g / {v['pulse_ms']} ms")
               for k, v in MSHOCK_CONDITIONS.items()],
    "hts":    [(k, f"Cond {k} — +{v['temp_c']}°C / 1000h")
               for k, v in HTS_CONDITIONS.items()],
    "pc":     [(k, f"Cond {k} — {v['tmin']}–{v['tmax']}°C / ΔT {v['delta_t']}°C")
               for k, v in PC_CONDITIONS.items()],
    "ptc":    [(k, f"Cond {k} — {v['tmin']:+d}°C to {v['tmax']:+d}°C"
                    f" / trans ≤{v['trans_min']}min / dwell ≥{v['dwell_min']}min")
               for k, v in PTC_CONDITIONS.items()],
    "thb":    [(k, f"Cond {k} — {v['temp_db']}°C / {v['rh']}% RH / {v['duration']}")
               for k, v in THB_CONDITIONS.items()],
    "vib":    (
        [(f"sin_{k}", f"Sin {k} — {v['fmin_hz']}–{v['fmax_hz']} Hz / {v['accel_g']}g / {v['disp_mm']} mm")
         for k, v in VIB_SIN_CONDITIONS.items()]
        + [(f"ran_{k}", f"Ran {k} — {v['rms_g']}g RMS")
           for k, v in VIB_RAN_CONDITIONS.items()]
    ),
}

# ── In-memory sessions ─────────────────────────────────────────────────────────
SESSIONS: dict = {}

COOKIE_SECRET = "jedec-df-calc-2025-secret-key-static"

def _get_or_create(handler):
    raw = handler.get_secure_cookie("sid")
    if raw:
        sid = raw.decode()
        if sid in SESSIONS:
            return sid, SESSIONS[sid]
    sid = str(uuid.uuid4())
    SESSIONS[sid] = {"part_type": "ttv", "last_report": None}
    handler.set_secure_cookie("sid", sid, expires_days=1)
    return sid, SESSIONS[sid]

# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_gantt_anchor(pid: int):
    """Return the Monday-aligned gantt anchor date for a project."""
    from datetime import date, timedelta
    meta = _db.get_meta(pid)
    start_date_str = (meta.get("gantt_start_date") or "").strip()
    try:
        return date.fromisoformat(start_date_str)
    except ValueError:
        today = date.today()
        return today - timedelta(days=today.weekday())


def _iso_week_to_relative(anchor, iso_week: int) -> int:
    """
    Convert a submitted ISO week number to a 1-based relative week from anchor.

    Year inference: if iso_week >= anchor's ISO week → same year as anchor.
    Otherwise → anchor year + 1 (task scheduled into the following year).
    """
    from datetime import date
    iso = anchor.isocalendar()
    anchor_iso_week = iso[1]
    anchor_year     = iso[0]
    year = anchor_year if iso_week >= anchor_iso_week else anchor_year + 1
    try:
        target_monday = date.fromisocalendar(year, iso_week, 1)
    except ValueError:
        return 1
    rel = (target_monday - anchor).days // 7 + 1
    return max(1, rel)


def _csam_eval(before_pc, after_pc, after_test):
    """
    Evaluate CSAM bonded area against threshold (95%).
    Returns (status_text, badge_class) tuple.
    """
    if before_pc is None:
        return ("No data", "secondary")
    if before_pc < CSAM_THRESHOLD:
        return ("Rejected — pre-PC < 95%", "danger")
    if after_pc is None:
        return ("Awaiting post-PC", "warning")
    if after_pc < CSAM_THRESHOLD:
        return ("Fail — Preconditioning", "danger")
    if after_test is None:
        return ("Awaiting post-test", "warning")
    if after_test < CSAM_THRESHOLD:
        return ("Fail — Post-test CSAM", "danger")
    return ("Pass", "success")

_DIE_EXCLUDED = {"pc", "shadow_moire"}

def applicable_tests(part_type: str) -> dict:
    if part_type == "ttv":
        return {k: v for k, v in TESTS.items() if not v["active_devices"]}
    if part_type == "die":
        return {k: v for k, v in TESTS.items()
                if not v["active_devices"] and k not in _DIE_EXCLUDED}
    return TESTS

STATUS_OPTS = [
    ("ns", "Not Started",      "secondary"),
    ("ip", "In Progress",      "warning"),
    ("co", "Complete",         "success"),
    ("na", "N/A",              "light"),
]
# Status options for characterization-only tests (no Pass/Fail)
STATUS_OPTS_CHAR = [
    ("ns", "Not Started",      "secondary"),
    ("ip", "In Progress",      "warning"),
    ("ch", "Characterized",    "info"),
    ("na", "N/A",              "light"),
]
STATUS_LABEL = {s[0]: s[1].upper() for s in STATUS_OPTS + STATUS_OPTS_CHAR}
STATUS_COLOR = {s[0]: s[2] for s in STATUS_OPTS + STATUS_OPTS_CHAR}

# ── Base HTML Layout ───────────────────────────────────────────────────────────

def _page(active: str, part_type: str, body: str, title: str = "Package Reliability",
          project: dict = None, active_sub: str = "") -> str:
    top_nav_links = [
        ("lookup",   "Test Lookup",           "/lookup"),
        ("projects", "Qualification Projects", "/projects"),
    ]
    nav = "".join(
        f'<li class="nav-item">'
        f'<a class="nav-link {"active" if k==active else ""} df-topnav-link"'
        f' href="{href}" data-tab="{k}">{label}</a></li>'
        for k, label, href in top_nav_links
    )
    # Project sub-nav (only rendered when inside a project)
    if project:
        pid = project["id"]
        pname = project["name"]
        sub_links = [
            ("overview",    "Overview",  f"/projects/{pid}"),
            ("sample-size", "Planner",   f"/projects/{pid}/sample-size"),
            ("report",      "Reporting", f"/projects/{pid}/report"),
            ("csam",        "CSAM Gallery", f"/projects/{pid}/csam"),
            ("tracker",     "Schedule",  f"/projects/{pid}/tracker"),
        ]
        sub_nav_items = "".join(
            f'<li class="nav-item" style="margin-right:.25rem">'
            f'<a class="nav-link px-3 py-1" href="{href}" '
            f'style="font-size:.8rem;font-weight:{"600" if k==active_sub else "400"};'
            f'color:{"var(--df-accent)" if k==active_sub else "var(--df-mid)"};'
            f'border-bottom:{"2px solid var(--df-accent)" if k==active_sub else "2px solid transparent"};'
            f'border-radius:0;white-space:nowrap">{label}</a></li>'
            for k, label, href in sub_links
        )
        project_subnav = f"""
<div style="background:var(--df-white);border-bottom:1px solid var(--df-border);padding:0">
  <div class="container-xl d-flex align-items-center" style="min-height:44px;gap:.5rem">
    <a href="/projects" class="text-decoration-none"
       style="font-size:.78rem;color:var(--df-grey);white-space:nowrap;padding-right:.75rem;
              border-right:1px solid var(--df-border)">
      <i class="bi bi-arrow-left me-1"></i>Projects
    </a>
    <span style="font-size:.82rem;font-weight:600;color:var(--df-charcoal);
                 white-space:nowrap;padding-right:.75rem;border-right:1px solid var(--df-border)"
          title="{pname}">{pname[:40]}{"…" if len(pname)>40 else ""}</span>
    <ul class="navbar-nav flex-row mb-0" style="gap:.1rem">{sub_nav_items}</ul>
  </div>
</div>"""
    else:
        project_subnav = ""
    pt_label = {"ttv": "TTV", "die": "Die"}.get(part_type, "Active")
    pt_full  = {"ttv": "Thermal Test Vehicle", "die": "Die"}.get(part_type, "Active Device")
    # Pill display: "TTV — Thermal Test Vehicle", "Die", "Active Device"
    _pt_pill_label = (f"{pt_label} \u2014 {pt_full}" if pt_label != pt_full else pt_full)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title} — Package Reliability</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
  <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css" rel="stylesheet">
  <style>
    /* ── DF Brand Tokens ───────────────────────────── */
    :root {{
      --df-black:   #111111;
      --df-charcoal:#1c1c1c;
      --df-mid:     #555555;
      --df-grey:    #888888;
      --df-border:  #e0e0e0;
      --df-bg:      #f8f8f8;
      --df-white:   #ffffff;
      --df-accent:  #c8432a;   /* terracotta — from DF "READ MORE" links */
      --df-accent-h:#a83523;   /* hover */
      --df-pass:    #2d7a4f;
      --df-fail:    #c8432a;
    }}

    /* ── Base ───────────────────────────────────────── */
    body {{
      background: var(--df-white);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
      color: var(--df-black);
      font-size: .9rem;
    }}

    /* ── Navbar ─────────────────────────────────────── */
    .navbar {{
      background: var(--df-white) !important;
      border-bottom: 1px solid var(--df-border);
      padding-top: .9rem;
      padding-bottom: .9rem;
    }}
    .navbar-brand {{
      font-size: .9rem;
      font-weight: 500;
      color: var(--df-black) !important;
      display: flex;
      align-items: center;
      gap: .4rem;
    }}
    .navbar .nav-link {{
      color: var(--df-mid) !important;
      font-size: .875rem;
      font-weight: 400;
      padding: .5rem .9rem;
      border-bottom: 2px solid transparent;
    }}
    .navbar .nav-link:hover {{ color: var(--df-accent) !important; }}
    .navbar .nav-link.active {{
      color: var(--df-accent) !important;
      border-bottom: 2px solid var(--df-accent);
    }}
    .pt-pill {{
      font-size: .8rem;
      color: var(--df-grey);
      border: 1px solid var(--df-border);
      padding: .25rem .7rem;
      text-decoration: none;
      display: flex;
      align-items: center;
      gap: .4rem;
      transition: border-color .15s;
    }}
    .pt-pill:hover {{ border-color: var(--df-accent); color: var(--df-accent); }}
    .pt-pill .pt-dot {{
      width: 7px; height: 7px;
      border-radius: 50%;
      display: inline-block;
    }}

    /* ── Cards ──────────────────────────────────────── */
    .card {{
      border: 1px solid var(--df-border) !important;
      border-radius: 0 !important;
      box-shadow: none !important;
    }}
    .card-df {{
      background: var(--df-bg);
      color: var(--df-charcoal);
      padding: .65rem 1.25rem;
      border-radius: 0;
      font-size: .82rem;
      font-weight: 600;
      border-bottom: 1px solid var(--df-border);
    }}
    .card-df h5, .card-df h6 {{ font-size: .82rem; font-weight: 600; margin: 0; }}

    /* ── Buttons ────────────────────────────────────── */
    .btn {{
      border-radius: 0 !important;
      font-size: .85rem;
      font-weight: 400;
    }}
    .btn-primary {{
      background: var(--df-accent) !important;
      border-color: var(--df-accent) !important;
      color: #fff !important;
    }}
    .btn-primary:hover {{
      background: var(--df-accent-h) !important;
      border-color: var(--df-accent-h) !important;
    }}
    .btn-danger {{
      background: var(--df-charcoal) !important;
      border-color: var(--df-charcoal) !important;
      color: #fff !important;
    }}
    .btn-danger:hover {{ background: #333 !important; border-color: #333 !important; }}
    .btn-outline-secondary {{
      border-color: var(--df-border) !important;
      color: var(--df-black) !important;
    }}
    .btn-outline-secondary:hover {{ background: var(--df-bg) !important; }}
    .btn-outline-primary {{
      border-color: var(--df-black) !important;
      color: var(--df-black) !important;
      background: transparent !important;
    }}
    .btn-outline-primary:hover {{ background: var(--df-black) !important; color: #fff !important; }}
    .btn-lg {{ font-size: .78rem !important; padding: .75rem 2rem !important; }}

    /* ── Form controls ──────────────────────────────── */
    .form-control, .form-select {{
      border-radius: 0 !important;
      border-color: var(--df-border) !important;
      font-size: .88rem;
    }}
    .form-control:focus, .form-select:focus {{
      border-color: var(--df-black) !important;
      box-shadow: none !important;
    }}
    .form-label {{ font-size: .82rem; color: var(--df-mid); margin-bottom: .3rem; }}
    .form-control.is-invalid {{ border-color: var(--df-accent) !important; }}
    .form-control.is-invalid:focus {{ box-shadow: none !important; }}

    /* ── Tables ─────────────────────────────────────── */
    .tbl-header th {{
      background: var(--df-bg) !important;
      color: var(--df-mid) !important;
      font-size: .8rem !important;
      font-weight: 600 !important;
      border-bottom: 1px solid var(--df-border) !important;
    }}
    .table {{ border-color: var(--df-border); }}
    .table-sm td, .table-sm th {{ padding: .55rem .75rem; }}

    /* ── Badges ─────────────────────────────────────── */
    .badge {{ border-radius: 2px !important; font-weight: 400; font-size: .75rem; }}
    .bg-success {{ background: #d4edda !important; color: #1a5c35 !important; }}
    .bg-danger  {{ background: #f8d7da !important; color: #842029 !important; }}
    .bg-warning {{ background: #fff3cd !important; color: #664d03 !important; }}
    .bg-secondary {{ background: var(--df-bg) !important; color: var(--df-grey) !important; border: 1px solid var(--df-border); }}
    .bg-light   {{ background: var(--df-bg) !important; color: var(--df-mid) !important; border: 1px solid var(--df-border); }}

    /* ── Precond bar ────────────────────────────────── */
    .precond-bar {{
      background: #fdf6f4;
      border-left: 3px solid var(--df-accent);
      padding: .75rem 1rem;
      margin-bottom: 1.5rem;
      font-size: .85rem;
      color: var(--df-mid);
    }}

    /* ── Stat numbers ───────────────────────────────── */
    .stat-num {{ font-size: 2.4rem; font-weight: 300; line-height: 1.1; }}
    .result-pass {{ color: var(--df-pass); font-size: 1.5rem; font-weight: 300; letter-spacing: -.01em; }}
    .result-fail {{ color: var(--df-fail); font-size: 1.5rem; font-weight: 300; letter-spacing: -.01em; }}

    /* ── Report wrapper ─────────────────────────────── */
    .report-wrap {{ background: var(--df-white); border: 1px solid var(--df-border); padding: 2.5rem 3rem; }}
    .rpt-header {{
      background: var(--df-charcoal);
      color: #fff;
      padding: 1.4rem 2rem;
      margin-bottom: 1.5rem;
    }}
    .rpt-header .rpt-title {{
      font-size: .72rem;
      font-weight: 400;
      letter-spacing: .18em;
      text-transform: uppercase;
    }}
    .rpt-header .rpt-sub {{
      font-size: .75rem;
      color: rgba(255,255,255,.5);
      margin-top: .25rem;
      letter-spacing: .06em;
    }}

    /* ── Misc ───────────────────────────────────────── */
    h4, h5 {{ font-weight: 500; }}
    h6 {{ font-weight: 600; }}
    .text-muted {{ color: var(--df-grey) !important; }}
    hr {{ border-color: var(--df-border); }}
    .alert {{ border-radius: 0 !important; }}
    .alert-warning {{ background: #fdf6f4; border-color: var(--df-accent); color: var(--df-mid); }}
    .accordion-button {{ border-radius: 0 !important; font-size: .88rem; }}
    .accordion-item {{ border-radius: 0 !important; border-color: var(--df-border); }}
    .accordion-button:not(.collapsed) {{ background: var(--df-bg); color: var(--df-black); box-shadow: none; }}
    .accordion-button:focus {{ box-shadow: none; }}

    @media print {{
      .no-print {{ display: none !important; }}
      body {{ background: #fff; }}
      .report-wrap {{ border: none; padding: 0; }}
    }}
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg">
  <div class="container-xl">
    <a class="navbar-brand" href="/">
      Package Reliability
    </a>
    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#nb"
            style="border:1px solid var(--df-border)">
      <span class="navbar-toggler-icon"></span>
    </button>
    <div class="collapse navbar-collapse" id="nb">
      <ul class="navbar-nav me-auto mb-0">{nav}</ul>
      <button id="navbarAdminBtn" class="btn btn-sm me-2" onclick="ganttAdminLogin()"
        style="border:1px solid #d1d5db;color:#6b7280;background:#fff;font-size:.78rem;padding:.25rem .6rem">
        <i class="bi bi-shield-lock me-1"></i>Admin
      </button>
      <a href="/part-type" class="pt-pill text-decoration-none">
        <span class="pt-dot" style="background:{"#c8432a" if part_type=="ttv" else "#8b5cf6" if part_type=="die" else "#2d7a4f"}"></span>
        {_pt_pill_label}
        <i class="bi bi-pencil-square ms-1" style="font-size:.65rem"></i>
      </a>
    </div>
  </div>
</nav>
{project_subnav}
<div class="container-xl py-5">
{body}
</div>
<!-- Global Admin Login Modal -->
<div class="modal fade" id="adminLoginModal" tabindex="-1">
  <div class="modal-dialog modal-sm">
    <div class="modal-content">
      <div class="modal-header">
        <h6 class="modal-title mb-0"><i class="bi bi-shield-lock me-2"></i>Admin Login</h6>
        <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <label class="form-label" style="font-size:.83rem">Password</label>
        <input type="password" class="form-control form-control-sm" id="adminPwInput"
               placeholder="Enter password" onkeydown="if(event.key==='Enter')ganttAdminSubmit()">
        <div id="adminPwError" style="display:none;font-size:.78rem;color:#dc2626;margin-top:6px">
          Incorrect password.
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
        <button type="button" class="btn btn-sm" onclick="ganttAdminSubmit()"
          style="background:var(--df-accent);color:#fff;border:none">Unlock</button>
      </div>
    </div>
  </div>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script>
// ── Global admin state ───────────────────────────────────────────────────────
var adminUnlocked = sessionStorage.getItem('ganttAdminUnlocked') === '1';

function _applyNavbarAdminState() {{
  var btn = document.getElementById('navbarAdminBtn');
  if (!btn) return;
  if (adminUnlocked) {{
    btn.innerHTML = '<i class="bi bi-shield-fill-check me-1"></i>Admin';
    btn.style.cssText = 'border:1px solid #16a34a;color:#15803d;background:#f0fdf4;font-size:.78rem;padding:.25rem .6rem';
    btn.onclick = ganttAdminLogout;
  }} else {{
    btn.innerHTML = '<i class="bi bi-shield-lock me-1"></i>Admin';
    btn.style.cssText = 'border:1px solid #d1d5db;color:#6b7280;background:#fff;font-size:.78rem;padding:.25rem .6rem';
    btn.onclick = ganttAdminLogin;
  }}
}}

function ganttAdminLogin() {{
  document.getElementById('adminPwInput').value = '';
  document.getElementById('adminPwError').style.display = 'none';
  bootstrap.Modal.getOrCreateInstance(document.getElementById('adminLoginModal')).show();
  setTimeout(function() {{ document.getElementById('adminPwInput').focus(); }}, 400);
}}
function ganttAdminSubmit() {{
  var pw = document.getElementById('adminPwInput').value;
  if (pw !== 'password') {{
    document.getElementById('adminPwError').style.display = '';
    return;
  }}
  adminUnlocked = true;
  sessionStorage.setItem('ganttAdminUnlocked', '1');
  bootstrap.Modal.getOrCreateInstance(document.getElementById('adminLoginModal')).hide();
  _applyNavbarAdminState();
  _applyAdminGates();
}}
function ganttAdminLogout() {{
  adminUnlocked = false;
  sessionStorage.removeItem('ganttAdminUnlocked');
  if (typeof editActive !== 'undefined' && editActive && typeof ganttDeactivateEdit === 'function') {{
    ganttDeactivateEdit();
  }}
  _applyNavbarAdminState();
  _applyAdminGates();
}}

// ── Admin-gate elements ──────────────────────────────────────────────────────
// Elements with data-admin-gate="show"  → visible only for admins
// Elements with data-admin-gate="hide"  → hidden only for admins  (unused currently)
function _applyAdminGates() {{
  document.querySelectorAll('[data-admin-gate="show"]').forEach(function(el) {{
    el.style.display = adminUnlocked ? '' : 'none';
  }});
}}

// ── Tab memory: remember last URL per top-level tab ─────────────────────────
(function() {{
  _applyNavbarAdminState();
  _applyAdminGates();

  // Record current URL for whichever tab is active
  var activeTab = '{active}';
  if (activeTab) {{
    sessionStorage.setItem('tabLast_' + activeTab, location.href);
  }}

  // Intercept top-nav clicks: if we have a saved URL for the target tab, go there
  document.querySelectorAll('.df-topnav-link').forEach(function(a) {{
    a.addEventListener('click', function(e) {{
      var tab = a.dataset.tab;
      if (tab === activeTab) return;   // already here, let default href handle it
      var saved = sessionStorage.getItem('tabLast_' + tab);
      if (saved) {{
        e.preventDefault();
        location.href = saved;
      }}
    }});
  }});
}})();
</script>
</body></html>"""


# ── Base Handler ───────────────────────────────────────────────────────────────

class Base(tornado.web.RequestHandler):
    def sess(self):
        return _get_or_create(self)

    def emit(self, body: str, title: str = "Package Reliability", active: str = "",
             project: dict = None, active_sub: str = ""):
        _, s = self.sess()
        pt = s.get("part_type", "ttv")
        self.finish(_page(active, pt, body, title, project=project, active_sub=active_sub))


# ── / → redirect ──────────────────────────────────────────────────────────────

class IndexHandler(Base):
    def get(self):
        self.redirect("/lookup")


# ── /part-type ────────────────────────────────────────────────────────────────

class PartTypeHandler(Base):
    def get(self):
        _, s = self.sess()
        pt = s.get("part_type", "ttv")
        nxt = self.get_argument("next", "/lookup")

        def sel(val):
            return "checked" if pt == val else ""
        def sel_style(val):
            if pt == val:
                return "border-left: 3px solid var(--df-accent); background: #fdf6f4;"
            return "border-left: 3px solid transparent;"

        body = f"""
        <div class="row justify-content-center">
          <div class="col-md-6 col-lg-5">
            <div class="card">
              <div class="card-df">Sample Part Type</div>
              <div class="card-body p-4">
                <p class="mb-4" style="font-size:.85rem;color:var(--df-grey)">Select the sample type for this session. Controls which qualification tests are shown throughout the app.</p>
                <form method="post">
                  <input type="hidden" name="next" value="{nxt}">
                  <div class="p-3 border mb-3" style="{sel_style('active')}">
                    <div class="form-check">
                      <input class="form-check-input" type="radio" name="part_type" id="pt_a" value="active" {sel('active')}>
                      <label class="form-check-label ms-1" for="pt_a">
                        <div style="font-size:.88rem;font-weight:500">Active Device</div>
                        <div style="font-size:.8rem;color:var(--df-grey);margin-top:.25rem">Full qualification suite — all JEDEC tests, including HTOL, ELFR, THB, ESD CDM/HBM, Latch-Up</div>
                      </label>
                    </div>
                  </div>
                  <div class="p-3 border mb-3" style="{sel_style('ttv')}">
                    <div class="form-check">
                      <input class="form-check-input" type="radio" name="part_type" id="pt_t" value="ttv" {sel('ttv')}>
                      <label class="form-check-label ms-1" for="pt_t">
                        <div style="font-size:.88rem;font-weight:500">Thermal Test Vehicle (TTV) — Inactive</div>
                        <div style="font-size:.8rem;color:var(--df-grey);margin-top:.25rem">Mechanical &amp; thermal tests only: uHAST, TC, T-Shock, M-Shock, Vibration, Power Cycling, HTS, Shadow Moiré</div>
                      </label>
                    </div>
                  </div>
                  <div class="p-3 border mb-4" style="{sel_style('die')}">
                    <div class="form-check">
                      <input class="form-check-input" type="radio" name="part_type" id="pt_d" value="die" {sel('die')}>
                      <label class="form-check-label ms-1" for="pt_d">
                        <div style="font-size:.88rem;font-weight:500">Die</div>
                        <div style="font-size:.8rem;color:var(--df-grey);margin-top:.25rem">TTV tests excluding Power Cycling and Shadow Moiré: uHAST, TC, T-Shock, M-Shock, Vibration, PTC, HTS</div>
                      </label>
                    </div>
                  </div>
                  <button type="submit" class="btn btn-primary w-100 py-2">
                    Set Part Type &amp; Continue &rarr;
                  </button>
                </form>
              </div>
            </div>
          </div>
        </div>"""
        self.emit(body, "Part Type")

    def post(self):
        _, s = self.sess()
        pt = self.get_argument("part_type", "active")
        if pt in ("active", "ttv", "die"):
            s["part_type"] = pt
        self.redirect(self.get_argument("next", "/lookup"))


# ── /lookup ───────────────────────────────────────────────────────────────────

class LookupHandler(Base):
    def get(self):
        _, s = self.sess()
        tests = applicable_tests(s.get("part_type", "ttv"))

        import json as _json
        tc_js_data = _json.dumps(
            {ltr: {"tmin": v["tmin"], "tmax": v["tmax"],
                   "cycles": v["cycles"],
                   "soak": v["soak"],
                   "soakTimes": {str(m): TC_SOAK_TIMES[m] for m in v["soak"]},
                   "t4": v["t4"]}
             for ltr, v in TC_CONDITIONS.items()}
        )
        tshock_js_data = _json.dumps(TSHOCK_CONDITIONS)
        uhast_js_data   = _json.dumps(UHAST_CONDITIONS)
        thb_js_data     = _json.dumps(THB_CONDITIONS)
        mshock_js_data  = _json.dumps(MSHOCK_CONDITIONS)
        vib_sin_js_data = _json.dumps(VIB_SIN_CONDITIONS)
        vib_ran_js_data = _json.dumps(VIB_RAN_CONDITIONS)
        pc_js_data      = _json.dumps(PC_CONDITIONS)
        pc_method_js    = _json.dumps([{"label": m[0], "desc": m[1]} for m in PC_METHODS])
        ptc_js_data     = _json.dumps(PTC_CONDITIONS)
        hts_js_data     = _json.dumps(HTS_CONDITIONS)

        rows = ""
        for key, t in tests.items():
            badges = ""
            if t["destructive"]:
                badges += '<span class="badge bg-warning text-dark ms-2" style="font-size:.7rem">Destructive</span>'
            else:
                badges += '<span class="badge bg-success bg-opacity-75 ms-2" style="font-size:.7rem">Non-Destructive</span>'
            if t["active_devices"]:
                badges += '<span class="badge bg-info text-dark ms-2" style="font-size:.7rem">Active Device</span>'
            # "Additional Reli Test Prohibited" shown only in the expanded detail panel, not as a header badge
            notes_row = f'<tr><th class="fw-normal text-muted pe-3">Notes</th><td>{t["notes"]}</td></tr>' if t["notes"] else ""
            spec_docs = SPEC_URLS.get(key, [])
            # Header badge links (small, shown in accordion button)
            spec_link = " ".join(
                f'<a href="{url}" target="_blank" rel="noopener" '
                f'class="text-muted small ms-2" title="View {lbl}">'
                f'<i class="bi bi-file-earmark-text"></i> {lbl}</a>'
                for lbl, url in spec_docs
            )
            # Standard row: individual clickable labels for each available PDF
            if spec_docs:
                std_link = " / ".join(
                    f'<a href="{url}" target="_blank" rel="noopener" '
                    f'class="text-decoration-none">{lbl}</a>'
                    for lbl, url in spec_docs
                )
                # Append any listed standard names not covered by a link
                # (so unlinkable docs like JESD85 still appear as plain text)
                linked_labels = {lbl for lbl, _ in spec_docs}
                raw_standards = [s.strip() for s in t["standard"].replace(" /", "/").split("/")]
                missing = [s for s in raw_standards if not any(s.startswith(ll) or ll.startswith(s.split("-")[0]) for ll in linked_labels)]
                if missing:
                    std_link += " / " + " / ".join(missing)
            else:
                std_link = t["standard"]

            # TC and T-Shock get interactive condition pickers; all others use static text
            if key == "tc":
                cond_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "H" else ""}>'
                    f'Condition {ltr} &nbsp;({v["tmin"]:+d}°C / {v["tmax"]:+d}°C)'
                    f'</option>'
                    for ltr, v in TC_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="tc-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:260px" onchange="updateTcCond(this.value)">
                    {cond_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="tc-range" class="fw-semibold">+−55°C to +150°C</span>
                    &nbsp;<span class="text-muted">(Condition <span id="tc-ltr">H</span>)</span><br>
                    <span class="text-muted">Soak modes:</span> <span id="tc-soak">1 &amp; 2</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Min soak time:</span> <span id="tc-soaktime">1 or 5 min</span><br>
                    <span id="tc-cycles-wrap">
                      <span class="text-muted">Cycles/hr (Table 3):</span> <span id="tc-cycles">2</span>
                    </span>
                    <span id="tc-t4-wrap" style="display:none">
                      <span class="text-muted">Cycles/hr (Table 4 — solder interconnect):</span>
                      <span id="tc-t4-detail"></span>
                    </span>
                  </div>
                </td>"""
            elif key == "tshock":
                ts_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "C" else ""}>'
                    f'Condition {ltr} &nbsp;({v["cold"]:+d}°C / {v["hot"]:+d}°C)'
                    f'</option>'
                    for ltr, v in TSHOCK_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="ts-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:260px" onchange="updateTsCond(this.value)">
                    {ts_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="ts-range" class="fw-semibold">−55°C to +125°C</span>
                    &nbsp;<span class="text-muted">(Condition <span id="ts-ltr">C</span>)</span><br>
                    <span class="text-muted">Step 1 (hot):</span> <span id="ts-hot">+125°C</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Step 2 (cold):</span> <span id="ts-cold">−55°C</span><br>
                    <span class="text-muted">Fluid S1:</span> <span id="ts-fluid1">Perfluorocarbon</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Fluid S2:</span> <span id="ts-fluid2">Perfluorocarbon</span>
                  </div>
                </td>"""
            elif key == "vib":
                sin_opts = "".join(
                    f'<option value="{n}"{"  selected" if n == "1" else ""}>'
                    f'Condition {n} &nbsp;({v["accel_g"]}g / {v["fmin_hz"]}–{v["fmax_hz"]} Hz)'
                    f'</option>'
                    for n, v in VIB_SIN_CONDITIONS.items()
                )
                ran_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "A" else ""}>'
                    f'Condition {ltr} &nbsp;({v["rms_g"]}g RMS)'
                    f'</option>'
                    for ltr, v in VIB_RAN_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <div class="btn-group btn-group-sm mb-2" role="group">
                    <input type="radio" class="btn-check" name="vib-type" id="vib-sin-radio" value="sin" checked
                           onchange="switchVibType('sin')">
                    <label class="btn btn-outline-secondary" for="vib-sin-radio">Sinusoidal</label>
                    <input type="radio" class="btn-check" name="vib-type" id="vib-ran-radio" value="ran"
                           onchange="switchVibType('ran')">
                    <label class="btn btn-outline-secondary" for="vib-ran-radio">Random</label>
                  </div>

                  <div id="vib-sin-panel">
                    <select id="vib-sin-sel" class="form-select form-select-sm mb-2"
                            style="max-width:310px" onchange="updateVibSin(this.value)">
                      {sin_opts}
                    </select>
                    <div style="font-size:.82rem;line-height:1.7">
                      <span id="vs-label" class="fw-semibold">20g peak, 20–2000 Hz</span>
                      &nbsp;<span class="text-muted">(Condition <span id="vs-num">1</span>)</span><br>
                      <span class="text-muted">Peak acceleration:</span> <span id="vs-accel">20 g</span>
                      &ensp;|&ensp;
                      <span class="text-muted">Frequency range:</span> <span id="vs-freq">20–2000 Hz</span><br>
                      <span class="text-muted">Displacement pk-pk:</span> <span id="vs-disp">0.060 in / 1.5 mm</span>
                      &ensp;|&ensp;
                      <span class="text-muted">Crossover:</span> <span id="vs-xover">80 Hz</span>
                    </div>
                  </div>

                  <div id="vib-ran-panel" style="display:none">
                    <select id="vib-ran-sel" class="form-select form-select-sm mb-2"
                            style="max-width:310px" onchange="updateVibRan(this.value)">
                      {ran_opts}
                    </select>
                    <div style="font-size:.82rem;line-height:1.7">
                      <span id="vr-label" class="fw-semibold">6.27g RMS</span>
                      &nbsp;<span class="text-muted">(Condition <span id="vr-ltr">A</span>)</span><br>
                      <span class="text-muted">RMS acceleration:</span> <span id="vr-accel">6.27 g</span>
                      &ensp;|&ensp;
                      <span class="text-muted">RMS velocity:</span> <span id="vr-vel">29.0 in/sec</span><br>
                      <span class="text-muted">RMS displacement:</span> <span id="vr-disp">0.926 in</span>
                      &ensp;|&ensp;
                      <span class="text-muted">6&times;RMS (3&sigma; pk-pk):</span> <span id="vr-sigma">5.55 in</span>
                    </div>
                  </div>
                </td>"""
            elif key == "mshock":
                ms_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "B" else ""}>'
                    f'Service Condition {ltr} &nbsp;({v["accel_g"]}g / {v["pulse_ms"]} ms)'
                    f'</option>'
                    for ltr, v in MSHOCK_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="ms-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:310px" onchange="updateMsCond(this.value)">
                    {ms_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="ms-label" class="fw-semibold">1500g peak, 0.5 ms half-sine</span>
                    &nbsp;<span class="text-muted">(Service Condition <span id="ms-ltr">B</span>)</span><br>
                    <span class="text-muted">Acceleration peak:</span> <span id="ms-accel">1500 g</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Pulse duration:</span> <span id="ms-pulse">0.5 ms</span><br>
                    <span class="text-muted">Velocity change:</span> <span id="ms-vel">468 cm/s &nbsp;(184 in/s)</span><br>
                    <span class="text-muted">Equiv. drop height:</span> <span id="ms-drop">112 cm &nbsp;(44 in)</span>
                  </div>
                </td>"""
            elif key == "uhast":
                uh_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "A" else ""}>'
                    f'Condition {ltr} &nbsp;({v["temp_db"]}°C / {v["rh"]}% RH)'
                    f'</option>'
                    for ltr, v in UHAST_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="uh-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:280px" onchange="updateUhCond(this.value)">
                    {uh_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="uh-label" class="fw-semibold">130°C / 85% RH</span>
                    &nbsp;<span class="text-muted">(Condition <span id="uh-ltr">A</span>)</span><br>
                    <span class="text-muted">Temp (dry-bulb):</span> <span id="uh-tdb">130 ± 2°C</span>
                    &ensp;|&ensp;
                    <span class="text-muted">RH:</span> <span id="uh-rh">85 ± 5%</span><br>
                    <span class="text-muted">Temp (wet-bulb):</span> <span id="uh-twb">124.7°C</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Vapor pressure:</span> <span id="uh-vp">230 kPa (33.3 psia)</span><br>
                    <span class="text-muted">Duration:</span> <span id="uh-dur">96 hours (−0, +2)</span>
                  </div>
                </td>"""
            elif key == "hts":
                hts_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "B" else ""}>'
                    f'Condition {ltr} &nbsp;(+{v["temp_c"]}°C)'
                    f'</option>'
                    for ltr, v in HTS_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="hts-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:240px" onchange="updateHtsCond(this.value)">
                    {hts_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="hts-label" class="fw-semibold">+150°C storage</span>
                    &nbsp;<span class="text-muted">(Condition <span id="hts-ltr">B</span>)</span><br>
                    <span class="text-muted">Temperature:</span>
                    <span id="hts-temp">+150°C (−0 / +10°C)</span><br>
                    <span class="text-muted">Duration (JESD47 default):</span>
                    <span>1000 hours</span>
                    &ensp;<span class="text-muted small">(other durations acceptable per product requirements)</span>
                  </div>
                </td>"""
            elif key == "thb":
                thb_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "A" else ""}>'
                    f'Condition {ltr} &nbsp;({v["temp_db"]}°C / {v["rh"]}% RH)'
                    f'</option>'
                    for ltr, v in THB_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="thb-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:280px" onchange="updateThbCond(this.value)">
                    {thb_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="thb-label" class="fw-semibold">130°C / 85% RH</span>
                    &nbsp;<span class="text-muted">(Condition <span id="thb-ltr">A</span>)</span><br>
                    <span class="text-muted">Temp (dry-bulb):</span> <span id="thb-tdb">130 ± 2°C</span>
                    &ensp;|&ensp;
                    <span class="text-muted">RH:</span> <span id="thb-rh">85 ± 5%</span><br>
                    <span class="text-muted">Temp (wet-bulb):</span> <span id="thb-twb">124.7°C</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Vapor pressure:</span> <span id="thb-vp">230 kPa (33.3 psia)</span><br>
                    <span class="text-muted">Duration:</span> <span id="thb-dur">96 hours (−0, +2)</span><br>
                    <span class="text-muted">Bias:</span> V<sub>dd</sub> applied at max operating voltage
                  </div>
                </td>"""
            elif key == "pc":
                pc_cond_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "B" else ""}>'
                    f'Condition {ltr} &nbsp;(Tmin {v["tmin"]}°C / Tmax {v["tmax"]}°C, ΔT {v["delta_t"]}°C)'
                    f'</option>'
                    for ltr, v in PC_CONDITIONS.items()
                )
                pc_method_opts = "".join(
                    f'<option value="{i}">{m[0]}</option>'
                    for i, m in enumerate(PC_METHODS)
                )
                condition_cell = f"""<td>
                  <div class="row g-2 mb-2">
                    <div class="col-auto">
                      <select id="pc-cond-sel" class="form-select form-select-sm"
                              style="max-width:340px" onchange="updatePcCond(this.value)">
                        {pc_cond_opts}
                      </select>
                    </div>
                  </div>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="pc-label" class="fw-semibold">Condition B — ΔT 100°C</span><br>
                    <span class="text-muted">Tcycle(min):</span> <span id="pc-tmin">25°C (+5, −5)</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Tcycle(max):</span> <span id="pc-tmax">+125°C (+5, −5)</span>
                    &ensp;|&ensp;
                    <span class="text-muted">ΔT:</span> <span id="pc-dt">100°C</span><br>
                    <span class="text-muted">Cycle rate:</span> <span>2–6 cycles/hr (typical)</span>
                  </div>
                  <div class="mt-2">
                    <label class="text-muted" style="font-size:.8rem">Test Method</label>
                    <select id="pc-method-sel" class="form-select form-select-sm mt-1"
                            style="max-width:340px" onchange="updatePcMethod(this.value)">
                      {pc_method_opts}
                    </select>
                    <div id="pc-method-desc" class="text-muted mt-1"
                         style="font-size:.8rem;font-style:italic">
                      Best for functional devices &amp; tightly controlled heater resistance; allows Tj variation modeling
                    </div>
                  </div>
                </td>"""
            elif key == "ptc":
                ptc_opts = "".join(
                    f'<option value="{ltr}"{"  selected" if ltr == "A" else ""}>'
                    f'Condition {ltr} &nbsp;({v["tmin"]:+d}°C to {v["tmax"]:+d}°C, {v["trans_min"]} min transition)'
                    f'</option>'
                    for ltr, v in PTC_CONDITIONS.items()
                )
                condition_cell = f"""<td>
                  <select id="ptc-cond-sel" class="form-select form-select-sm mb-2"
                          style="max-width:340px" onchange="updatePtcCond(this.value)">
                    {ptc_opts}
                  </select>
                  <div style="font-size:.82rem;line-height:1.7">
                    <span id="ptc-label" class="fw-semibold">−40°C to +85°C</span>
                    &nbsp;<span class="text-muted">(Condition <span id="ptc-ltr">A</span>)</span><br>
                    <span class="text-muted">T<sub>min</sub>:</span>
                    <span id="ptc-tmin">−40°C <span class="text-muted">(+0, −10)</span></span>
                    &ensp;|&ensp;
                    <span class="text-muted">T<sub>max</sub>:</span>
                    <span id="ptc-tmax">+85°C <span class="text-muted">(+10, −0)</span></span><br>
                    <span class="text-muted">Transition time (max):</span>
                    <span id="ptc-trans">20 minutes</span>
                    &ensp;|&ensp;
                    <span class="text-muted">Dwell time (min):</span>
                    <span id="ptc-dwell">10 minutes</span>
                  </div>
                </td>"""
            else:
                condition_cell = f"<td>{t['condition']}</td>"

            rows += f"""
            <div class="accordion-item">
              <h2 class="accordion-header">
                <button class="accordion-button collapsed py-3" type="button"
                        data-bs-toggle="collapse" data-bs-target="#a_{key}" aria-expanded="false">
                  <strong class="me-2">{t['name']}</strong>
                  <span class="text-muted" style="font-size:.83rem">{t['standard']}</span>
                  {badges}
                  {spec_link}
                </button>
              </h2>
              <div id="a_{key}" class="accordion-collapse collapse">
                <div class="accordion-body pt-2 pb-3">
                  <div class="text-muted small mb-3" style="font-style:italic">{t['full_name']}</div>
                  <div class="row">
                    <div class="col-md-6">
                      <table class="table table-sm table-borderless mb-0">
                        <tr><th class="fw-normal text-muted pe-3" style="width:130px;white-space:nowrap">Precursor</th>
                            <td>PC ({PRECOND['full_name']}; {PRECOND['standard']})</td></tr>
                        <tr><th class="fw-normal text-muted pe-3">Standard</th><td>{std_link}</td></tr>
                        <tr><th class="fw-normal text-muted pe-3">Condition</th>{condition_cell}</tr>
                        <tr><th class="fw-normal text-muted pe-3">Duration</th><td>{t['duration']}</td></tr>
                        <tr><th class="fw-normal text-muted pe-3">Additional Reli Test Prohibited</th><td>{'Yes' if t['destructive'] else 'No'}</td></tr>
                      </table>
                    </div>
                    <div class="col-md-6">
                      <table class="table table-sm table-borderless mb-0">
                        <tr><th class="fw-normal text-muted pe-3" style="width:130px;white-space:nowrap">Pre-test</th><td>{t['pre_testing']}</td></tr>
                        <tr><th class="fw-normal text-muted pe-3">Post-test</th><td>{t['post_testing']}</td></tr>
                        <tr><th class="fw-normal text-muted pe-3">Pass Criteria</th><td>{t['pass_criteria']}</td></tr>
                        {notes_row}
                      </table>
                    </div>
                  </div>
                </div>
              </div>
            </div>"""

        # Build deduplicated spec document list with associated test names
        _test_name_map = {k: t["full_name"] for k, t in TESTS.items()}
        _test_name_map["jesd47"] = "Stress-Test-Driven Qualification of Integrated Circuits"
        _test_name_map["precond"] = "Preconditioning (MSL)"
        _spec_map: dict = {}
        _spec_order: list = []
        for _key, _docs in SPEC_URLS.items():
            _tname = _test_name_map.get(_key, _key)
            for _lbl, _url in _docs:
                if _lbl not in _spec_map:
                    _spec_map[_lbl] = {"url": _url, "tests": []}
                    _spec_order.append(_lbl)
                if _tname not in _spec_map[_lbl]["tests"]:
                    _spec_map[_lbl]["tests"].append(_tname)
        spec_list_rows = ""
        for _lbl in sorted(_spec_order):
            _info = _spec_map[_lbl]
            _tests_str = " / ".join(_info["tests"])
            spec_list_rows += f"""
            <div class="d-flex align-items-center gap-3 py-2 border-bottom">
              <i class="bi bi-file-earmark-pdf text-danger flex-shrink-0 fs-5"></i>
              <div class="flex-grow-1">
                <span class="fw-semibold">{_lbl}</span>
                <span class="text-muted ms-2" style="font-size:.82rem">&mdash; {_tests_str}</span>
              </div>
              <a href="{_info['url']}" download="{_lbl}.pdf"
                 class="btn btn-sm btn-outline-secondary flex-shrink-0">
                <i class="bi bi-download me-1"></i>{_lbl}
              </a>
            </div>"""

        body = f"""
        <div class="d-flex align-items-center mb-3">
          <h4 class="mb-0" style="font-weight:300">Test Condition Lookup</h4>
        </div>
        <div class="precond-bar">
          <strong>Precursor for all tests &mdash;</strong>
          PC ({PRECOND['full_name']};
          {" / ".join(f'<a href="{url}" target="_blank" style="color:inherit;text-decoration:underline dotted">{lbl}</a>' for lbl, url in SPEC_URLS.get("precond", []))})
          &nbsp;&middot;&nbsp; {PRECOND['condition']}
          &nbsp;&middot;&nbsp; {PRECOND['duration']}
          &nbsp;&middot;&nbsp; Pass: {PRECOND['pass_criteria']}
        </div>
        <div class="alert alert-info d-flex align-items-start gap-2 mb-3 mt-2 py-2 px-3" role="alert" style="font-size:.85rem">
          <i class="bi bi-info-circle-fill mt-1 flex-shrink-0"></i>
          <span>Test conditions shown reflect one recommended test case for thermal test vehicle.
          Other conditions may be selected from JEDEC standard based on device type, use environment, and applicable specifications.</span>
        </div>
        <div class="accordion shadow-sm" id="accTest">{rows}</div>

        <div class="card mt-4 mb-2 shadow-sm border-secondary">
          <div class="card-header bg-secondary bg-opacity-10 py-2 px-3 d-flex align-items-center gap-2">
            <i class="bi bi-tools text-secondary"></i>
            <strong style="font-size:.9rem">Engineering Analysis</strong>
          </div>
          <div class="card-body px-3 py-2" style="font-size:.84rem; line-height:1.6">
            <p class="mb-2 text-muted">The following techniques are used for root-cause investigation and failure analysis. They are not qualification tests — samples are consumed and <strong>cannot be reused for qualification</strong>.</p>
            <div class="accordion accordion-flush border rounded" id="eng-analysis-acc">
              <div class="accordion-item">
                <h2 class="accordion-header">
                  <button class="accordion-button collapsed py-2" type="button" data-bs-toggle="collapse" data-bs-target="#ea-pull" style="font-size:.84rem">
                    <strong class="me-2">Pull-Test</strong>
                    <span class="text-muted" style="font-size:.83rem">Bond wire / die-attach strength</span>
                    <span class="badge bg-danger bg-opacity-75 ms-2" style="font-size:.7rem">Destructive</span>
                  </button>
                </h2>
                <div id="ea-pull" class="accordion-collapse collapse" data-bs-parent="#eng-analysis-acc">
                  <div class="accordion-body py-2 px-3" style="font-size:.83rem; line-height:1.7">
                    <p class="mb-2">A mechanical pull or shear force is applied to a bond wire or die-attach joint until failure. The force at failure (gram-force or millinewton) and the failure mode (e.g., wire heel break, bond lift, intermetallic fracture, die-attach cohesive failure) are recorded.</p>
                    <ul class="mb-2 ps-3">
                      <li><strong>Wire bond pull:</strong> Per MIL-STD-883 Method 2011 / ASTM F459. A hook is placed under the wire mid-span and pulled perpendicular to the substrate.</li>
                      <li><strong>Ball/wedge shear:</strong> Per ASTM F1269 / JEDEC JESD22-B116. A shear tool pushes laterally against the ball or wedge bond.</li>
                      <li><strong>Die shear:</strong> A shear tool pushes against the die sidewall to measure die-attach adhesion strength.</li>
                    </ul>
                    <div class="alert alert-danger py-1 px-2 mb-0 d-flex gap-2 align-items-start" style="font-size:.8rem" role="alert">
                      <i class="bi bi-exclamation-triangle-fill text-danger flex-shrink-0 mt-1"></i>
                      <span>The sample is physically destroyed in the process. Tested samples may not be returned to the qualification lot.</span>
                    </div>
                  </div>
                </div>
              </div>
              <div class="accordion-item">
                <h2 class="accordion-header">
                  <button class="accordion-button collapsed py-2" type="button" data-bs-toggle="collapse" data-bs-target="#ea-xsem" style="font-size:.84rem">
                    <strong class="me-2">X-SEM</strong>
                    <span class="text-muted" style="font-size:.83rem">Cross-sectional scanning electron microscopy</span>
                    <span class="badge bg-danger bg-opacity-75 ms-2" style="font-size:.7rem">Destructive</span>
                  </button>
                </h2>
                <div id="ea-xsem" class="accordion-collapse collapse" data-bs-parent="#eng-analysis-acc">
                  <div class="accordion-body py-2 px-3" style="font-size:.83rem; line-height:1.7">
                    <p class="mb-2">The package is mechanically or focused-ion-beam (FIB) sectioned through the region of interest, then imaged under a scanning electron microscope (SEM). Energy-dispersive X-ray spectroscopy (EDS/EDX) may be performed simultaneously to map elemental composition.</p>
                    <ul class="mb-2 ps-3">
                      <li>Reveals internal voids, cracks, delamination, intermetallic growth, and solder joint morphology not visible by CSAM.</li>
                      <li>FIB-SEM allows site-specific cross-sections to sub-micron precision; mechanical polishing is used for broader area surveys.</li>
                      <li>Common targets: bond interface, SCD-to-die bond line, solder joints, via fill integrity.</li>
                    </ul>
                    <div class="alert alert-danger py-1 px-2 mb-0 d-flex gap-2 align-items-start" style="font-size:.8rem" role="alert">
                      <i class="bi bi-exclamation-triangle-fill text-danger flex-shrink-0 mt-1"></i>
                      <span>The sample is physically destroyed in the process. Tested samples may not be returned to the qualification lot.</span>
                    </div>
                  </div>
                </div>
              </div>
            </div>
          </div>
        </div>

        <div class="card mt-2 mb-2 shadow-sm border-warning">
          <div class="card-header bg-warning bg-opacity-10 py-2 px-3 d-flex align-items-center gap-2">
            <i class="bi bi-exclamation-triangle-fill text-warning"></i>
            <strong style="font-size:.9rem">In Case of Part Failure</strong>
          </div>
          <div class="card-body px-3 py-2" style="font-size:.84rem; line-height:1.6">
            <p class="mb-2">The following guidance is drawn from <strong>JESD47I</strong>:</p>
            <ul class="mb-2 ps-3">
              <li><strong>Discounting failures (§3.6):</strong> A failure may be discounted from the sample count if it can be documented that the root cause is unrelated to the test conditions (e.g., handling damage, significant delamination, SCD/Si cracking, pre-existing defects). Evidence of the unrelated cause is required.</li>
              <li><strong>Sample reusability (§3.5):</strong> Devices used in <em>nondestructive</em> tests may be reused in subsequent stress tests. Devices subjected to <em>destructive</em> analysis may not be reused for qualification — they are limited to engineering analysis only.</li>
              <li><strong>Failure analysis &amp; requalification (§4.2.3):</strong> Failed devices should be analyzed for root cause; only a <em>representative sample</em> needs to be analyzed, not every failed part. Successful requalification requires demonstrating corrective and preventive actions. Only the tests affected by the change that caused the failure need to be repeated — a full requalification from scratch is not required.</li>
            </ul>
            <p class="mb-0 text-muted" style="font-size:.8rem">Refer to JESD47I for full normative requirements. The same §3.8 sample size formula applies to any requalification run.</p>
          </div>
        </div>

        <div class="card mt-2 mb-2 shadow-sm border-info">
          <div class="card-header bg-info bg-opacity-10 py-2 px-3 d-flex align-items-center gap-2">
            <i class="bi bi-arrow-repeat text-info"></i>
            <strong style="font-size:.9rem">Reasons for Requalification</strong>
          </div>
          <div class="card-body px-3 py-2" style="font-size:.84rem; line-height:1.6">
            <p class="mb-2">Any of the following changes trigger a requalification requirement per JESD47I. Expand the applicable device category below.</p>

            <!-- TTV accordion -->
            <div class="accordion accordion-flush border rounded mb-2" id="requal-acc">
              <div class="accordion-item">
                <h2 class="accordion-header">
                  <button class="accordion-button collapsed py-2" type="button" data-bs-toggle="collapse" data-bs-target="#requal-ttv" style="font-size:.84rem; font-weight:600">
                    <i class="bi bi-cpu me-2 text-secondary"></i>Thermal Test Vehicle (TTV)
                  </button>
                </h2>
                <div id="requal-ttv" class="accordion-collapse collapse" data-bs-parent="#requal-acc">
                  <div class="accordion-body py-2 px-3">
                    <ol class="mb-0 ps-3" style="font-size:.83rem; line-height:1.7">
                      <li>New TTV (PCB, layout, solder)</li>
                      <li><strong>Metallization:</strong> New materials or a significant change in composition</li>
                      <li><strong>Bonding:</strong> Process and/or technique</li>
                      <li>Die thickness</li>
                      <li>SCD thickness</li>
                    </ol>
                  </div>
                </div>
              </div>

              <!-- Active device accordion -->
              <div class="accordion-item">
                <h2 class="accordion-header">
                  <button class="accordion-button collapsed py-2" type="button" data-bs-toggle="collapse" data-bs-target="#requal-active" style="font-size:.84rem; font-weight:600">
                    <i class="bi bi-diagram-3 me-2 text-secondary"></i>Active Device
                  </button>
                </h2>
                <div id="requal-active" class="accordion-collapse collapse" data-bs-parent="#requal-acc">
                  <div class="accordion-body py-2 px-3">
                    <ol class="mb-0 ps-3" style="font-size:.83rem; line-height:1.7">
                      <li><strong>Active Circuit Element:</strong> New type of circuit element or modification of transistors beyond original qualification or spec limits</li>
                      <li><strong>Major Circuit Elements:</strong> Addition of a major new circuit block to an existing circuit such as adding a Digital Signal Processor or embedded memory block to an existing product</li>
                      <li><strong>Wafer Diameter Change / Metallization:</strong> New materials or a significant change in composition</li>
                      <li><strong>Change In Minimum Feature Size:</strong> A reduction of greater than 20% shall be considered a new process</li>
                      <li><strong>Wafer Fab Process:</strong> Utilizing different process techniques at critical points (excluding wafer transport equipment)</li>
                      <li><strong>Diffusion/Dopant:</strong> New material or technique</li>
                      <li><strong>Polysilicon or other MOSFET gate material:</strong> Composition, design rules, process</li>
                      <li><strong>Lithography:</strong> Change in wavelength, method (air / immersion / e-beam), or etch technique</li>
                      <li><strong>Wafer Frontside Metallization:</strong> Composition, design rules, process and/or technique</li>
                      <li><strong>VIA:</strong> Composition, design rules, process and/or technique</li>
                      <li><strong>Passivation Overcoat:</strong> Either glass or organic material composition, design rules, process and/or technique</li>
                      <li><strong>Dielectric Materials:</strong> Composition, design rules, process and/or technique</li>
                      <li><strong>Low-K Dielectric:</strong> A dielectric material used for inter-metal isolation with a K value less than 3.2</li>
                      <li><strong>Wafer Backside Operation:</strong> Metal composition, design rules, process and/or technique</li>
                      <li><strong>New Wafer Manufacturing Line:</strong> Not already qualified for the fabrication process</li>
                      <li><strong>Assembly Process:</strong> Utilizing different process techniques at critical points</li>
                      <li><strong>Die Coating:</strong> Material, process, and/or technique</li>
                      <li><strong>Lead Frame:</strong> Base material, finish, and critical dimensions</li>
                      <li><strong>Bond Wire:</strong> Material, diameter</li>
                      <li><strong>Bonding / Die Preparation:</strong> Process and/or technique; separation and clean methods</li>
                      <li><strong>Die Attach:</strong> Material, process, and/or technique</li>
                      <li><strong>Encapsulation:</strong> Material, composition, process and/or technique</li>
                      <li><strong>Hermetic Package:</strong> Material, composition, seal material, process and/or technique</li>
                      <li><strong>Wafer Bumping Material:</strong> Process or technique (including flip chip assembly process)</li>
                      <li><strong>Package Dimension Change:</strong> Larger package body size or reduction in lead or solder ball pitch</li>
                      <li>Die Thickness</li>
                      <li>New Chip-Package Combination</li>
                    </ol>
                  </div>
                </div>
              </div>
            </div>

            <p class="mb-0 text-muted" style="font-size:.8rem">Source: JESD47I Table 1 &amp; §4.2. Only tests affected by the triggering change need to be repeated.</p>
          </div>
        </div>

        <div class="card mt-2 mb-2 shadow-sm">
          <div class="card-body py-2 px-3 d-flex align-items-center gap-3" style="font-size:.85rem">
            <i class="bi bi-file-earmark-pdf text-danger fs-5"></i>
            <div>
              <strong>JESD47I</strong> — Stress-Test-Driven Qualification of Integrated Circuits
              <span class="text-muted ms-2" style="font-size:.8rem">Generic qualification guidelines &amp; pass/fail criteria</span>
            </div>
            <a href="{SPEC_URLS['jesd47'][0][1]}" target="_blank" class="btn btn-sm btn-outline-secondary ms-auto">
              <i class="bi bi-download me-1"></i>Download PDF
            </a>
          </div>
        </div>

        <div class="accordion shadow-sm mt-2 mb-2" id="accSpecs">
          <div class="accordion-item">
            <h2 class="accordion-header">
              <button class="accordion-button collapsed py-2 px-3" type="button"
                      data-bs-toggle="collapse" data-bs-target="#collapseSpecs"
                      style="font-size:.9rem">
                <i class="bi bi-journals text-secondary me-2"></i>
                <strong>JEDEC Specification Documents</strong>
              </button>
            </h2>
            <div id="collapseSpecs" class="accordion-collapse collapse" data-bs-parent="#accSpecs">
              <div class="accordion-body px-3 py-1">
                {spec_list_rows}
              </div>
            </div>
          </div>
        </div>

        <script>
        const TC_DATA     = {tc_js_data};
        const TSHOCK_DATA = {tshock_js_data};
        const UHAST_DATA   = {uhast_js_data};
        const THB_DATA     = {thb_js_data};
        const MSHOCK_DATA   = {mshock_js_data};
        const VIB_SIN_DATA  = {vib_sin_js_data};
        const VIB_RAN_DATA  = {vib_ran_js_data};
        const PC_DATA       = {pc_js_data};
        const PC_METHODS    = {pc_method_js};
        const PTC_DATA      = {ptc_js_data};
        const HTS_DATA      = {hts_js_data};

        function fmtTemp(v) {{
          return (v >= 0 ? "+" : "\u2212") + Math.abs(v) + "\u00b0C";
        }}

        function updateTcCond(ltr) {{
          const c = TC_DATA[ltr];
          if (!c) return;
          document.getElementById("tc-ltr").textContent   = ltr;
          document.getElementById("tc-range").textContent = fmtTemp(c.tmin) + " to " + fmtTemp(c.tmax);

          // Soak modes label (e.g. "1, 2 & 3")
          const modes = c.soak;
          const modeStr = modes.length > 1
            ? modes.slice(0, -1).join(", ") + " & " + modes[modes.length - 1]
            : "" + modes[0];
          document.getElementById("tc-soak").textContent = modeStr;

          // Min soak times (e.g. "1, 5, or 10 min")
          const times = modes.map(m => c.soakTimes[String(m)]);
          const timeStr = times.length > 1
            ? times.slice(0, -1).join(", ") + " or " + times[times.length - 1] + " min"
            : times[0] + " min";
          document.getElementById("tc-soaktime").textContent = timeStr;

          // Cycles/hr — Table 4 if available, else Table 3
          const t4 = c.t4 || {{}};
          const t4Keys = Object.keys(t4);
          if (t4Keys.length > 0) {{
            // Show per-soak-mode rates from Table 4
            const detail = t4Keys.map(m => "Soak " + m + ": " + t4[m] + " cph").join(" \u2022 ");
            document.getElementById("tc-t4-detail").textContent = " " + detail;
            document.getElementById("tc-cycles-wrap").style.display = "none";
            document.getElementById("tc-t4-wrap").style.display    = "";
          }} else {{
            // Fall back to Table 3
            document.getElementById("tc-cycles").textContent        = c.cycles;
            document.getElementById("tc-cycles-wrap").style.display = "";
            document.getElementById("tc-t4-wrap").style.display     = "none";
          }}
        }}

        function updateTsCond(ltr) {{
          const c = TSHOCK_DATA[ltr];
          if (!c) return;
          document.getElementById("ts-ltr").textContent    = ltr;
          document.getElementById("ts-range").textContent  = fmtTemp(c.cold) + " to " + fmtTemp(c.hot);
          document.getElementById("ts-hot").textContent    = fmtTemp(c.hot);
          document.getElementById("ts-cold").textContent   = fmtTemp(c.cold);
          document.getElementById("ts-fluid1").textContent = c.fluid_s1;
          document.getElementById("ts-fluid2").textContent = c.fluid_s2;
        }}

        function updateHtsCond(ltr) {{
          const c = HTS_DATA[ltr];
          if (!c) return;
          document.getElementById("hts-ltr").textContent   = ltr;
          document.getElementById("hts-label").textContent = "+" + c.temp_c + "\u00b0C storage";
          document.getElementById("hts-temp").textContent  = "+" + c.temp_c + "\u00b0C (\u22120 / +10\u00b0C)";
        }}

        function updatePcCond(ltr) {{
          const c = PC_DATA[ltr];
          if (!c) return;
          document.getElementById("pc-label").textContent = "Condition " + ltr + " \u2014 \u0394T " + c.delta_t + "\u00b0C";
          document.getElementById("pc-tmin").textContent  = c.tmin + "\u00b0C (+5, \u22125)";
          document.getElementById("pc-tmax").textContent  = "+" + c.tmax + "\u00b0C (+5, \u22125)";
          document.getElementById("pc-dt").textContent    = c.delta_t + "\u00b0C";
        }}

        function updatePcMethod(idx) {{
          const m = PC_METHODS[parseInt(idx)];
          if (!m) return;
          document.getElementById("pc-method-desc").innerHTML = m.desc;
        }}

        function updatePtcCond(ltr) {{
          const c = PTC_DATA[ltr];
          if (!c) return;
          document.getElementById("ptc-ltr").textContent   = ltr;
          document.getElementById("ptc-label").textContent = c.tmin + "\u00b0C to +" + c.tmax + "\u00b0C";
          document.getElementById("ptc-tmin").innerHTML    = c.tmin + "\u00b0C <span class='text-muted'>" + c.tmin_tol + "</span>";
          document.getElementById("ptc-tmax").innerHTML    = "+" + c.tmax + "\u00b0C <span class='text-muted'>" + c.tmax_tol + "</span>";
          document.getElementById("ptc-trans").textContent = c.trans_min + " minutes";
          document.getElementById("ptc-dwell").textContent = c.dwell_min + " minutes";
        }}

        function switchVibType(type) {{
          document.getElementById("vib-sin-panel").style.display = type === "sin" ? "" : "none";
          document.getElementById("vib-ran-panel").style.display = type === "ran" ? "" : "none";
        }}

        function updateVibSin(num) {{
          const c = VIB_SIN_DATA[num];
          if (!c) return;
          document.getElementById("vs-num").textContent   = num;
          document.getElementById("vs-label").textContent = c.accel_g + "g peak, " + c.fmin_hz + "\u2013" + c.fmax_hz + " Hz";
          document.getElementById("vs-accel").textContent = c.accel_g + " g";
          document.getElementById("vs-freq").textContent  = c.fmin_hz + "\u2013" + c.fmax_hz + " Hz";
          document.getElementById("vs-disp").textContent  = c.disp_in + " in / " + c.disp_mm + " mm";
          document.getElementById("vs-xover").textContent = c.xover_hz + " Hz";
        }}

        function updateVibRan(ltr) {{
          const c = VIB_RAN_DATA[ltr];
          if (!c) return;
          document.getElementById("vr-ltr").textContent   = ltr;
          document.getElementById("vr-label").textContent = c.rms_g + "g RMS";
          document.getElementById("vr-accel").textContent = c.rms_g + " g";
          document.getElementById("vr-vel").textContent   = c.vel_ins + " in/sec";
          document.getElementById("vr-disp").textContent  = c.disp_in + " in";
          document.getElementById("vr-sigma").textContent = c.sigma_in + " in";
        }}

        function updateMsCond(ltr) {{
          const c = MSHOCK_DATA[ltr];
          if (!c) return;
          document.getElementById("ms-ltr").textContent   = ltr;
          document.getElementById("ms-label").textContent = c.accel_g + "g peak, " + c.pulse_ms + " ms half-sine";
          document.getElementById("ms-accel").textContent = c.accel_g + " g";
          document.getElementById("ms-pulse").textContent = c.pulse_ms + " ms";
          document.getElementById("ms-vel").textContent   = c.vel_cms + " cm/s \u00a0(" + c.vel_ins + " in/s)";
          document.getElementById("ms-drop").textContent  = c.drop_cm + " cm \u00a0(" + c.drop_in + " in)";
        }}

        function updateUhCond(ltr) {{
          const c = UHAST_DATA[ltr];
          if (!c) return;
          document.getElementById("uh-ltr").textContent  = ltr;
          document.getElementById("uh-label").textContent = c.temp_db + "\u00b0C / " + c.rh + "% RH";
          document.getElementById("uh-tdb").textContent  = c.temp_db + " \u00b1 2\u00b0C";
          document.getElementById("uh-rh").textContent   = c.rh + " \u00b1 5%";
          document.getElementById("uh-twb").textContent  = c.temp_wb + "\u00b0C";
          document.getElementById("uh-vp").textContent   = c.vp_kpa + " kPa (" + c.vp_psia + " psia)";
          document.getElementById("uh-dur").textContent  = c.duration;
        }}

        function updateThbCond(ltr) {{
          const c = THB_DATA[ltr];
          if (!c) return;
          document.getElementById("thb-ltr").textContent  = ltr;
          document.getElementById("thb-label").textContent = c.temp_db + "\u00b0C / " + c.rh + "% RH";
          document.getElementById("thb-tdb").textContent  = c.temp_db + " \u00b1 2\u00b0C";
          document.getElementById("thb-rh").textContent   = c.rh + " \u00b1 5%";
          document.getElementById("thb-twb").textContent  = c.temp_wb + "\u00b0C";
          document.getElementById("thb-vp").textContent   = c.vp_kpa + " kPa (" + c.vp_psia + " psia)";
          document.getElementById("thb-dur").textContent  = c.duration;
        }}
        </script>"""
        self.emit(body, "Test Lookup", "lookup")


# ── /sample-size ──────────────────────────────────────────────────────────────

class SampleSizeHandler(Base):
    def get(self):
        self._render()

    def post(self):
        try:
            k    = int(self.get_argument("failures", "0"))
            ltpd = float(self.get_argument("ltpd", "5"))
            if k < 0:              raise ValueError("Acceptance number C must be ≥ 0")
            if not (0.01 <= ltpd <= 99.99): raise ValueError("LTPD must be between 0.01% and 99.99%")
            n_jesd47 = min_sample_size_ltpd(ltpd, k)
            r_equiv  = 1.0 - ltpd / 100.0
            n_exact  = min_sample_size(r_equiv, 0.90, k)
            self._render(k=k, ltpd=ltpd, n_jesd47=n_jesd47, n_exact=n_exact)
        except Exception as e:
            self._render(error=str(e))

    def _render(self, k=0, ltpd=5, n_jesd47=None, n_exact=None, error=None):
        # Build Table A HTML — highlight nearest standard LTPD column and selected C row
        ltpd_cols = TABLE_A_LTPD   # [10, 7, 5, 3, 2, 1.5, 1]
        nearest_ltpd = min(ltpd_cols, key=lambda l: abs(l - ltpd)) if n_jesd47 is not None else None
        hdr_cells = "".join(
            f'<th class="{"table-primary fw-bold" if l == nearest_ltpd else ""}">'
            f'LTPD {l}%</th>'
            for l in ltpd_cols
        )
        tbl_rows = ""
        for c_val, row in TABLE_A.items():
            row_cls = 'class="table-primary"' if (n_jesd47 is not None and c_val == k) else ""
            cells = ""
            for ci, n_val in enumerate(row):
                is_sel = (n_jesd47 is not None and c_val == k and ltpd_cols[ci] == nearest_ltpd)
                td_cls = ' class="fw-bold text-primary"' if is_sel else ""
                cells += f"<td{td_cls}>{n_val}</td>"
            tbl_rows += f"<tr {row_cls}><td><strong>{c_val}</strong></td>{cells}</tr>"

        table_a_html = f"""
        <div class="card mt-4">
          <div class="card-df d-flex align-items-center">
            <h6 class="mb-0">Table A — JESD47I §3.8</h6>
            <span class="text-white-50 ms-2" style="font-size:.78rem">Sample Size for Maximum % Defective at 90% Confidence</span>
          </div>
          <div class="table-responsive">
            <table class="table table-sm table-bordered mb-0" style="font-size:.8rem">
              <thead class="tbl-header">
                <tr><th>Accept # (C)</th>{hdr_cells}</tr>
              </thead>
              <tbody>{tbl_rows}</tbody>
            </table>
          </div>
          <div class="card-footer text-muted py-1 px-3" style="font-size:.75rem">
            Highlighted cell = selected C &amp; LTPD.
            <a href="{SPEC_URLS['jesd47'][0][1]}" target="_blank" class="ms-2">
              <i class="bi bi-file-earmark-pdf me-1"></i>JESD47I PDF
            </a>
          </div>
        </div>"""

        result_html = ""
        if error:
            result_html = f'<div class="alert alert-danger"><i class="bi bi-exclamation-triangle me-2"></i>{error}</div>'
        elif n_jesd47 is not None:
            r_equiv = 1.0 - ltpd / 100.0
            result_html = f"""
            <div class="card mb-3">
              <div class="card-body text-center py-4">
                <div class="text-muted small mb-1">Minimum Sample Size (JESD47I)</div>
                <div class="stat-num" style="color:var(--df-navy)">{n_jesd47}</div>
                <div class="text-muted mt-2 small">
                  to demonstrate ≤<strong>{ltpd}%</strong> defective at
                  <strong>90%</strong> confidence with <strong>{k}</strong> allowed failure(s)
                </div>
              </div>
            </div>
            <div class="card mb-2">
              <div class="card-body py-3 px-4" style="font-size:.82rem">
                <div class="mb-1"><strong>JESD47I §3.8 formula:</strong></div>
                <div class="text-muted font-monospace">N &ge; 0.5 &times; &chi;&sup2;(2C+2,&thinsp;0.1) &times; (1/LTPD &minus; 0.5) + C</div>
                <div class="mt-2 text-muted">
                  Exact chi-squared (R={r_equiv:.3f}, C=90%): <strong>{n_exact}</strong>
                  {"&ensp;<span class='badge bg-secondary'>same</span>" if n_exact == n_jesd47 else f"&ensp;<span class='text-muted'>({'+' if n_jesd47>n_exact else ''}{n_jesd47-n_exact} vs JESD47I)</span>"}
                </div>
              </div>
            </div>"""
        else:
            result_html = """<div class="card text-center p-5" style="color:var(--df-grey)">
              <p class="mb-0" style="font-size:.85rem">Select parameters and click Calculate &rarr;</p></div>"""

        body = f"""
        <h4 class="mb-4" style="font-weight:300">Sample Size Planner</h4>
        <div class="row g-4">
          <div class="col-md-4">
            <div class="card">
              <div class="card-df"><h6 class="mb-0">Parameters</h6></div>
              <div class="card-body p-4">
                <p class="text-muted small mb-3">
                  Per <strong>JESD47I §3.8</strong> — how many units are required to satisfy
                  a maximum defect level at 90% confidence?
                </p>
                <form method="post">
                  <div class="mb-3">
                    <label class="form-label">LTPD — Max % Defective</label>
                    <div class="input-group">
                      <input type="number" class="form-control" name="ltpd"
                             value="{ltpd}" min="0.01" max="99.99" step="0.01" required>
                      <span class="input-group-text">%</span>
                    </div>
                    <div class="form-text">Lot Tolerance Percent Defective (any value)</div>
                  </div>
                  <div class="mb-4">
                    <label class="form-label">Acceptance Number (C)</label>
                    <input type="number" class="form-control" name="failures"
                           value="{k}" min="0" step="1" required>
                    <div class="form-text">Max allowed failures (0 = zero-failure plan)</div>
                  </div>
                  <button type="submit" class="btn btn-primary w-100">Calculate &rarr;</button>
                </form>
                <hr class="my-3">
                <p class="text-muted mb-0" style="font-size:.78rem">
                  Confidence is fixed at 90% per JESD47I.<br>
                  LTPD = (1&minus;R) &times; 100, e.g. LTPD&nbsp;5% &equiv; R&nbsp;=&nbsp;95%.
                </p>
              </div>
            </div>
          </div>
          <div class="col-md-8">
            {result_html}
            {table_a_html}
          </div>
        </div>"""
        self.emit(body, "Sample Size Planner", "sample-size")


# ── /pass-fail ────────────────────────────────────────────────────────────────

class PassFailHandler(Base):
    def get(self):
        self._render()

    def post(self):
        try:
            n        = int(self.get_argument("n",          ""))
            k        = int(self.get_argument("failures",   "0"))
            c        = float(self.get_argument("confidence", "0.90"))
            ltpd_inp = float(self.get_argument("ltpd_pct", "5.0"))
            if n < 1:        raise ValueError("Need at least 1 sample")
            if k > n:        raise ValueError("Failures cannot exceed samples tested")
            if not (0.50 <= c <= 0.9999):            raise ValueError("Confidence must be 0.50–0.9999")
            if not (0.01 <= ltpd_inp <= 99.99):      raise ValueError("Defective Rate must be 0.01–99.99%")
            r_req = 1.0 - ltpd_inp / 100.0
            passed, r_demo = _pf(n, k, c, r_req)
            sens = [(f, demonstrated_reliability(n, f, c)) for f in range(min(n+1, 9))]
            self._render(n=n, k=k, c=c, r_req=r_req, ltpd_inp=ltpd_inp,
                         passed=passed, r_demo=r_demo, sens=sens)
        except (ValueError, TypeError) as e:
            self._render(error=str(e))

    def _render(self, n="", k=0, c=0.90, r_req=0.95, ltpd_inp=5.0,
                passed=None, r_demo=None, sens=None, error=None):
        result_html = ""
        if error:
            result_html = f'<div class="alert alert-danger"><i class="bi bi-exclamation-triangle me-2"></i>{error}</div>'
        elif passed is not None:
            verdict_cls  = "result-pass" if passed else "result-fail"
            verdict_icon = "check-circle-fill" if passed else "x-circle-fill"
            verdict_text = "PASS" if passed else "FAIL"
            defect_demo  = (1.0 - r_demo) * 100.0
            detail_text  = (f"Demonstrated defect rate {defect_demo:.2f}% ≤ LTPD {ltpd_inp:.2f}%" if passed
                            else f"Demonstrated defect rate {defect_demo:.2f}% exceeds LTPD {ltpd_inp:.2f}%")
            # JESD47I Table A minimum n for equivalent LTPD
            ltpd_pct   = ltpd_inp
            # Find closest standard LTPD column
            std_ltpd   = min(TABLE_A_LTPD, key=lambda l: abs(l - ltpd_pct))
            jesd47_n   = min_sample_size_ltpd(std_ltpd, k) if k <= 12 else None
            jesd47_note = ""
            if jesd47_n is not None:
                ltpd_ok = n >= jesd47_n
                ltpd_badge = "success" if ltpd_ok else "danger"
                jesd47_note = (f'<div class="text-muted small mt-2">'
                               f'JESD47I Table A (LTPD&nbsp;{std_ltpd}%, C={k}): '
                               f'min&nbsp;n&nbsp;=&nbsp;<strong>{jesd47_n}</strong> &ensp;'
                               f'<span class="badge bg-{ltpd_badge}">{"meets" if ltpd_ok else "below"} JESD47I</span>'
                               f'</div>')

            extra = ""
            if not passed:
                n_needed = min_sample_size(r_req, c, k)
                gap = n_needed - n
                if gap > 0:
                    extra = f'<div class="text-muted small mt-1">→ {gap} more samples needed (0 additional failures) to reach target</div>'

            def _pf_row(f, r_f, k=k, r_req=r_req):
                tr_cls   = 'class="table-primary"' if f == k else ""
                bg       = "bg-success" if r_f >= r_req else "bg-danger"
                verdict  = "PASS" if r_f >= r_req else "FAIL"
                actual   = '<span class="badge bg-primary ms-1">actual</span>' if f == k else ""
                return (f'<tr {tr_cls}><td>{f}</td><td>{r_f*100:.2f}%</td>'
                        f'<td><span class="badge {bg}">{verdict}</span>{actual}</td></tr>')
            sens_rows = "".join(_pf_row(f, r_f) for f, r_f in (sens or []))
            result_html = f"""
            <div class="card mb-3">
              <div class="card-body text-center py-4">
                <div class="{verdict_cls} mb-1"><i class="bi bi-{verdict_icon} me-2"></i>{verdict_text}</div>
                <div class="text-muted small">{detail_text}</div>
                {extra}
                {jesd47_note}
                <hr class="my-3">
                <div class="row g-3 text-center">
                  <div class="col">
                    <div class="text-muted" style="font-size:.75rem">DEFECT RATE (DEMO)</div>
                    <div class="fw-bold fs-5">{defect_demo:.2f}%</div>
                  </div>
                  <div class="col">
                    <div class="text-muted" style="font-size:.75rem">LTPD (MAX)</div>
                    <div class="fw-bold fs-5">{ltpd_inp:.2f}%</div>
                  </div>
                  <div class="col">
                    <div class="text-muted" style="font-size:.75rem">CONFIDENCE</div>
                    <div class="fw-bold fs-5">{c*100:.0f}%</div>
                  </div>
                  <div class="col">
                    <div class="text-muted" style="font-size:.75rem">n / failures</div>
                    <div class="fw-bold fs-5">{n} / {k}</div>
                  </div>
                </div>
              </div>
            </div>
            <div class="card">
              <div class="card-header bg-light py-2">
                <strong>Sensitivity Analysis</strong>
                <span class="text-muted small ms-2">same n={n}, same confidence — what if failures differed?</span>
              </div>
              <table class="table table-hover table-sm mb-0">
                <thead class="tbl-header"><tr><th>Failures</th><th>Demonstrated R</th><th>Result vs Target</th></tr></thead>
                <tbody>{sens_rows}</tbody>
              </table>
            </div>"""
        else:
            result_html = """<div class="card text-center p-5" style="color:var(--df-grey)">
              <p class="mb-0" style="font-size:.85rem;letter-spacing:.05em">Enter test results and click Evaluate &rarr;</p></div>"""

        body = f"""
        <h4 class="mb-4" style="font-weight:300">Pass / Fail Determination</h4>
        <div class="row g-4">
          <div class="col-md-5">
            <div class="card">
              <div class="card-df"><h6 class="mb-0">Test Results</h6></div>
              <div class="card-body p-4">
                <p class="text-muted small mb-3">Given actual results, what reliability was demonstrated?</p>
                <form method="post">
                  <div class="mb-3">
                    <label class="form-label">Samples Tested</label>
                    <input type="number" class="form-control" name="n" value="{n}" min="1" required placeholder="e.g. 77">
                  </div>
                  <div class="mb-3">
                    <label class="form-label">Failures Observed</label>
                    <input type="number" class="form-control" name="failures" value="{k}" min="0" required>
                  </div>
                  <div class="mb-3">
                    <label class="form-label">Confidence Level</label>
                    <div class="input-group">
                      <input type="number" class="form-control" name="confidence" value="{c}" step="0.001" min="0.50" max="0.9999" required>
                      <span class="input-group-text text-muted small">e.g. 0.90</span>
                    </div>
                  </div>
                  <div class="mb-4">
                    <label class="form-label">Defective Rate % (LTPD)</label>
                    <div class="input-group">
                      <input type="number" class="form-control" name="ltpd_pct" value="{ltpd_inp}" step="0.01" min="0.01" max="99.99" required>
                      <span class="input-group-text text-muted small">% defective</span>
                    </div>
                  </div>
                  <button type="submit" class="btn btn-primary w-100">
                    <i class="bi bi-check2-circle me-2"></i>Evaluate
                  </button>
                </form>
              </div>
            </div>
          </div>
          <div class="col-md-7">{result_html}</div>
        </div>"""
        self.emit(body, "Pass / Fail", "pass-fail")


# ── /report (GET = form, POST = generate) ─────────────────────────────────────

class ReportHandler(Base):
    def get(self):
        _, s = self.sess()
        pt = s.get("part_type", "ttv")
        tests = applicable_tests(pt)
        self._render_form(tests)

    def post(self):
        _, s = self.sess()
        pt    = s.get("part_type", "ttv")
        tests = applicable_tests(pt)

        customer = self.get_argument("customer", "Customer").strip() or "Customer"
        product  = self.get_argument("product",  "SCD-on-Si Package").strip()
        author   = self.get_argument("author",   "").strip()

        # Parse part description
        part_desc = {}
        if pt == "ttv":
            part_desc["si_thick"]      = self.get_argument("part_si_thick", "50").strip() or "50"
            part_desc["bond_type"]     = self.get_argument("part_bond_type", "Ag Sinter").strip() or "Ag Sinter"
            part_desc["diamond_thick"] = self.get_argument("part_diamond_thick", "300").strip() or "300"
        else:
            part_desc["part_number"] = self.get_argument("part_number", "").strip()
            part_desc["part_tech"]   = self.get_argument("part_tech", "").strip()
            part_desc["part_package"] = self.get_argument("part_package", "").strip()

        entries = []
        samples = {}
        for key, t in tests.items():
            sc    = self.get_argument(f"status_{key}", "ns")
            label = STATUS_LABEL.get(sc, "NOT STARTED")
            n_raw = self.get_argument(f"n_{key}", "").strip()
            k_raw = self.get_argument(f"k_{key}", "0").strip()
            notes = self.get_argument(f"notes_{key}", "").strip()

            try:    n_val = int(n_raw) if n_raw else None
            except: n_val = None
            try:    k_val = int(k_raw)
            except: k_val = 0

            is_char = t.get("characterization_only", False)
            char_result = self.get_argument(f"char_result_{key}", "").strip() if is_char else ""

            r_demo = stat_pass = None
            corrected = False
            if not is_char:
                # Compute reliability stats whenever test is in-progress or complete
                # Use project-specific LTPD and confidence if injected by ProjectReportHandler
                _conf = getattr(self, "_qual_confidence", 0.90)
                _ltpd = getattr(self, "_qual_ltpd", 5.0)
                _r_req = 1.0 - _ltpd / 100.0
                if sc in ("co", "ip") and n_val:
                    r_demo    = demonstrated_reliability(n_val, k_val, _conf)
                    stat_pass, _ = _pf(n_val, k_val, _conf, _r_req)

            # Manual pass/fail override (form field from clickable badge)
            _ov_raw = self.get_argument(f"pass_override_{key}", "auto").strip()
            if _ov_raw == "force_pass":
                pass_override  = "force_pass"
                effective_pass = True if r_demo is not None else None
            elif _ov_raw == "force_fail":
                pass_override  = "force_fail"
                effective_pass = False if r_demo is not None else None
            else:
                pass_override  = "auto"
                effective_pass = stat_pass

            entries.append({
                "key": key, "test": t, "sc": sc, "status": label,
                "n": n_val, "k": k_val, "notes": notes,
                "r_demo": r_demo, "stat_pass": stat_pass,
                "pass_override": pass_override, "effective_pass": effective_pass,
                "corrected": corrected,
                "is_char": is_char, "char_result": char_result,
            })

            # Parse sample records (TTV and Die)
            if pt in ("ttv", "die"):
                is_ttv = (pt == "ttv")
                samples_for_test = []
                n_samples_key = f"n_samples_{key}"
                n_samples_str = self.get_argument(n_samples_key, "0").strip() or "0"
                try:
                    n_samples = int(n_samples_str)
                except:
                    n_samples = 0

                for i in range(n_samples):
                    sid = self.get_argument(f"sid_{key}_{i}", f"SN-{i+1:03d}").strip() or f"SN-{i+1:03d}"

                    # CSAM values
                    csam_bpc_str = self.get_argument(f"csam_bpc_{key}_{i}", "").strip()
                    csam_apc_str = self.get_argument(f"csam_apc_{key}_{i}", "").strip()
                    csam_atst_str = self.get_argument(f"csam_atst_{key}_{i}", "").strip()

                    try:
                        csam_bpc = float(csam_bpc_str) if csam_bpc_str else None
                    except:
                        csam_bpc = None
                    try:
                        csam_apc = float(csam_apc_str) if csam_apc_str else None
                    except:
                        csam_apc = None
                    try:
                        csam_atst = float(csam_atst_str) if csam_atst_str else None
                    except:
                        csam_atst = None

                    csam_status, csam_badge = _csam_eval(csam_bpc, csam_apc, csam_atst)

                    # Thermal and Func (TTV only; Die uses CSAM-only pass logic)
                    if is_ttv:
                        thermal = self.get_argument(f"thermal_{key}_{i}", "pass").strip() or "pass"
                        func    = self.get_argument(f"func_{key}_{i}", "pass").strip() or "pass"
                        failed_rtds_str = self.get_argument(f"failed_rtds_{key}_{i}", "").strip()
                        failed_rtds = [x.strip().upper() for x in failed_rtds_str.split(",") if x.strip()]
                    else:
                        thermal = "pass"
                        func    = "pass"
                        failed_rtds = []

                    # File uploads (base64)
                    img_bpc = ""
                    img_apc = ""
                    img_atst = ""

                    for suffix, var_name in [("bpc", "img_bpc"), ("apc", "img_apc"), ("atst", "img_atst")]:
                        files = self.request.files.get(f"img_{suffix}_{key}_{i}", [])
                        if files:
                            f = files[0]
                            mime = f.get("content_type", "image/jpeg")
                            b64 = base64.b64encode(f["body"]).decode()
                            img_uri = f"data:{mime};base64,{b64}"
                            if var_name == "img_bpc":
                                img_bpc = img_uri
                            elif var_name == "img_apc":
                                img_apc = img_uri
                            elif var_name == "img_atst":
                                img_atst = img_uri

                    samples_for_test.append({
                        "id": sid,
                        "csam_bpc": csam_bpc,
                        "csam_apc": csam_apc,
                        "csam_atst": csam_atst,
                        "csam_status": csam_status,
                        "csam_badge": csam_badge,
                        "thermal": thermal,
                        "failed_rtds": failed_rtds,
                        "func": func,
                        "img_bpc": img_bpc,
                        "img_apc": img_apc,
                        "img_atst": img_atst,
                    })

                samples[key] = samples_for_test

        _r_conf  = getattr(self, "_qual_confidence", 0.90)
        _r_ltpd  = getattr(self, "_qual_ltpd", 5.0)
        _r_conds = getattr(self, "_test_conditions", {})
        _stat_mode = self.get_argument("stat_mode", "full").strip()
        _show_stats = (_stat_mode != "pf_only")

        # Pass/fail counts depend on mode:
        # full    → chi-squared effective_pass (respects manual override)
        # pf_only → 0 failures = pass, any failures = fail (for Complete tests)
        if _show_stats:
            _n_pass = sum(1 for e in entries if e["sc"] == "co" and e.get("effective_pass") is True)
            _n_fail = sum(1 for e in entries if e["sc"] == "co" and e.get("effective_pass") is False)
        else:
            _n_pass = sum(1 for e in entries if e["sc"] == "co" and not e.get("is_char") and (e.get("k") or 0) == 0)
            _n_fail = sum(1 for e in entries if e["sc"] == "co" and not e.get("is_char") and (e.get("k") or 0) > 0)

        report = {
            "customer": customer, "product": product,
            "author": author, "part_type": pt,
            "part_label": PART_TYPE_LABELS.get(pt, "Active Device"),
            "part_desc": part_desc,
            "date": datetime.now().strftime("%B %d, %Y"),
            "entries": entries,
            "samples": samples,
            # Project qualification criteria (from Planner)
            "qual_ltpd":       _r_ltpd,
            "qual_confidence": _r_conf,
            "qual_r_req":      1.0 - _r_ltpd / 100.0,
            "test_conditions": _r_conds,
            "show_stats":      _show_stats,
            "n_pass": _n_pass,
            "n_fail": _n_fail,
            "n_co":   sum(1 for e in entries if e["sc"] == "co"),
            "n_ip":   sum(1 for e in entries if e["sc"] == "ip"),
            "n_ns":   sum(1 for e in entries if e["sc"] == "ns"),
            "n_char": sum(1 for e in entries if e["sc"] == "ch"),
            "total":  len(entries),
        }
        s["last_report"] = report
        self._render_report(report)

    # ── form ──────────────────────────────────────────────────────────────────

    def _render_form(self, tests: dict, saved_counts: dict = None):
        _, s = self.sess()
        pt = s.get("part_type", "ttv")
        # Allow subclasses to inject per-test sample counts via instance attr
        if saved_counts is None:
            saved_counts = getattr(self, "_saved_n", {})

        # Pre-compute min sample sizes for k=0..20 failures using project LTPD + confidence
        # Fall back to JESD47 defaults (95% R, 90% CL) when not in a project context
        _qual_ltpd       = getattr(self, "_qual_ltpd", 5.0)
        _qual_confidence = getattr(self, "_qual_confidence", 0.90)
        _r_req           = 1.0 - _qual_ltpd / 100.0
        min_n_table = [min_sample_size(_r_req, _qual_confidence, k) for k in range(21)]
        min_n_js = str(min_n_table)
        # Human-readable labels for JS badge tooltip
        _ltpd_pct_js   = _qual_ltpd
        _conf_pct_js   = round(_qual_confidence * 100, 1)
        _r_req_pct_js  = round(_r_req * 100, 1)

        status_opts_html = "".join(
            f'<option value="{sc}">{lbl}</option>'
            for sc, lbl, _ in STATUS_OPTS
        )
        status_opts_char_html = "".join(
            f'<option value="{sc}">{lbl}</option>'
            for sc, lbl, _ in STATUS_OPTS_CHAR
        )

        test_rows = ""
        for key, t in tests.items():
            # Prefer saved planner count, fall back to TESTS dict default.
            # saved_counts[key] may be a plain int/str (planner count) or a list
            # of sample dicts (stored from a previous report submission) — handle both.
            sz = str(t["sample_size"]) if t["sample_size"] else "5"
            if key in saved_counts and saved_counts[key]:
                val = saved_counts[key]
                if isinstance(val, list):
                    # Report-format: list of sample dicts → use length as count
                    sz = str(len(val)) if val else sz
                else:
                    try:
                        sz = str(int(val))
                    except (TypeError, ValueError):
                        pass  # keep default
            is_char = t.get("characterization_only", False)
            opts_html = status_opts_char_html if is_char else status_opts_html

            # Pre-compute conditional cell HTML (no backslash allowed inside f-strings)
            if is_char:
                k_cell = ""
                char_hint = '<div class="text-muted small mt-1" style="font-style:italic">Characterization only — record warpage result in field above</div>'
            else:
                k_cell = (
                    '<td class="align-middle sf-' + key + '" style="min-width:130px">'
                    '<div class="d-flex align-items-center gap-1">'
                    '<input type="number" class="form-control form-control-sm k-input" name="k_' + key + '"'
                    ' id="k_' + key + '" value="0" min="0" placeholder="k" data-key="' + key + '"'
                    ' oninput="updatePF(\'' + key + '\')">'
                    '<input type="hidden" name="pass_override_' + key + '" id="override_' + key + '" value="auto">'
                    '<span id="pf_' + key + '" style="font-size:.72rem;font-weight:700;'
                    'white-space:nowrap;padding:2px 7px;border-radius:4px;'
                    'background:#f3f4f6;color:#6b7280;cursor:pointer;user-select:none" '
                    'onclick="cycleOverride(\'' + key + '\')" title="Click to override">—</span>'
                    '</div>'
                    '</td>'
                )
                char_hint = ""

            # For characterization tests: replace n/k with a warpage result field
            if is_char:
                nk_cells = f"""
              <td class="align-middle sf-{key}" colspan="2" style="min-width:160px">
                <input type="text" class="form-control form-control-sm" name="char_result_{key}"
                       id="n_{key}" placeholder="e.g. +125 µm at 260°C" data-key="{key}"
                       title="Record peak warpage result (signed, with units)">
              </td>"""
            else:
                nk_cells = f"""
              <td class="align-middle sf-{key}" style="min-width:80px">
                <input type="number" class="form-control form-control-sm n-input" name="n_{key}"
                       id="n_{key}" value="{sz}" min="1" placeholder="n" data-key="{key}"
                       oninput="updatePF('{key}')">
              </td>"""

            _spec_docs = SPEC_URLS.get(key, [])
            if _spec_docs:
                std_cell = " / ".join(
                    f'<a href="{url}" target="_blank" rel="noopener" '
                    f'class="text-muted text-decoration-none" title="Open {lbl}">'
                    f'{lbl} <i class="bi bi-box-arrow-up-right" style="font-size:.65rem"></i></a>'
                    for lbl, url in _spec_docs
                )
            else:
                std_cell = t["standard"]

            test_rows += f"""
            <tr id="row-{key}">
              <td class="align-middle fw-semibold" style="min-width:170px">{t['name']}</td>
              <td class="align-middle text-muted small">{std_cell}</td>
              <td class="align-middle" style="min-width:140px">
                <select class="form-select form-select-sm status-sel" name="status_{key}" data-key="{key}" {"data-char=true" if is_char else ""}>
                  {opts_html}
                </select>
              </td>
              {nk_cells}
              {k_cell}
              <td>
                <input type="text" class="form-control form-control-sm" name="notes_{key}" placeholder="optional">
                {char_hint}
                <div id="warn-{key}" class="text-danger small mt-1" style="display:none">
                  <i class="bi bi-exclamation-triangle-fill me-1"></i>
                  <span id="warn-msg-{key}"></span>
                </div>
              </td>
            </tr>"""

        form_action = getattr(self, "_form_action", "/report")
        body = f"""
        <h4 class="mb-4" style="font-weight:300">Qualification Report</h4>
        <form method="post" action="{form_action}" enctype="multipart/form-data" id="report-form">
          <!-- Customer info -->
          <div class="card mb-4">
            <div class="card-df"><h6 class="mb-0">Report Information</h6></div>
            <div class="card-body p-4">
              <div class="row g-3">
                <div class="col-md-4">
                  <label class="form-label">Customer Name</label>
                  <input type="text" class="form-control" name="customer" placeholder="Customer">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Product / Program</label>
                  <input type="text" class="form-control" name="product" value="SCD-on-Si Package">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Prepared By</label>
                  <input type="text" class="form-control" name="author" placeholder="e.g. Reliability Engineering">
                </div>
                <div class="col-12">
                  <label class="form-label mb-1">Report Mode</label>
                  <div class="d-flex gap-3">
                    <div class="form-check">
                      <input class="form-check-input" type="radio" name="stat_mode" id="sm_full" value="full" checked>
                      <label class="form-check-label" for="sm_full" style="font-size:.9rem">
                        Include statistical analysis
                        <small class="text-muted d-block" style="font-size:.78rem">Shows demonstrated reliability, confidence interval, and chi-squared pass/fail criteria</small>
                      </label>
                    </div>
                    <div class="form-check">
                      <input class="form-check-input" type="radio" name="stat_mode" id="sm_pf" value="pf_only">
                      <label class="form-check-label" for="sm_pf" style="font-size:.9rem">
                        Pass / Fail only
                        <small class="text-muted d-block" style="font-size:.78rem">Shows n and failures per test — no confidence interval or statistical criteria</small>
                      </label>
                    </div>
                  </div>
                </div>
              </div>
            </div>
          </div>
          <!-- Part Description -->
          <div class="card mb-4">
            <div class="card-df"><h6 class="mb-0">Part Description</h6></div>
            <div class="card-body p-4">"""

        if pt == "ttv":
            body += f"""
              <p style="font-style:italic;color:var(--df-grey);margin-bottom:1rem;font-size:.85rem">Nanotest TTV10-NT20 — 24.9 × 24.9mm flip-chip package, SAC305 solder on PCB. Die has 16 RTD temperature sensors (T1–T16), 4 heater zones, and 6 hotspot cells.</p>
              <div class="row g-3">
                <div class="col-md-4">
                  <label class="form-label">Si Die Thickness (µm)</label>
                  <input type="text" class="form-control" name="part_si_thick" value="50">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Bond Technique</label>
                  <select class="form-select" name="part_bond_type">
                    <option value="Cu TCB">Cu TCB</option>
                    <option value="Ag Sinter" selected>Ag Sinter</option>
                    <option value="Ag TCB">Ag TCB</option>
                  </select>
                </div>
                <div class="col-md-4">
                  <label class="form-label">SCD Thickness (µm)</label>
                  <input type="text" class="form-control" name="part_diamond_thick" value="300">
                </div>
              </div>"""
        else:
            body += f"""
              <p style="color:var(--df-grey);margin-bottom:1rem;font-size:.85rem">Active device — part details are customer-specific.</p>
              <div class="row g-3">
                <div class="col-md-4">
                  <label class="form-label">Part Number</label>
                  <input type="text" class="form-control" name="part_number" placeholder="">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Technology Node</label>
                  <input type="text" class="form-control" name="part_tech" placeholder="">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Package Type</label>
                  <input type="text" class="form-control" name="part_package" placeholder="">
                </div>
              </div>"""

        body += f"""
            </div>
          </div>
          <!-- Test status table -->
          <div class="card mb-4">
            <div class="card-df"><h6 class="mb-0">Test Status</h6></div>
            <div class="table-responsive">
              <table class="table table-sm table-hover align-middle mb-0" style="min-width:700px">
                <thead class="tbl-header">
                  <tr>
                    <th>Test</th>
                    <th>Standard</th>
                    <th>Status</th>
                    <th title="Samples tested">n</th>
                    <th title="Failures observed">failures</th>
                    <th>Notes</th>
                  </tr>
                </thead>
                <tbody>{test_rows}</tbody>
              </table>
            </div>
          </div>
          <!-- Sample Records (TTV and Die) -->"""

        if pt in ("ttv", "die"):
            _is_ttv = (pt == "ttv")
            _sr_desc = (
                "<strong>CSAM:</strong> ≥95% bond area at all three timepoints.&ensp;"
                "<strong>Functionality:</strong> sensor reading within 5% of original (no heating).&ensp;"
                "<strong>Thermal:</strong> sensor reading within 5% of original (with heating)."
                if _is_ttv else
                "<strong>CSAM:</strong> ≥95% bond area at all three timepoints."
            )
            body += f"""
          <div class="card mb-4">
            <div class="card-df"><h6 class="mb-0">Sample Records</h6></div>
            <div class="card-body p-4">
              <p style="font-size:.85rem;color:var(--df-grey);margin-bottom:1rem">
                {_sr_desc}
              </p>"""

            _ttv_extra_th = """
                              <th style="min-width:70px">Thermal</th>
                              <th style="min-width:120px">Failed Sensors</th>
                              <th style="min-width:70px">Func</th>""" if _is_ttv else ""

            for key, t in tests.items():
                body += f"""
              <div class="accordion mb-3" style="border:1px solid var(--df-border)">
                <div class="accordion-item">
                  <h2 class="accordion-header">
                    <button class="accordion-button collapsed py-2" type="button" data-bs-toggle="collapse" data-bs-target="#acc_{key}">
                      <strong>{t['name']}</strong>
                    </button>
                  </h2>
                  <div id="acc_{key}" class="accordion-collapse collapse">
                    <div class="accordion-body pt-3 pb-2">
                      <div class="row g-2 align-items-end mb-3">
                        <div class="col-auto">
                          <label class="form-label mb-2">Number of samples</label>
                          <input type="number" class="form-control form-control-sm n-samples-input" name="n_samples_{key}" value="{t.get('sample_size', 0)}" min="0" max="20" data-key="{key}" style="width:100px">
                        </div>
                        <div class="col-auto">
                          <button type="button" class="btn btn-sm btn-outline-secondary" onclick="buildSampleRows('{key}', parseInt(document.querySelector('[name=\\'n_samples_{key}\\']').value) || 0)">
                            Generate Rows
                          </button>
                        </div>
                      </div>
                      <div class="table-responsive" style="font-size:.8rem">
                        <table class="table table-sm table-bordered mb-0">
                          <thead class="tbl-header">
                            <tr>
                              <th style="min-width:100px">Sample ID</th>
                              <th style="min-width:110px">CSAM Before PC (%)</th>
                              <th style="min-width:110px">CSAM After PC (%)</th>
                              <th style="min-width:110px">CSAM After Test (%)</th>{_ttv_extra_th}
                              <th style="min-width:240px">Images</th>
                            </tr>
                          </thead>
                          <tbody id="stbody_{key}">
                          </tbody>
                        </table>
                      </div>
                    </div>
                  </div>
                </div>
              </div>"""

            body += f"""
            </div>
          </div>
          <input type="hidden" id="part-type-input" value="{pt}">
          <script>
            const _isTTV = {str(_is_ttv).lower()};
            function buildSampleRows(key, n) {{
              const tbody = document.getElementById('stbody_' + key);
              if (!tbody) return;
              tbody.innerHTML = '';
              for (let i = 0; i < n; i++) {{
                const sid = 'SN-' + String(i+1).padStart(3,'0');
                const ttvCells = _isTTV ? `
                  <td><select class="form-select form-select-sm" name="thermal_${{key}}_${{i}}"><option value="pass">Pass</option><option value="fail">Fail</option></select></td>
                  <td><input type="text" class="form-control form-control-sm" name="failed_rtds_${{key}}_${{i}}" placeholder="e.g. T3,T7"></td>
                  <td><select class="form-select form-select-sm" name="func_${{key}}_${{i}}"><option value="pass">Pass</option><option value="fail">Fail</option></select></td>` : '';
                tbody.innerHTML += `<tr>
                  <td><input type="text" class="form-control form-control-sm" name="sid_${{key}}_${{i}}" value="${{sid}}"></td>
                  <td><input type="number" class="form-control form-control-sm" name="csam_bpc_${{key}}_${{i}}" step="0.1" min="0" max="100" placeholder="—"></td>
                  <td><input type="number" class="form-control form-control-sm" name="csam_apc_${{key}}_${{i}}" step="0.1" min="0" max="100" placeholder="—"></td>
                  <td><input type="number" class="form-control form-control-sm" name="csam_atst_${{key}}_${{i}}" step="0.1" min="0" max="100" placeholder="—"></td>
                  ${{ttvCells}}
                  <td style="min-width:240px">
                    <div class="mb-1"><label style="font-size:.7rem">Pre-PC</label><input type="file" class="form-control form-control-sm" name="img_bpc_${{key}}_${{i}}" accept="image/*"></div>
                    <div class="mb-1"><label style="font-size:.7rem">Post-PC</label><input type="file" class="form-control form-control-sm" name="img_apc_${{key}}_${{i}}" accept="image/*"></div>
                    <div><label style="font-size:.7rem">Post-Test</label><input type="file" class="form-control form-control-sm" name="img_atst_${{key}}_${{i}}" accept="image/*"></div>
                  </td>
                </tr>`;
              }}
            }}

            // Auto-init: listen to n_samples inputs
            document.querySelectorAll('.n-samples-input').forEach(inp => {{
              const key = inp.dataset.key;
              buildSampleRows(key, parseInt(inp.value) || 0);
              inp.addEventListener('change', function() {{
                buildSampleRows(key, parseInt(this.value) || 0);
              }});
            }});
          </script>"""

        else:
            body += f"""
          <div class="card mb-4">
            <div class="card-body p-4">
              <p style="font-size:.85rem;color:var(--df-grey)">Sample tracking not yet configured for active devices.</p>
            </div>
          </div>"""

        body += f"""
          <button type="submit" class="btn btn-primary btn-lg px-5">
            <i class="bi bi-file-earmark-check me-2"></i>Generate Report
          </button>
        </form>

        <script>
        // Pre-computed min sample sizes using project LTPD ({_ltpd_pct_js}%) and
        // confidence ({_conf_pct_js}%): MIN_N[k] = min n to demonstrate ≤{_ltpd_pct_js}% defective
        const MIN_N = {min_n_js};
        const QUAL_LTPD_PCT  = {_ltpd_pct_js};
        const QUAL_CONF_PCT  = {_conf_pct_js};
        const QUAL_R_PCT     = {_r_req_pct_js};

        // cycleOverride: click badge to cycle auto → force_pass → force_fail → auto
        function cycleOverride(key) {{
          const hidden = document.getElementById('override_' + key);
          const nEl   = document.getElementById('n_' + key);
          if (!hidden || !nEl) return;
          if ((parseInt(nEl.value) || 0) === 0) return;
          const cur = hidden.value;
          hidden.value = cur === 'auto' ? 'force_pass' : cur === 'force_pass' ? 'force_fail' : 'auto';
          updatePF(key);
        }}

        // updatePF: compute and display pass/fail badge, respecting manual override
        function updatePF(key) {{
          const badge  = document.getElementById('pf_' + key);
          if (!badge) return;  // characterization test, no badge
          const nEl    = document.getElementById('n_' + key);
          const kEl    = document.getElementById('k_' + key);
          const hidden = document.getElementById('override_' + key);
          if (!nEl || !kEl) return;
          const n = parseInt(nEl.value) || 0;
          const k = parseInt(kEl.value) || 0;
          if (n === 0) {{
            badge.textContent = '—';
            badge.style.background = '#f3f4f6';
            badge.style.color      = '#6b7280';
            badge.style.outline    = '';
            badge.title = 'Click to override';
            if (hidden) hidden.value = 'auto';
            return;
          }}
          const minN     = (k < MIN_N.length) ? MIN_N[k] : MIN_N[MIN_N.length - 1];
          const autoPass = n >= minN;
          const override = hidden ? hidden.value : 'auto';
          const isOverride = override !== 'auto';
          const isPass = override === 'force_pass' ? true : override === 'force_fail' ? false : autoPass;

          badge.title = isOverride
            ? `Manually overridden — click to cycle (auto→pass→fail). Auto result: ${{autoPass ? 'PASS' : 'FAIL'}} (need n≥${{minN}})`
            : `n≥${{minN}} needed to pass (≤${{QUAL_LTPD_PCT}}% defective @ ${{QUAL_CONF_PCT}}% CL). Click to override.`;
          badge.style.outline = isOverride ? '2px solid #f97316' : '';
          if (isPass) {{
            badge.textContent    = isOverride ? '✓ PASS' : 'PASS';
            badge.style.background = '#dcfce7';
            badge.style.color      = '#15803d';
          }} else {{
            badge.textContent    = isOverride ? '✗ FAIL' : 'FAIL';
            badge.style.background = '#fee2e2';
            badge.style.color      = '#b91c1c';
          }}
        }}

        function refreshFields() {{
          document.querySelectorAll('.status-sel').forEach(sel => {{
            const key = sel.dataset.key;
            const isChar = sel.dataset.char === 'true';
            // Show n/k cells only when status is complete (or characterized for char tests)
            const show = isChar ? ['ip','ch'].includes(sel.value) : sel.value === 'co';
            document.querySelectorAll('.sf-' + key).forEach(cell => {{
              cell.style.opacity = show ? '1' : '0.4';
              cell.querySelectorAll('input').forEach(i => i.disabled = !show);
            }});
            if (!isChar) updatePF(key);
          }});
        }}

        document.querySelectorAll('.status-sel').forEach(s => {{
          s.addEventListener('change', function() {{
            const key = this.dataset.key;
            const isChar = this.dataset.char === 'true';
            const show = isChar ? ['ip','ch'].includes(this.value) : this.value === 'co';
            document.querySelectorAll('.sf-' + key).forEach(cell => {{
              cell.style.opacity = show ? '1' : '0.4';
              cell.querySelectorAll('input').forEach(i => i.disabled = !show);
            }});
            if (!isChar) updatePF(key);
          }});
        }});

        refreshFields();
        </script>"""
        self.emit(body, "Qual Report", "report")

    # ── report display ────────────────────────────────────────────────────────

    def _render_report(self, r: dict):
        # Banner for any auto-corrected entries
        corrected_entries = [e for e in r["entries"] if e.get("corrected")]
        correction_banner = ""
        if corrected_entries:
            names = ", ".join(e["test"]["name"] for e in corrected_entries)
            correction_banner = f"""
            <div class="alert alert-warning d-flex align-items-start gap-2 mb-4" role="alert">
              <i class="bi bi-exclamation-triangle-fill fs-5 mt-1 flex-shrink-0"></i>
              <div>
                <strong>Status corrected on {len(corrected_entries)} test(s):</strong> {names}<br>
                <small>These were submitted as <em>Pass</em> but the n / failures values do not demonstrate
                ≥95% reliability at 90% confidence (JEDEC chi-squared method). Status has been
                overridden to <strong>Fail</strong>.</small>
              </div>
            </div>"""

        # Summary count tiles
        tiles = ""
        for count, label, color in [
            (r["n_pass"],             "PASS",           "success"),
            (r["n_fail"],             "FAIL",           "danger"),
            (r.get("n_co", 0),        "COMPLETE",       "success"),
            (r["n_ip"],               "IN PROGRESS",    "warning"),
            (r["n_ns"],               "NOT STARTED",    "secondary"),
            (r.get("n_char", 0),      "CHARACTERIZED",  "info"),
        ]:
            # Only show PASS/FAIL tiles when some tests are complete with stats
            if label in ("PASS", "FAIL") and r.get("n_co", 0) == 0:
                continue
            if count == 0 and label in ("CHARACTERIZED", "PASS", "FAIL"):
                continue  # hide tile if count is zero
            tiles += f"""
            <div class="col d-flex flex-column align-items-center justify-content-center py-3" style="gap:.4rem">
              <div class="stat-num text-{color}">{count}</div>
              <div class="text-muted" style="font-size:.72rem;line-height:1">{label}</div>
            </div>"""

        # Summary table rows
        table_rows = ""
        for e in r["entries"]:
            is_char = e.get("is_char", False)
            color = STATUS_COLOR.get(e["sc"], "secondary")
            badge = f'<span class="badge bg-{color} {"text-dark" if color in ("light","info") else ""}">{e["status"]}</span>'

            if is_char:
                # Characterization test: show warpage result instead of n/k and stat
                char_res = e.get("char_result", "") or "—"
                nk = "—"
                stat_cell = f' <small class="text-muted">{char_res}</small>'
            else:
                nk = f"{e['n']} / {e['k']}" if e["n"] is not None else "—"
                stat_cell = ""
            notes_td = f'<td class="text-muted small">{e["notes"]}</td>' if e["notes"] else '<td class="text-muted">—</td>'
            table_rows += f"""
            <tr>
              <td class="fw-semibold">{e['test']['name']}</td>
              <td class="text-muted small">{e['test']['standard']}</td>
              <td>{badge}{stat_cell}</td>
              <td class="text-muted small">{nk}</td>
              {notes_td}
            </tr>"""

        # Statistical details — pull project criteria from report dict
        _q_ltpd = r.get("qual_ltpd", 5.0)
        _q_conf = r.get("qual_confidence", 0.90)
        _q_rreq = r.get("qual_r_req", 0.95)
        _q_conf_lbl = f'{_q_conf*100:.0f}%'
        _q_rreq_lbl = f'{_q_rreq*100:.0f}%'
        _q_ltpd_lbl = f'{_q_ltpd:g}%'
        stat_entries = [e for e in r["entries"] if e["r_demo"] is not None]
        stat_section = ""
        if stat_entries and r.get("show_stats", True):
            srows = ""
            for e in stat_entries:
                ok  = e.get("effective_pass", e.get("stat_pass"))
                _is_ov = e.get("pass_override", "auto") != "auto"
                _ov_badge = (' <span class="badge text-bg-warning" style="font-size:.62rem">override</span>'
                             if _is_ov else "")
                _stat_note = (f' <span class="text-muted" style="font-size:.78rem">'
                              f'(auto: {"PASS" if e.get("stat_pass") else "FAIL"})</span>'
                              if _is_ov else "")
                srows += (
                    f'<tr><td class="fw-semibold">{e["test"]["name"]}</td>'
                    f'<td>{e["n"]}</td><td>{e["k"]}</td>'
                    f'<td>{e["r_demo"]*100:.2f}%</td>'
                    f'<td><span class="badge {"bg-success" if ok else "bg-danger"}">'
                    f'{"PASS" if ok else "FAIL"}</span>{_ov_badge}'
                    f' vs {_q_rreq_lbl} R @ {_q_conf_lbl} CL{_stat_note}</td></tr>'
                )
            stat_section = f"""
            <h6 class="text-muted mt-4 mb-2" style="font-size:.8rem;letter-spacing:.05em;text-transform:uppercase">Statistical Details ({_q_conf_lbl} confidence, {_q_rreq_lbl} reliability target — LTPD {_q_ltpd_lbl})</h6>
            <table class="table table-sm table-bordered">
              <thead class="tbl-header"><tr><th>Test</th><th>n</th><th>Failures</th><th>Demonstrated R</th><th>Result</th></tr></thead>
              <tbody>{srows}</tbody>
            </table>"""

        body = f"""
        <!-- Action bar -->
        <div class="d-flex align-items-center justify-content-between mb-4 no-print">
          <h4 class="mb-0" style="font-weight:300;letter-spacing:-.01em">Qualification Report</h4>
          <div class="d-flex gap-2">
            <a href="/report/pdf" class="btn btn-primary">Download PDF &darr;</a>
            <button onclick="window.print()" class="btn btn-outline-secondary">Print</button>
            <a href="/report" class="btn btn-outline-primary">New Report</a>
          </div>
        </div>

        <!-- Report body -->
        <div class="report-wrap">
          <div class="rpt-header">
            <div class="d-flex justify-content-between align-items-start">
              <div>
                <div class="rpt-title">Reliability Qualification Report</div>
              </div>
              <div class="rpt-sub text-end">{r['date']}</div>
            </div>
          </div>

          <!-- Metadata -->
          <div class="row g-3 mb-4">
            <div class="col-6 col-md-3">
              <div class="text-muted small">Customer</div>
              <div class="fw-semibold">{r['customer']}</div>
            </div>
            <div class="col-6 col-md-3">
              <div class="text-muted small">Product</div>
              <div class="fw-semibold">{r['product']}</div>
            </div>
            <div class="col-6 col-md-3">
              <div class="text-muted small">Part Type</div>
              <div class="fw-semibold">{r['part_label']}</div>
            </div>
            <div class="col-6 col-md-3">
              <div class="text-muted small">Prepared By</div>
              <div class="fw-semibold">{r['author']}</div>
            </div>
          </div>

          <!-- Part Description -->"""

        if r['part_type'] == 'ttv':
            pd = r.get('part_desc', {})
            si_thick = pd.get('si_thick', '50')
            bond_type = pd.get('bond_type', 'Ag Sinter')
            diamond_thick = pd.get('diamond_thick', '300')
            body += f"""
          <div style="background:var(--df-bg);border-left:3px solid var(--df-border);padding:.75rem 1rem;margin-bottom:1.5rem;font-size:.85rem">
            <div style="font-weight:600;margin-bottom:.5rem">Nanotest TTV10-NT20 Thermal Test Vehicle</div>
            <div style="color:var(--df-grey);margin-bottom:.25rem">25mm die package — flip-chip on PCB, SAC305 solder</div>
            <div style="color:var(--df-grey);margin-bottom:.5rem">Si die: {si_thick} µm | Bond: {bond_type} | SCD: {diamond_thick} µm</div>
            <div style="color:var(--df-grey);font-size:.8rem">16 RTD sensors (T1–T16) | 4 heater zones (S1–S4) | 6 hotspot cells (HS1–HS6)</div>
          </div>"""
        else:
            pd = r.get('part_desc', {})
            pn = pd.get('part_number', '—')
            pt_node = pd.get('part_tech', '—')
            pkg = pd.get('part_package', '—')
            body += f"""
          <div style="background:var(--df-bg);border-left:3px solid var(--df-border);padding:.75rem 1rem;margin-bottom:1.5rem;font-size:.85rem">
            <div style="font-weight:600;margin-bottom:.5rem">Active Device</div>
            <div style="color:var(--df-grey);margin-bottom:.25rem">Part Number: {pn}</div>
            <div style="color:var(--df-grey)">Package Type: {pkg}</div>
          </div>"""

        # Build "Tests Performed" rows
        _tc = r.get("test_conditions", {})
        def _cond_label(key):
            ckey = _tc.get(key, "")
            opts = _TEST_CONDITION_OPTIONS.get(key, [])
            lbl  = next((lbl for k, lbl in opts if k == ckey), ckey)
            return lbl or "—"
        def _dur_label(key):
            t_info = TESTS.get(key, {})
            return t_info.get("duration", "—")

        tp_rows = ""
        for e in r["entries"]:
            tp_rows += (
                f'<tr>'
                f'<td class="fw-semibold" style="white-space:nowrap">{e["test"]["name"]}</td>'
                f'<td class="text-muted small">{_cond_label(e["key"])}</td>'
                f'<td class="text-muted small">{_dur_label(e["key"])}</td>'
                f'</tr>'
            )

        body += f"""
          <!-- Precond note -->
          <div class="precond-bar mb-4">
            <strong>Precursor for all tests:</strong>
            PC ({PRECOND['full_name']}; {PRECOND['standard']}) &mdash;
            {PRECOND['condition']} &mdash; {PRECOND['duration']}
          </div>

          <!-- Tests Performed -->
          <h6 class="text-muted mb-2" style="font-size:.8rem;letter-spacing:.05em;text-transform:uppercase">Tests Performed</h6>
          <table class="table table-sm table-bordered mb-4">
            <thead class="tbl-header">
              <tr><th>Test</th><th>Condition</th><th>Duration</th></tr>
            </thead>
            <tbody>{tp_rows}</tbody>
          </table>

          {correction_banner}

          <!-- Summary counts -->
          <div class="row g-0 mb-4 text-center border rounded overflow-hidden">
            {tiles}
          </div>

          <!-- Summary table -->
          <table class="table table-bordered table-hover table-sm mb-2">
            <thead class="tbl-header">
              <tr><th>Test</th><th>Standard</th><th>Status</th><th>n / fails</th><th>Notes</th></tr>
            </thead>
            <tbody>{table_rows}</tbody>
          </table>

          {stat_section}

          <!-- Sample Records Tables (TTV and Die) -->"""

        samples = r.get("samples", {})
        _rpt_is_ttv = r['part_type'] == 'ttv'
        if r['part_type'] in ('ttv', 'die') and samples:
            for key in samples:
                test_samples = samples[key]
                if not test_samples:
                    continue

                test_name = None
                for e in r["entries"]:
                    if e["key"] == key:
                        test_name = e["test"]["name"]
                        break

                if not test_name:
                    continue

                _extra_th = """
                <th style="min-width:80px">Thermal</th>
                <th style="min-width:80px">Func</th>""" if _rpt_is_ttv else ""

                body += f"""
          <h6 class="text-muted mt-4 mb-2" style="font-size:.8rem;letter-spacing:.05em;text-transform:uppercase">{test_name} — Sample Records</h6>
          <table class="table table-sm table-bordered mb-3">
            <thead class="tbl-header">
              <tr>
                <th style="min-width:100px">Sample</th>
                <th style="min-width:120px">CSAM Before PC</th>
                <th style="min-width:120px">CSAM After PC</th>
                <th style="min-width:120px">CSAM After Test</th>
                <th style="min-width:100px">CSAM Status</th>{_extra_th}
                <th style="min-width:80px">Overall</th>
              </tr>
            </thead>
            <tbody>"""

                for sample in test_samples:
                    sample_id = sample["id"]
                    csam_bpc = sample["csam_bpc"]
                    csam_apc = sample["csam_apc"]
                    csam_atst = sample["csam_atst"]
                    csam_status = sample["csam_status"]
                    csam_badge = sample["csam_badge"]
                    thermal = sample["thermal"]
                    func = sample["func"]

                    # Determine overall status
                    if csam_status == "Rejected — pre-PC < 95%":
                        overall = "Rejected"
                        overall_badge = "danger"
                        show_dash = True
                    else:
                        show_dash = False
                        if _rpt_is_ttv:
                            passed = csam_status == "Pass" and thermal == "pass" and func == "pass"
                        else:
                            passed = csam_status == "Pass"
                        overall = "Pass" if passed else "Fail"
                        overall_badge = "success" if passed else "danger"

                    # RTD failure note (TTV thermal fail only)
                    _failed_rtds = sample.get("failed_rtds", [])
                    if _rpt_is_ttv and not show_dash and thermal == "fail" and _failed_rtds:
                        _rtd_note = (f' <small class="text-muted" style="font-size:.8em;white-space:nowrap">'
                                     f'{len(_failed_rtds)}/16 RTD</small>')
                    else:
                        _rtd_note = ""

                    _extra_td = f"""
                <td class="text-muted small">{"—" if show_dash else thermal.upper()}</td>
                <td class="text-muted small">{"—" if show_dash else func.upper()}</td>""" if _rpt_is_ttv else ""

                    body += f"""
              <tr>
                <td class="fw-semibold">{sample_id}</td>
                <td class="text-muted small">{f"{csam_bpc:.1f}%" if csam_bpc is not None else "—"}</td>
                <td class="text-muted small">{f"{csam_apc:.1f}%" if csam_apc is not None else "—"}</td>
                <td class="text-muted small">{f"{csam_atst:.1f}%" if csam_atst is not None else "—"}</td>
                <td><span class="badge bg-{csam_badge}">{csam_status}</span></td>{_extra_td}
                <td style="white-space:nowrap"><span class="badge bg-{overall_badge}">{overall.upper()}</span>{_rtd_note}</td>
              </tr>"""

                body += f"""
            </tbody>
          </table>"""

        # CSAM Images Appendix
        has_images = False
        for key in samples:
            for sample in samples.get(key, []):
                if sample["img_bpc"] or sample["img_apc"] or sample["img_atst"]:
                    has_images = True
                    break
            if has_images:
                break

        if r['part_type'] == 'ttv' and has_images:
            body += f"""
          <h6 class="text-muted mt-5 mb-3" style="font-size:.8rem;letter-spacing:.05em;text-transform:uppercase">Appendix — CSAM Images</h6>"""

            for key in samples:
                test_samples = samples[key]
                test_name = None
                for e in r["entries"]:
                    if e["key"] == key:
                        test_name = e["test"]["name"]
                        break

                if not test_name:
                    continue

                # Check if any sample in this test has images
                test_has_images = any(s["img_bpc"] or s["img_apc"] or s["img_atst"] for s in test_samples)
                if not test_has_images:
                    continue

                body += f"""
          <div style="margin-top:1.5rem;page-break-inside:avoid">
            <div style="font-size:.85rem;font-weight:600;margin-bottom:.75rem;color:var(--df-black)">{test_name}</div>"""

                for sample in test_samples:
                    sample_id = sample["id"]
                    img_bpc = sample["img_bpc"]
                    img_apc = sample["img_apc"]
                    img_atst = sample["img_atst"]

                    if not (img_bpc or img_apc or img_atst):
                        continue

                    body += f"""
            <div style="margin-bottom:1.5rem">
              <div style="font-size:.8rem;color:var(--df-grey);margin-bottom:.5rem;font-weight:500">{sample_id}</div>
              <div style="display:flex;gap:1rem;align-items:flex-start">"""

                    for label, img_uri, csam_val in [("Before PC", img_bpc, sample["csam_bpc"]), ("After PC", img_apc, sample["csam_apc"]), ("After Test", img_atst, sample["csam_atst"])]:
                        if img_uri:
                            csam_text = f"{csam_val:.1f}%" if csam_val is not None else ""
                            body += f"""
                <div style="flex:1;text-align:center">
                  <img src="{img_uri}" style="max-width:250px;border:1px solid #e0e0e0;margin-bottom:.5rem">
                  <div style="font-size:.75rem;color:var(--df-grey)">{label}</div>
                  <div style="font-size:.7rem;color:var(--df-grey)">{csam_text}</div>
                </div>"""

                    body += f"""
              </div>
            </div>"""

                body += f"""
          </div>"""

        body += f"""
          <div class="text-muted text-end mt-4" style="font-size:.75rem">
            Generated by Package Reliability Qualification Suite &mdash; {r['date']}
          </div>
        </div>"""
        self.emit(body, "Qual Report", "report")


# ── /report/pdf ───────────────────────────────────────────────────────────────

class ReportPdfHandler(Base):
    def get(self):
        import traceback
        _, s = self.sess()
        rpt = s.get("last_report")
        if not rpt:
            self.redirect("/report")
            return
        try:
            pdf_bytes = _make_pdf(rpt)
        except Exception:
            err = traceback.format_exc()
            self.set_status(500)
            self.set_header("Content-Type", "text/plain")
            self.finish(f"PDF generation error:\n\n{err}")
            return
        fname = f"QualReport_{rpt['customer'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        self.set_header("Content-Type", "application/pdf")
        self.set_header("Content-Disposition", f'attachment; filename="{fname}"')
        self.finish(pdf_bytes)


# ── PDF generation (ReportLab) ────────────────────────────────────────────────

_NAVY   = colors.HexColor("#0f2744")
_LIGHT  = colors.HexColor("#e8f0fe")
_PASS_C = colors.HexColor("#d1e7dd")
_FAIL_C = colors.HexColor("#f8d7da")
_IP_C   = colors.HexColor("#fff3cd")
_NS_C   = colors.HexColor("#e9ecef")

_STATUS_BG = {"PASS": _PASS_C, "FAIL": _FAIL_C,
              "IN PROGRESS": _IP_C, "NOT STARTED": _NS_C, "N/A": _NS_C}


def _b64_to_rl_image(b64_uri: str, w: float, h: float):
    """Convert a base64 data URI to a ReportLab Image flowable."""
    try:
        header, data = b64_uri.split(",", 1)
        img_bytes = base64.b64decode(data)
        return RLImage(io.BytesIO(img_bytes), width=w, height=h)
    except Exception:
        return None


def _make_pdf(r: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=2*cm, bottomMargin=2*cm,
                             leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()

    # Custom paragraph styles
    def S(name, **kw):
        base = kw.pop("parent", "Normal")
        return ParagraphStyle(name, parent=styles[base], **kw)

    hdr_title  = S("HT", fontSize=16, fontName="Helvetica-Bold",
                   textColor=colors.white, spaceAfter=7, leading=20)
    hdr_sub    = S("HS", fontSize=9,  textColor=colors.HexColor("#aac4e0"),
                   leading=13, spaceBefore=2)
    meta_label = S("ML", fontSize=8,  textColor=colors.grey)
    meta_val   = S("MV", fontSize=10, fontName="Helvetica-Bold")
    body_s     = S("B",  fontSize=9)
    small_s    = S("Sm", fontSize=8,  textColor=colors.grey)
    section_h  = S("SH", fontSize=10, fontName="Helvetica-Bold",
                   textColor=_NAVY, spaceBefore=12, spaceAfter=4)
    sub_h      = S("SUB", fontSize=9, fontName="Helvetica-Bold",
                   textColor=colors.HexColor("#444444"), spaceBefore=8, spaceAfter=2)
    img_cap    = S("IC", fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    W = 17 * cm   # usable width

    story = []

    # ── Helper: page-wide navy header bar (reused on each page break) ──────────
    def _hdr_bar():
        d = [[
            [Paragraph("RELIABILITY QUALIFICATION REPORT", hdr_title)],
            [Paragraph(r["date"], hdr_sub)],
        ]]
        t = Table(d, colWidths=[12*cm, 5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), _NAVY),
            ("TOPPADDING",   (0,0), (-1,-1), 16),
            ("BOTTOMPADDING",(0,0), (-1,-1), 16),
            ("LEFTPADDING",  (0,0), (-1,-1), 16),
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
            ("ALIGN",        (1,0), (1,0),   "RIGHT"),
        ]))
        return t

    # ── 1. Header bar ─────────────────────────────────────────────────────────
    story.append(_hdr_bar())
    story.append(Spacer(1, 0.5*cm))

    # ── 2. Metadata block ─────────────────────────────────────────────────────
    meta_rows = [
        [Paragraph("Customer",   meta_label), Paragraph(r["customer"],  meta_val),
         Paragraph("Product",    meta_label), Paragraph(r["product"],   meta_val)],
        [Paragraph("Part Type",  meta_label), Paragraph(r["part_label"],meta_val),
         Paragraph("Prepared By",meta_label), Paragraph(r["author"],    meta_val)],
    ]
    # Part description rows
    pd = r.get("part_desc", {})
    if r.get("part_type") == "ttv":
        meta_rows.append([
            Paragraph("Si Thickness",      meta_label), Paragraph(pd.get("si_thick","—") + " µm", meta_val),
            Paragraph("Bond Type",         meta_label), Paragraph(pd.get("bond_type","—"),         meta_val),
        ])
        meta_rows.append([
            Paragraph("SCD Thickness",     meta_label), Paragraph(pd.get("diamond_thick","—") + " µm", meta_val),
            Paragraph("",                  meta_label), Paragraph("",                                   meta_val),
        ])
    else:
        meta_rows.append([
            Paragraph("Part Number",  meta_label), Paragraph(pd.get("part_number","—"), meta_val),
            Paragraph("Package",      meta_label), Paragraph(pd.get("part_package","—"), meta_val),
        ])
    meta_tbl = Table(meta_rows, colWidths=[2.5*cm, 6*cm, 2.5*cm, 6*cm])
    meta_tbl.setStyle(TableStyle([
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width=W, color=colors.HexColor("#dee2e6")))
    story.append(Spacer(1, 0.4*cm))

    # ── 3. Precond callout ────────────────────────────────────────────────────
    precond_data = [[
        Paragraph(
            f"<b>Universal Precursor for all tests:</b> "
            f"PC ({PRECOND['full_name']}; {PRECOND['standard']}) — "
            f"{PRECOND['condition']} — {PRECOND['duration']}",
            S("PC", fontSize=8, textColor=colors.HexColor("#5a4000"))
        )
    ]]
    precond_tbl = Table(precond_data, colWidths=[W])
    precond_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#fff8e1")),
        ("LEFTPADDING",  (0,0),(-1,-1), 10),
        ("TOPPADDING",   (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(precond_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── 3b. Tests Performed ───────────────────────────────────────────────────
    _tc = r.get("test_conditions", {})
    def _pdf_cond_label(key: str) -> str:
        ckey = _tc.get(key, "")
        opts = _TEST_CONDITION_OPTIONS.get(key, [])
        lbl  = next((lb for k, lb in opts if k == ckey), ckey)
        return lbl or "—"
    def _pdf_dur_label(key: str) -> str:
        return TESTS.get(key, {}).get("duration", "—")

    tp_hdr = [
        Paragraph("<b>Test</b>",      S("TPH", fontSize=7.5, textColor=colors.white)),
        Paragraph("<b>Condition</b>", S("TPH", fontSize=7.5, textColor=colors.white)),
        Paragraph("<b>Duration</b>",  S("TPH", fontSize=7.5, textColor=colors.white)),
    ]
    tp_data = [tp_hdr]
    for e in r["entries"]:
        key = e["key"]
        tp_data.append([
            Paragraph(e["test"]["name"],      S("TPD", fontSize=7.5)),
            Paragraph(_pdf_cond_label(key),   S("TPD", fontSize=7)),
            Paragraph(_pdf_dur_label(key),    S("TPD", fontSize=7)),
        ])
    tp_tbl = Table(tp_data, colWidths=[4*cm, 8*cm, 5.5*cm])
    tp_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  _NAVY),
        ("BACKGROUND",    (0, 1), (-1,-1),  colors.white),
        ("ROWBACKGROUNDS",(0, 1), (-1,-1),  [colors.white, colors.HexColor("#f8f9fa")]),
        ("TOPPADDING",    (0, 0), (-1,-1),  5),
        ("BOTTOMPADDING", (0, 0), (-1,-1),  5),
        ("LEFTPADDING",   (0, 0), (-1,-1),  6),
        ("RIGHTPADDING",  (0, 0), (-1,-1),  6),
        ("GRID",          (0, 0), (-1,-1),  0.25, colors.HexColor("#dee2e6")),
        ("VALIGN",        (0, 0), (-1,-1),  "TOP"),
    ]))
    story.append(Paragraph(
        "Tests Performed",
        S("TPSEC", fontSize=9, fontName="Helvetica-Bold", textColor=_NAVY, spaceAfter=4)
    ))
    story.append(tp_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── 4. Summary counts ─────────────────────────────────────────────────────
    counts = [
        (r["n_pass"], "PASS",        colors.HexColor("#198754")),
        (r["n_fail"], "FAIL",        colors.HexColor("#dc3545")),
        (r["n_ip"],   "IN PROGRESS", colors.HexColor("#fd7e14")),
        (r["n_ns"],   "NOT STARTED", colors.grey),
        (r["total"],  "TOTAL",       _NAVY),
    ]
    col_w = W / len(counts)
    # Two separate rows: numbers on top, labels below
    num_row   = [Paragraph(str(c),  S(f"CN{i}", fontSize=22, fontName="Helvetica-Bold",
                                       textColor=color, alignment=TA_CENTER))
                 for i, (c, label, color) in enumerate(counts)]
    label_row = [Paragraph(label, S(f"CL{i}", fontSize=7, textColor=colors.grey,
                                     alignment=TA_CENTER))
                 for i, (c, label, color) in enumerate(counts)]
    cnt_tbl = Table([num_row, label_row],
                    colWidths=[col_w] * len(counts),
                    rowHeights=[1.1*cm, 0.55*cm])
    cnt_tbl.setStyle(TableStyle([
        ("TOPPADDING",    (0,0), (-1,0),  6),
        ("BOTTOMPADDING", (0,0), (-1,0),  6),
        ("TOPPADDING",    (0,1), (-1,1),  2),
        ("BOTTOMPADDING", (0,1), (-1,1),  6),
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#f8f9fa")),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#dee2e6")),
        ("VALIGN",        (0,0), (-1,0),  "MIDDLE"),
        ("VALIGN",        (0,1), (-1,1),  "TOP"),
    ]))
    story.append(cnt_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── 5. Summary table ──────────────────────────────────────────────────────
    story.append(Paragraph("Test Results Summary", section_h))
    th_s = S("TH", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white)
    tbl_data = [[
        Paragraph("Test",      th_s),
        Paragraph("Standard",  th_s),
        Paragraph("Status",    th_s),
        Paragraph("n / fails", th_s),
        Paragraph("Notes",     th_s),
    ]]
    row_style_cmds = [
        ("BACKGROUND",    (0,0), (-1,0), _NAVY),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("FONTSIZE",      (0,1), (-1,-1), 9),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]
    for i, e in enumerate(r["entries"], start=1):
        if e.get("is_char"):
            nk = e.get("char_result") or "—"
        elif e["n"] is not None:
            nk = f"{e['n']} / {e['k']}"
        else:
            nk = "—"
        bg = _STATUS_BG.get(e["status"], _NS_C)
        tbl_data.append([
            Paragraph(e["test"]["name"],     body_s),
            Paragraph(e["test"]["standard"], small_s),
            Paragraph(e["status"],           body_s),
            Paragraph(nk,                    body_s),
            Paragraph(e["notes"] or "—",     small_s),
        ])
        row_style_cmds.append(("BACKGROUND", (2, i), (2, i), bg))

    sum_tbl = Table(tbl_data, colWidths=[5*cm, 2.8*cm, 3.5*cm, 2.2*cm, 3.5*cm])
    sum_tbl.setStyle(TableStyle(row_style_cmds))
    story.append(sum_tbl)

    # ── 6. Statistical details (only when show_stats=True) ───────────────────
    _pdf_ltpd = r.get("qual_ltpd", 5.0)
    _pdf_conf = r.get("qual_confidence", 0.90)
    _pdf_rreq = r.get("qual_r_req", 0.95)
    _pdf_conf_lbl = f'{_pdf_conf*100:.0f}%'
    _pdf_rreq_lbl = f'{_pdf_rreq*100:.0f}%'
    _pdf_ltpd_lbl = f'{_pdf_ltpd:g}%'
    stat_entries = [e for e in r["entries"] if e.get("r_demo") is not None]
    if stat_entries and r.get("show_stats", True):
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f"Statistical Details  ({_pdf_conf_lbl} confidence level, {_pdf_rreq_lbl} reliability target — LTPD {_pdf_ltpd_lbl})", section_h))
        s_data = [[
            Paragraph(h, S(f"SH{j}", fontSize=9, fontName="Helvetica-Bold",
                           textColor=colors.white))
            for j, h in enumerate(["Test", "n", "Failures",
                                    "Demonstrated R", f"vs {_pdf_rreq_lbl} Target"])
        ]]
        s_cmds = [
            ("BACKGROUND",    (0,0), (-1,0), _NAVY),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("FONTSIZE",      (0,1), (-1,-1), 9),
        ]
        for i, e in enumerate(stat_entries, start=1):
            ok  = e.get("effective_pass", e.get("stat_pass"))
            bg  = _PASS_C if ok else _FAIL_C
            _is_ov = e.get("pass_override", "auto") != "auto"
            res = ("PASS*" if _is_ov else "PASS") if ok else ("FAIL*" if _is_ov else "FAIL")
            s_data.append([
                Paragraph(e["test"]["name"],          body_s),
                Paragraph(str(e["n"]),                body_s),
                Paragraph(str(e["k"]),                body_s),
                Paragraph(f"{e['r_demo']*100:.2f}%",  body_s),
                Paragraph(res,                        body_s),
            ])
            s_cmds.append(("BACKGROUND", (4, i), (4, i), bg))
        stat_tbl = Table(s_data, colWidths=[5.5*cm, 2*cm, 2*cm, 4*cm, 3.5*cm])
        stat_tbl.setStyle(TableStyle(s_cmds))
        story.append(stat_tbl)

    # ── 7. Sample records ─────────────────────────────────────────────────────
    all_samples = r.get("samples", {})
    tests_with_samples = [(e, all_samples[e["key"]])
                          for e in r["entries"]
                          if e["key"] in all_samples and all_samples[e["key"]]]
    if tests_with_samples:
        story.append(PageBreak())
        story.append(_hdr_bar())
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Sample Records", section_h))

        sr_th = S("SRTH", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white)
        sr_td = S("SRTD", fontSize=8)
        sr_sm = S("SRSM", fontSize=7, textColor=colors.grey)

        for entry, slist in tests_with_samples:
            story.append(Paragraph(entry["test"]["name"], sub_h))

            _pdf_is_ttv = r.get("part_type") == "ttv"
            # Table header
            _sr_hdr = [
                Paragraph("Sample ID",          sr_th),
                Paragraph("CSAM Before PC (%)", sr_th),
                Paragraph("CSAM After PC (%)",  sr_th),
                Paragraph("CSAM After Test (%)",sr_th),
                Paragraph("CSAM Status",        sr_th),
            ]
            if _pdf_is_ttv:
                _sr_hdr += [
                    Paragraph("Thermal",        sr_th),
                    Paragraph("Failed Sensors", sr_th),
                    Paragraph("Func.",          sr_th),
                    Paragraph("Overall",        sr_th),
                ]
            sr_data = [_sr_hdr]
            sr_cmds = [
                ("BACKGROUND",    (0,0), (-1,0), _NAVY),
                ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#dee2e6")),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING",   (0,0), (-1,-1), 5),
                ("FONTSIZE",      (0,1), (-1,-1), 8),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ]
            for ri, s in enumerate(slist, start=1):
                def _pct(v):
                    return f"{v:.1f}%" if v is not None else "—"
                def _tf(v):
                    return v.upper() if v else "—"
                csam_ok = s.get("csam_status","") == "Pass"

                _row = [
                    Paragraph(s["id"],                         sr_td),
                    Paragraph(_pct(s.get("csam_bpc")),         sr_td),
                    Paragraph(_pct(s.get("csam_apc")),         sr_td),
                    Paragraph(_pct(s.get("csam_atst")),        sr_td),
                    Paragraph(s.get("csam_status", "—"),       sr_td),
                ]
                if _pdf_is_ttv:
                    therm_ok = s.get("thermal","") == "pass"
                    func_ok  = s.get("func","") == "pass"
                    _failed_rtds_pdf = s.get("failed_rtds", [])
                    rtds = ", ".join(_failed_rtds_pdf) or "None"
                    _overall_ok = csam_ok and therm_ok and func_ok
                    if not therm_ok and _failed_rtds_pdf:
                        _overall_txt = f"FAIL  {len(_failed_rtds_pdf)}/16 RTD"
                    else:
                        _overall_txt = "PASS" if _overall_ok else "FAIL"
                    _row += [
                        Paragraph(_tf(s.get("thermal")), sr_td),
                        Paragraph(rtds,                  sr_sm),
                        Paragraph(_tf(s.get("func")),    sr_td),
                        Paragraph(_overall_txt,          sr_td),
                    ]
                    sr_cmds += [
                        ("BACKGROUND", (5, ri), (5, ri), _PASS_C if therm_ok   else _FAIL_C),
                        ("BACKGROUND", (7, ri), (7, ri), _PASS_C if func_ok    else _FAIL_C),
                        ("BACKGROUND", (8, ri), (8, ri), _PASS_C if _overall_ok else _FAIL_C),
                    ]
                sr_data.append(_row)
                sr_cmds.append(("BACKGROUND", (4, ri), (4, ri), _PASS_C if csam_ok else _FAIL_C))

            if _pdf_is_ttv:
                _sr_widths = [2.0*cm, 1.9*cm, 1.9*cm, 1.9*cm, 2.8*cm, 1.4*cm, 1.6*cm, 1.4*cm, 2.1*cm]
            else:
                _sr_widths = [2.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 6.9*cm]
            sr_tbl = Table(sr_data, colWidths=_sr_widths)
            sr_tbl.setStyle(TableStyle(sr_cmds))
            story.append(sr_tbl)
            story.append(Spacer(1, 0.4*cm))

    # ── 8. CSAM Image Registry ────────────────────────────────────────────────
    csam_entries = []
    for entry, slist in tests_with_samples:
        for s in slist:
            if s.get("img_bpc") or s.get("img_apc") or s.get("img_atst"):
                csam_entries.append((entry, s))

    if csam_entries:
        story.append(PageBreak())
        story.append(_hdr_bar())
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("CSAM Image Registry", section_h))
        story.append(Paragraph(
            "Acoustic microscopy images for each sample — Before Preconditioning | "
            "After Preconditioning | After Test.  Bonded area threshold: ≥95%.",
            S("CSUB", fontSize=8, textColor=colors.grey, spaceAfter=8)
        ))

        SIDE_W  = 5.0 * cm
        IMG_COL = (W - SIDE_W) / 3       # exactly fills remaining width
        IMG_W   = IMG_COL - 0.3 * cm     # image slightly smaller than cell
        IMG_H   = IMG_W                  # keep square

        for entry, s in csam_entries:
            # Section label: Test — Sample ID
            story.append(Paragraph(
                f"{entry['test']['name']}  —  {s['id']}",
                sub_h
            ))

            cells = []
            captions = []
            for label, key in [("Before PC", "img_bpc"),
                                ("After PC",  "img_apc"),
                                ("After Test","img_atst")]:
                uri = s.get(key, "")
                if uri:
                    img = _b64_to_rl_image(uri, IMG_W, IMG_H)
                else:
                    img = Paragraph("No image", img_cap)
                cells.append(img or Paragraph("—", img_cap))
                captions.append(Paragraph(label, img_cap))

            img_tbl = Table(
                [cells, captions],
                colWidths=[IMG_COL, IMG_COL, IMG_COL],
            )
            img_tbl.setStyle(TableStyle([
                ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                ("VALIGN",        (0,0), (-1,-1), "TOP"),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 2),
                ("LEFTPADDING",   (0,0), (-1,-1), 4),
                ("RIGHTPADDING",  (0,0), (-1,-1), 4),
                ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#dee2e6")),
            ]))

            # Metadata sidebar
            csam_ok = s.get("csam_status","") == "Pass"
            status_color = colors.HexColor("#198754") if csam_ok else colors.HexColor("#dc3545")
            side_data = [
                [Paragraph("CSAM Status", S("SL", fontSize=7, textColor=colors.grey)),
                 Paragraph(s.get("csam_status","—"),
                           S("SV", fontSize=8, fontName="Helvetica-Bold",
                             textColor=status_color))],
                [Paragraph("Before PC",   S("SL", fontSize=7, textColor=colors.grey)),
                 Paragraph(f"{s['csam_bpc']:.1f}%" if s.get("csam_bpc") is not None else "—",
                           S("SV2", fontSize=8))],
                [Paragraph("After PC",    S("SL2", fontSize=7, textColor=colors.grey)),
                 Paragraph(f"{s['csam_apc']:.1f}%" if s.get("csam_apc") is not None else "—",
                           S("SV3", fontSize=8))],
                [Paragraph("After Test",  S("SL3", fontSize=7, textColor=colors.grey)),
                 Paragraph(f"{s['csam_atst']:.1f}%" if s.get("csam_atst") is not None else "—",
                           S("SV4", fontSize=8))],
            ]
            side_tbl = Table(side_data, colWidths=[2.2*cm, 2.8*cm])
            side_tbl.setStyle(TableStyle([
                ("TOPPADDING",    (0,0),(-1,-1), 3),
                ("BOTTOMPADDING", (0,0),(-1,-1), 3),
                ("LEFTPADDING",   (0,0),(-1,-1), 4),
                ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#dee2e6")),
                ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor("#f8f9fa")),
            ]))

            row_tbl = Table([[img_tbl, side_tbl]],
                             colWidths=[3*IMG_COL, SIDE_W])
            row_tbl.setStyle(TableStyle([
                ("VALIGN", (0,0),(-1,-1), "TOP"),
                ("LEFTPADDING",  (0,0),(-1,-1), 0),
                ("RIGHTPADDING", (0,0),(-1,-1), 0),
                ("TOPPADDING",   (0,0),(-1,-1), 0),
                ("BOTTOMPADDING",(0,0),(-1,-1), 0),
            ]))
            story.append(row_tbl)
            story.append(Spacer(1, 0.6*cm))

    # ── 9. Footer ─────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width=W, color=colors.HexColor("#dee2e6")))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Package Reliability Qualification Suite — {r['date']}",
        S("FTR", fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── Project handlers ──────────────────────────────────────────────────────────

def _fmt_dt(iso: str) -> str:
    """Format ISO datetime string to readable form."""
    try:
        return datetime.fromisoformat(iso).strftime("%b %d, %Y")
    except Exception:
        return iso or "—"

class ProjectListHandler(Base):
    def get(self):
        projects = _db.list_projects()
        if projects:
            rows = ""
            for p in projects:
                status_color = {"active": "success", "complete": "info", "archived": "secondary"}.get(p["status"], "secondary")
                _pt_badge = {"ttv": "TTV", "die": "Die"}.get(p["part_type"], "Active")
                rows += f"""
                <tr>
                  <td><a href="/projects/{p['id']}" class="fw-semibold text-decoration-none"
                         style="color:var(--df-black)">{p['name']}</a></td>
                  <td class="text-muted" style="font-size:.83rem">{p['description'] or '—'}</td>
                  <td><span class="badge" style="font-size:.72rem;background:var(--df-bg);color:var(--df-mid);border:1px solid var(--df-border)">
                    {_pt_badge}</span></td>
                  <td><span class="badge bg-{status_color}">{p['status'].capitalize()}</span></td>
                  <td class="text-muted" style="font-size:.82rem">{_fmt_dt(p['updated_at'])}</td>
                  <td class="text-muted" style="font-size:.82rem">{_fmt_dt(p['created_at'])}</td>
                  <td>
                    <a href="/projects/{p['id']}" class="btn btn-sm btn-outline-secondary">Open</a>
                    <button class="btn btn-sm btn-outline-secondary ms-1"
                            style="color:var(--df-grey)"
                            onclick="confirmDelete({p['id']}, '{p['name'].replace(chr(39), chr(92)+chr(39))}')">
                      Delete
                    </button>
                  </td>
                </tr>"""
            table = f"""
            <div class="card shadow-sm">
              <table class="table table-hover mb-0">
                <thead class="tbl-header">
                  <tr><th>Project</th><th>Description</th><th>Type</th><th>Status</th>
                      <th>Last Updated</th><th>Created</th><th></th></tr>
                </thead>
                <tbody>{rows}</tbody>
              </table>
            </div>"""
        else:
            table = """<div class="card text-center py-5 text-muted shadow-sm">
              <i class="bi bi-folder2-open fs-1 mb-3"></i>
              <p class="mb-0">No projects yet — create your first one below.</p>
            </div>"""

        body = f"""
        <!-- Delete-project confirmation modal -->
        <div class="modal fade" id="deleteModal" tabindex="-1">
          <div class="modal-dialog modal-dialog-centered">
            <div class="modal-content">
              <div class="modal-header border-0 pb-0">
                <h6 class="modal-title text-danger">
                  <i class="bi bi-exclamation-triangle me-2"></i>Delete Project
                </h6>
                <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
              </div>
              <div class="modal-body pt-2">
                <p style="font-size:.9rem" class="mb-1">You are about to permanently delete:</p>
                <p id="del-project-name" class="fw-semibold mb-3" style="font-size:1rem"></p>
                <div class="alert alert-danger py-2 mb-0" style="font-size:.83rem">
                  This will delete all sample data, pass/fail records, schedule tasks, and reports
                  for this project. <strong>This cannot be undone.</strong>
                </div>
              </div>
              <div class="modal-footer">
                <button type="button" class="btn btn-sm btn-outline-secondary"
                        data-bs-dismiss="modal">Cancel</button>
                <form id="deleteForm" method="post" class="d-inline">
                  <button type="submit" class="btn btn-sm btn-danger">
                    Yes, delete permanently
                  </button>
                </form>
              </div>
            </div>
          </div>
        </div>

        <div class="d-flex align-items-center justify-content-between mb-4">
          <h4 class="mb-0" style="font-weight:300">Projects</h4>
        </div>
        {table}
        <div class="card mt-4 shadow-sm" data-admin-gate="show" style="display:none">
          <div class="card-df"><h6 class="mb-0">New Project</h6></div>
          <div class="card-body p-4">
            <form method="post" action="/projects/new">
              <div class="row g-3">
                <div class="col-md-4">
                  <label class="form-label">Project Name <span class="text-danger">*</span></label>
                  <input type="text" class="form-control" name="name" required placeholder="e.g. TTV Phase 1 — Cu TCB">
                </div>
                <div class="col-md-4">
                  <label class="form-label">Description</label>
                  <input type="text" class="form-control" name="description" placeholder="Optional short description">
                </div>
                <div class="col-md-2">
                  <label class="form-label">Device Type</label>
                  <select class="form-select" name="part_type">
                    <option value="ttv">TTV</option>
                    <option value="die">Die</option>
                    <option value="active">Active Device</option>
                  </select>
                </div>
                <div class="col-md-2 d-flex align-items-end">
                  <button type="submit" class="btn btn-primary w-100">
                    <i class="bi bi-plus-lg me-1"></i>Create Project
                  </button>
                </div>
              </div>
            </form>
          </div>
        </div>
        <script>
        function confirmDelete(pid, name) {{
          document.getElementById('del-project-name').textContent = name;
          document.getElementById('deleteForm').action = '/projects/' + pid + '/delete';
          bootstrap.Modal.getOrCreateInstance(document.getElementById('deleteModal')).show();
        }}
        </script>"""
        self.emit(body, "Projects", active="projects")

    def post(self):
        # Redirect to /projects/new handler
        self.redirect("/projects/new")

class ProjectNewHandler(Base):
    def post(self):
        name        = self.get_argument("name", "").strip()
        description = self.get_argument("description", "").strip()
        part_type   = self.get_argument("part_type", "ttv")
        if not name:
            self.redirect("/projects")
            return
        pid = _db.create_project(name, description, part_type)
        # Sync part_type to session for consistency
        _, s = self.sess()
        s["part_type"] = part_type
        self.redirect(f"/projects/{pid}")

class ProjectDeleteHandler(Base):
    def post(self, pid):
        _db.delete_project(int(pid))
        self.redirect("/projects")

class ProjectDetailHandler(Base):
    def get(self, pid):
        pid = int(pid)
        p   = _db.get_project(pid)
        if not p:
            self.send_error(404); return
        meta  = _db.get_meta(pid)
        ss    = _db.get_sample_size(pid)
        pf    = _db.get_pass_fail(pid)
        samps = _db.get_samples(pid)

        # Quick stats
        n_tests_with_data = len([k for k, v in samps.items() if v])
        pf_verdict = ""
        if pf.get("n"):
            from jedec_calc import pass_fail as _pff, demonstrated_reliability
            try:
                n_v  = int(pf["n"])
                r_req = 1.0 - float(pf["ltpd_pct"]) / 100.0
                passed, r_demo = _pff(n_v, int(pf["failures"]), float(pf["confidence"]), r_req)
                pf_verdict = f'<span class="badge bg-{"success" if passed else "danger"} ms-2">{"PASS" if passed else "FAIL"}</span>'
            except Exception:
                pass

        status_opts = "".join(
            f'<option value="{v}" {"selected" if p["status"]==v else ""}>{l}</option>'
            for v, l in [("active","Active"),("complete","Complete"),("archived","Archived")]
        )

        _pt_str      = {"ttv": "TTV", "die": "Die"}.get(p["part_type"], "Active Device")
        meta_device  = meta.get("device_name","")
        meta_pkg     = meta.get("device_pkg","")
        meta_bond    = meta.get("bond_type","")
        meta_eng     = meta.get("engineer","")
        meta_lot     = meta.get("lot_id","")
        meta_notes   = meta.get("notes","")

        body = f"""
        <div class="row g-4">
          <div class="col-lg-8">
            <div class="card shadow-sm mb-4">
              <div class="card-df d-flex align-items-center justify-content-between">
                <h6 class="mb-0">Project Details</h6>
                <button class="btn btn-sm btn-outline-secondary" data-admin-gate="show" style="display:none"
                        onclick="document.getElementById('edit-form').classList.toggle('d-none')">
                  <i class="bi bi-pencil me-1"></i>Edit
                </button>
              </div>
              <div class="card-body p-4">
                <!-- Read-only view -->
                <div id="view-details">
                  <div class="row g-3" style="font-size:.87rem">
                    <div class="col-12"><span class="text-muted d-block">Project Name</span><strong>{p['name']}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Device / Part</span><strong>{meta_device or '—'}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Package</span><strong>{meta_pkg or '—'}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Bond Type</span><strong>{meta_bond or '—'}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Engineer</span><strong>{meta_eng or '—'}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Lot / Wafer ID</span><strong>{meta_lot or '—'}</strong></div>
                    <div class="col-sm-4"><span class="text-muted d-block">Device Type</span>
                      <strong>{_pt_str}</strong></div>
                    {'<div class="col-12"><span class="text-muted d-block">Notes</span><span>' + meta_notes + '</span></div>' if meta_notes else ''}
                  </div>
                </div>
                <!-- Edit form (hidden by default) -->
                <form id="edit-form" class="d-none" method="post" action="/projects/{pid}/meta">
                  <div class="row g-3">
                    <div class="col-12">
                      <label class="form-label">Project Name <span class="text-danger">*</span></label>
                      <input type="text" class="form-control" name="project_name" value="{p['name']}" required>
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Device / Part</label>
                      <input type="text" class="form-control" name="device_name" value="{meta_device}">
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Package</label>
                      <input type="text" class="form-control" name="device_pkg" value="{meta_pkg}">
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Bond Type</label>
                      <input type="text" class="form-control" name="bond_type" value="{meta_bond}">
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Engineer</label>
                      <input type="text" class="form-control" name="engineer" value="{meta_eng}">
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Lot / Wafer ID</label>
                      <input type="text" class="form-control" name="lot_id" value="{meta_lot}">
                    </div>
                    <div class="col-md-4">
                      <label class="form-label">Status</label>
                      <select class="form-select" name="status">{status_opts}</select>
                    </div>
                    <div class="col-12">
                      <label class="form-label">Notes</label>
                      <textarea class="form-control" name="notes" rows="2">{meta_notes}</textarea>
                    </div>
                    <div class="col-12 d-flex gap-2">
                      <button type="submit" class="btn btn-primary">Save</button>
                      <button type="button" class="btn btn-outline-secondary"
                              onclick="document.getElementById('edit-form').classList.add('d-none')">Cancel</button>
                    </div>
                  </div>
                </form>
              </div>
            </div>
          </div>

          <div class="col-lg-4">
            <div class="card shadow-sm mb-3">
              <div class="card-df"><h6 class="mb-0">Quick Actions</h6></div>
              <div class="list-group list-group-flush">
                <a href="/projects/{pid}/sample-size" class="list-group-item list-group-item-action d-flex align-items-center gap-3 py-3">
                  <i class="bi bi-calculator fs-5 text-muted"></i>
                  <div><div class="fw-semibold" style="font-size:.87rem">Sample Size Planner</div>
                  <div class="text-muted" style="font-size:.78rem">LTPD={ss['ltpd']}%, C={ss['failures']} failures</div></div>
                  <i class="bi bi-chevron-right ms-auto text-muted"></i>
                </a>
                <a href="/projects/{pid}/report" class="list-group-item list-group-item-action d-flex align-items-center gap-3 py-3">
                  <i class="bi bi-file-earmark-text fs-5 text-muted"></i>
                  <div><div class="fw-semibold" style="font-size:.87rem">Report Generation</div>
                  <div class="text-muted" style="font-size:.78rem">{n_tests_with_data} test{'s' if n_tests_with_data!=1 else ''} with sample data</div></div>
                  <i class="bi bi-chevron-right ms-auto text-muted"></i>
                </a>
                <a href="/projects/{pid}/csam" class="list-group-item list-group-item-action d-flex align-items-center gap-3 py-3">
                  <i class="bi bi-images fs-5 text-muted"></i>
                  <div><div class="fw-semibold" style="font-size:.87rem">CSAM Gallery</div>
                  <div class="text-muted" style="font-size:.78rem">Browse images by test and sample</div></div>
                  <i class="bi bi-chevron-right ms-auto text-muted"></i>
                </a>
                <a href="/projects/{pid}/tracker" class="list-group-item list-group-item-action d-flex align-items-center gap-3 py-3">
                  <i class="bi bi-calendar3 fs-5 text-muted"></i>
                  <div><div class="fw-semibold" style="font-size:.87rem">Schedule</div>
                  <div class="text-muted" style="font-size:.78rem">GANTT chart &amp; task timeline</div></div>
                  <i class="bi bi-chevron-right ms-auto text-muted"></i>
                </a>
              </div>
            </div>
          </div>
        </div>"""
        self.emit(body, p["name"], active="projects", project=p, active_sub="overview")

    def post(self, pid):
        # POST to project root → redirect to projects list
        self.redirect("/projects")

class ProjectMetaHandler(Base):
    def post(self, pid):
        pid = int(pid)
        p   = _db.get_project(pid)
        if not p:
            self.redirect("/projects"); return
        _db.save_meta(pid,
            device_name = self.get_argument("device_name", ""),
            device_pkg  = self.get_argument("device_pkg",  ""),
            bond_type   = self.get_argument("bond_type",   ""),
            engineer    = self.get_argument("engineer",    ""),
            lot_id      = self.get_argument("lot_id",      ""),
            notes       = self.get_argument("notes",       ""),
        )
        status       = self.get_argument("status", p["status"])
        new_name     = self.get_argument("project_name", "").strip()
        update_kwargs = {"status": status}
        if new_name:
            update_kwargs["name"] = new_name
        _db.update_project(pid, **update_kwargs)
        self.redirect(f"/projects/{pid}")

# ── Project-scoped: Sample Size ───────────────────────────────────────────────

class ProjectSampleSizeHandler(Base):
    def _get_project_or_404(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404)
        return p

    def get(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        saved = _db.get_sample_size(int(pid))
        self._render(p, k=saved["failures"], ltpd=saved["ltpd"], confidence=saved.get("confidence", 0.90))

    def post(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        action = self.get_argument("action", "planner")
        if action == "save_prescreen":
            import json as _jps
            try:
                entries = _jps.loads(self.get_argument("prescreen_json", "[]"))
            except Exception:
                entries = []
            _db.save_nonjec_prescreen(int(pid), entries)
            self._render(p)
            return
        elif action == "save_postqual":
            import json as _jpq
            try:
                entries = _jpq.loads(self.get_argument("postqual_json", "[]"))
            except Exception:
                entries = []
            _db.save_nonjec_postqual(int(pid), entries)
            self._render(p)
            return
        elif action == "per_test":
            # Save per-test sample counts and selected conditions
            tests = applicable_tests(p["part_type"])
            existing = _db.get_samples(int(pid))
            new_conditions = {}
            for key in tests:
                val = self.get_argument(f"n_{key}", "")
                if val.strip().isdigit():
                    existing[key] = int(val.strip())
                cond = self.get_argument(f"cond_{key}", "").strip()
                if cond:
                    new_conditions[key] = cond
            _db.save_samples(int(pid), existing)
            if new_conditions:
                _db.save_test_conditions(int(pid), new_conditions)
            self._render(p)
            return
        try:
            k          = int(self.get_argument("failures", "0"))
            ltpd       = float(self.get_argument("ltpd", "5"))
            confidence = float(self.get_argument("confidence", "0.90"))
            if k < 0: raise ValueError("Acceptance number C must be ≥ 0")
            if not (0.01 <= ltpd <= 99.99): raise ValueError("LTPD must be between 0.01% and 99.99%")
            if not (0.50 <= confidence <= 0.9999): raise ValueError("Confidence must be between 0.50 and 0.9999")
            _db.save_sample_size(int(pid), ltpd, k, confidence)
            n_jesd47 = min_sample_size_ltpd(ltpd, k)
            r_equiv  = 1.0 - ltpd / 100.0
            n_exact  = min_sample_size(r_equiv, confidence, k)
            self._render(p, k=k, ltpd=ltpd, confidence=confidence, n_jesd47=n_jesd47, n_exact=n_exact)
        except Exception as e:
            self._render(p, error=str(e))

    def _render(self, p, k=0, ltpd=5.0, confidence=0.90, n_jesd47=None, n_exact=None, error=None):
        pid   = p["id"]
        tests = applicable_tests(p["part_type"])
        saved_counts      = _db.get_samples(int(pid))
        saved_conds       = _db.get_test_conditions(int(pid))
        prescreen_entries = _db.get_nonjec_prescreen(int(pid))
        postqual_entries  = _db.get_nonjec_postqual(int(pid))

        # ── Table A ────────────────────────────────────────────────────────
        ltpd_cols = TABLE_A_LTPD
        nearest_ltpd = min(ltpd_cols, key=lambda l: abs(l - ltpd)) if n_jesd47 is not None else None
        hdr_cells = "".join(
            f'<th style="padding:3px 5px;font-size:.7rem;text-align:center;white-space:nowrap;'
            f'{"background:#dbeafe;font-weight:700;" if l == nearest_ltpd else ""}">'
            f'{l}%</th>'
            for l in ltpd_cols
        )
        tbl_rows = ""
        for c_val, row in TABLE_A.items():
            row_bg = ' style="background:#eff6ff"' if (n_jesd47 is not None and c_val == k) else ""
            cells = ""
            for ci, n_val in enumerate(row):
                is_sel = (n_jesd47 is not None and c_val == k and ltpd_cols[ci] == nearest_ltpd)
                td_s = ' style="font-weight:700;color:#1d4ed8;padding:3px 5px;font-size:.75rem;text-align:center"' if is_sel else ' style="padding:3px 5px;font-size:.75rem;text-align:center"'
                cells += f"<td{td_s}>{n_val}</td>"
            tbl_rows += f'<tr{row_bg}><td style="padding:3px 5px;font-size:.75rem;font-weight:600">{c_val}</td>{cells}</tr>'

        table_a_html = f"""
        <div class="card mb-3" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex align-items-center gap-2">
            <h6 class="mb-0" style="font-size:.82rem">Table A — JESD47I §3.8</h6>
            <span class="text-white-50" style="font-size:.7rem">Max % defective at 90% confidence</span>
          </div>
          <div class="table-responsive">
            <table class="table table-sm table-bordered mb-0" style="font-size:.75rem">
              <thead class="tbl-header">
                <tr>
                  <th style="padding:3px 5px;font-size:.7rem;white-space:nowrap">C</th>
                  {hdr_cells}
                </tr>
              </thead>
              <tbody>{tbl_rows}</tbody>
            </table>
          </div>
          <div class="card-footer text-muted py-1 px-2" style="font-size:.68rem">
            Highlighted = selected C &amp; LTPD. &ensp;
            <a href="{SPEC_URLS['jesd47'][0][1]}" target="_blank" style="font-size:.68rem">
              <i class="bi bi-file-earmark-pdf me-1"></i>JESD47I PDF
            </a>
          </div>
        </div>"""

        # ── Result card ────────────────────────────────────────────────────
        if error:
            result_html = f'<div class="alert alert-danger" style="font-size:.82rem"><i class="bi bi-exclamation-triangle me-2"></i>{error}</div>'
        elif n_jesd47 is not None:
            r_equiv = 1.0 - ltpd / 100.0
            result_html = f"""
            <div class="card mb-3" style="border:1px solid var(--df-border)">
              <div class="card-body text-center py-3">
                <div class="text-muted" style="font-size:.73rem">Min Sample Size (JESD47I)</div>
                <div style="font-size:2rem;font-weight:700;color:var(--df-navy);line-height:1.2">{n_jesd47}</div>
                <div class="text-muted mt-1" style="font-size:.72rem">
                  ≤<strong>{ltpd}%</strong> defective · 90% CL · C={k}
                </div>
                <div class="text-muted" style="font-size:.7rem">
                  Exact χ²: <strong>{n_exact}</strong>
                  {"&ensp;<span class='badge bg-secondary' style='font-size:.65rem'>same</span>" if n_exact == n_jesd47 else f"&ensp;<span style='color:#6b7280'>({'+' if n_jesd47>n_exact else ''}{n_jesd47-n_exact} vs exact)</span>"}
                </div>
              </div>
            </div>"""
        else:
            result_html = ""

        # ── Per-test allocation ────────────────────────────────────────────
        alloc_rows = ""
        for key, t in tests.items():
            count = saved_counts.get(key, "")
            if isinstance(count, list):
                count = len(count) if count else ""
            # Condition dropdown (only for tests that have options)
            cond_opts = _TEST_CONDITION_OPTIONS.get(key, [])
            saved_cond = saved_conds.get(key, "")
            if cond_opts:
                opt_html = '<option value="">— select —</option>'
                for ckey, clabel in cond_opts:
                    sel = ' selected' if ckey == saved_cond else ''
                    opt_html += f'<option value="{ckey}"{sel}>{clabel}</option>'
                cond_cell = (
                    f'<td style="padding:5px 8px">'
                    f'<select name="cond_{key}" class="form-select form-select-sm" style="font-size:.72rem;width:100%">'
                    f'{opt_html}</select>'
                    f'</td>'
                )
            else:
                cond_cell = '<td style="padding:5px 8px;font-size:.72rem;color:var(--df-grey)">—</td>'
            alloc_rows += (
                f'<tr>'
                f'<td style="padding:5px 8px;font-size:.78rem;font-weight:500">{t["name"]}</td>'
                f'<td style="padding:5px 8px;font-size:.72rem;color:var(--df-grey)">{t.get("standard","")}</td>'
                f'<td style="padding:5px 8px">'
                f'<input type="number" name="n_{key}" value="{count}" min="0" '
                f'class="form-control form-control-sm" style="width:72px" placeholder="—">'
                f'</td>'
                f'{cond_cell}'
                f'</tr>'
            )

        per_test_panel = f"""
        <div class="card" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex justify-content-between align-items-center">
            <h6 class="mb-0" style="font-size:.82rem">Sample Allocation per Test</h6>
            <span style="font-size:.7rem;color:rgba(255,255,255,.6)">Shown on Schedule</span>
          </div>
          <div class="card-body p-0">
            <form method="post" action="/projects/{pid}/sample-size">
              <input type="hidden" name="action" value="per_test">
              <div style="overflow-x:auto">
              <table class="table table-sm mb-0" style="min-width:500px">
                <thead class="tbl-header">
                  <tr>
                    <th style="padding:5px 8px;font-size:.72rem">Test</th>
                    <th style="padding:5px 8px;font-size:.72rem">Std</th>
                    <th style="padding:5px 8px;font-size:.72rem">n</th>
                    <th style="padding:5px 8px;font-size:.72rem">Condition</th>
                  </tr>
                </thead>
                <tbody>{alloc_rows}</tbody>
              </table>
              </div>
              <div class="p-2 border-top" data-admin-gate="show" style="display:none">
                <button type="submit" class="btn btn-sm"
                  style="background:var(--df-accent);color:#fff;border:none;font-size:.78rem">
                  Save Sample Counts
                </button>
              </div>
            </form>
          </div>
        </div>"""

        # ── Pre-Qual Engineering Analysis panel ────────────────────────────
        def _ps_type_opts(sel="pull_test"):
            return "".join(
                f'<option value="{k}"{" selected" if k == sel else ""}>{v["label"]}</option>'
                for k, v in _NONJEC_TYPES.items()
            )
        def _ps_row(entry):
            ttype = entry.get("test_type", "pull_test")
            cname = entry.get("custom_name", "")
            samp  = entry.get("sample_count", 0)
            is_o  = ttype == "other"
            dv    = entry.get("duration_weeks") or _NONJEC_TYPES.get(ttype, {}).get("default_dur", 1) or 1
            nd_name = "" if is_o else " disabled"  # custom name only editable for "other"
            return (
                f'<tr>'
                f'<td style="padding:4px 6px"><select class="form-select form-select-sm ps-type" style="width:110px" onchange="onPsTypeChange(this)">{_ps_type_opts(ttype)}</select></td>'
                f'<td style="padding:4px 6px"><input type="text" class="form-control form-control-sm ps-name" value="{cname}" placeholder="Custom name" style="width:130px"{nd_name}></td>'
                f'<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm ps-dur" value="{dv}" min="1" style="width:60px"></td>'
                f'<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm ps-n" value="{samp}" min="0" style="width:72px" placeholder="0"></td>'
                f'<td style="padding:4px 6px"><button type="button" class="btn btn-sm" style="padding:2px 7px;font-size:.72rem;border:1px solid #fca5a5;color:#dc2626;background:#fff" onclick="this.closest(\'tr\').remove()">&#x2715;</button></td>'
                f'</tr>'
            )
        ps_rows = "".join(_ps_row(e) for e in prescreen_entries)

        prescreen_panel = f"""
        <div class="card mt-3" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex justify-content-between align-items-center">
            <h6 class="mb-0" style="font-size:.82rem">Pre-Qual Engineering Analysis</h6>
            <span style="font-size:.7rem;color:rgba(255,255,255,.6)">Samples not entering JEDEC tests</span>
          </div>
          <div class="card-body p-0">
            <form method="post" action="/projects/{pid}/sample-size" id="prescreenForm" onsubmit="collectPrescreenData()">
              <input type="hidden" name="action" value="save_prescreen">
              <input type="hidden" name="prescreen_json" id="prescreenJson">
              <div style="overflow-x:auto">
                <table class="table table-sm mb-0" style="min-width:520px">
                  <thead class="tbl-header">
                    <tr>
                      <th style="padding:5px 8px;font-size:.72rem">Test Type</th>
                      <th style="padding:5px 8px;font-size:.72rem">Custom Name</th>
                      <th style="padding:5px 8px;font-size:.72rem">Wks</th>
                      <th style="padding:5px 8px;font-size:.72rem">n</th>
                      <th style="padding:5px 8px;font-size:.72rem"></th>
                    </tr>
                  </thead>
                  <tbody id="prescreenBody">{ps_rows if ps_rows else '<tr id="ps-empty"><td colspan="5" class="text-center text-muted py-3" style="font-size:.8rem">No pre-screen tests added.</td></tr>'}</tbody>
                </table>
              </div>
              <div class="p-2 border-top d-flex gap-2">
                <button type="button" class="btn btn-sm btn-outline-secondary" data-admin-gate="show" style="display:none;font-size:.78rem" onclick="addPrescreenRow()">
                  <i class="bi bi-plus-lg me-1"></i>Add Row
                </button>
                <button type="submit" class="btn btn-sm" data-admin-gate="show" style="display:none;background:var(--df-accent);color:#fff;border:none;font-size:.78rem">
                  Save
                </button>
              </div>
            </form>
          </div>
        </div>"""

        # ── Post-Qual Engineering Analysis panel ─────────────────────────────
        _jedec_test_opts_base = '<option value="">— select JEDEC test —</option>' + "".join(
            f'<option value="{key}">{t["name"]}</option>' for key, t in tests.items()
        )
        def _pq_type_opts(sel="pull_test"):
            return "".join(
                f'<option value="{k}"{" selected" if k == sel else ""}>{v["label"]}</option>'
                for k, v in _NONJEC_TYPES.items()
            )
        def _pq_parent_opts(sel=""):
            return '<option value="">— select JEDEC test —</option>' + "".join(
                f'<option value="{key}"{" selected" if key == sel else ""}>{t["name"]}</option>'
                for key, t in tests.items()
            )
        def _pq_row(entry):
            ttype  = entry.get("test_type", "pull_test")
            cname  = entry.get("custom_name", "")
            parent = entry.get("parent_stress_test_key", "")
            samp   = entry.get("sample_count", 0)
            is_o   = ttype == "other"
            dv     = entry.get("duration_weeks") or _NONJEC_TYPES.get(ttype, {}).get("default_dur", 1) or 1
            nd_name = "" if is_o else " disabled"  # custom name only editable for "other"
            return (
                f'<tr>'
                f'<td style="padding:4px 6px"><select class="form-select form-select-sm pq-type" style="width:110px" onchange="onPqTypeChange(this)">{_pq_type_opts(ttype)}</select></td>'
                f'<td style="padding:4px 6px"><input type="text" class="form-control form-control-sm pq-name" value="{cname}" placeholder="Custom name" style="width:120px"{nd_name}></td>'
                f'<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm pq-dur" value="{dv}" min="1" style="width:55px"></td>'
                f'<td style="padding:4px 6px"><select class="form-select form-select-sm pq-parent" style="min-width:155px">{_pq_parent_opts(parent)}</select></td>'
                f'<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm pq-n" value="{samp}" min="0" style="width:68px" placeholder="0"></td>'
                f'<td style="padding:4px 6px"><button type="button" class="btn btn-sm" style="padding:2px 7px;font-size:.72rem;border:1px solid #fca5a5;color:#dc2626;background:#fff" onclick="this.closest(\'tr\').remove()">&#x2715;</button></td>'
                f'</tr>'
            )
        pq_rows = "".join(_pq_row(e) for e in postqual_entries)

        postqual_panel = f"""
        <div class="card mt-3" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex justify-content-between align-items-center">
            <h6 class="mb-0" style="font-size:.82rem">Post-Qual Engineering Analysis</h6>
            <span style="font-size:.7rem;color:rgba(255,255,255,.6)">Samples drawn from post-JEDEC pool</span>
          </div>
          <div class="card-body p-0">
            <form method="post" action="/projects/{pid}/sample-size" id="postqualForm" onsubmit="collectPostqualData()">
              <input type="hidden" name="action" value="save_postqual">
              <input type="hidden" name="postqual_json" id="postqualJson">
              <div style="overflow-x:auto">
                <table class="table table-sm mb-0" style="min-width:620px">
                  <thead class="tbl-header">
                    <tr>
                      <th style="padding:5px 8px;font-size:.72rem">Test Type</th>
                      <th style="padding:5px 8px;font-size:.72rem">Custom Name</th>
                      <th style="padding:5px 8px;font-size:.72rem">Wks</th>
                      <th style="padding:5px 8px;font-size:.72rem">Parent JEDEC Test</th>
                      <th style="padding:5px 8px;font-size:.72rem">n / set</th>
                      <th style="padding:5px 8px;font-size:.72rem"></th>
                    </tr>
                  </thead>
                  <tbody id="postqualBody">{pq_rows if pq_rows else '<tr id="pq-empty"><td colspan="6" class="text-center text-muted py-3" style="font-size:.8rem">No post-qual tests added.</td></tr>'}</tbody>
                </table>
              </div>
              <div class="p-2 border-top d-flex gap-2">
                <button type="button" class="btn btn-sm btn-outline-secondary" data-admin-gate="show" style="display:none;font-size:.78rem" onclick="addPostqualRow()">
                  <i class="bi bi-plus-lg me-1"></i>Add Row
                </button>
                <button type="submit" class="btn btn-sm" data-admin-gate="show" style="display:none;background:var(--df-accent);color:#fff;border:none;font-size:.78rem">
                  Save
                </button>
              </div>
            </form>
          </div>
        </div>"""

        # JS for both sections — embedded in the page body
        _nonjec_durs_js = '{' + ", ".join(
            f'"{k}": {v["default_dur"] if v["default_dur"] is not None else "null"}'
            for k, v in _NONJEC_TYPES.items()
        ) + '}'
        _jedec_opts_js = _jedec_test_opts_base.replace("'", "\\'")

        nonjec_script = f"""<script>
        var _NONJEC_DURS = {_nonjec_durs_js};
        var _JEDEC_TEST_OPTS = '{_jedec_opts_js}';

        function onPsTypeChange(sel) {{
          var tr = sel.closest('tr');
          var nameEl = tr.querySelector('.ps-name');
          var durEl  = tr.querySelector('.ps-dur');
          var isOther = (sel.value === 'other');
          nameEl.disabled = !isOther;
          // Duration is always editable; auto-populate default when type changes
          if (!isOther) {{
            nameEl.value = '';
            var d = _NONJEC_DURS[sel.value];
            if (d !== null && d !== undefined) durEl.value = d;
          }}
        }}
        function onPqTypeChange(sel) {{
          var tr = sel.closest('tr');
          var nameEl = tr.querySelector('.pq-name');
          var durEl  = tr.querySelector('.pq-dur');
          var isOther = (sel.value === 'other');
          nameEl.disabled = !isOther;
          // Duration is always editable; auto-populate default when type changes
          if (!isOther) {{
            nameEl.value = '';
            var d = _NONJEC_DURS[sel.value];
            if (d !== null && d !== undefined) durEl.value = d;
          }}
        }}

        function removeNonjecRow(btn) {{ btn.closest('tr').remove(); }}

        var _DEL_BTN = '<button type="button" class="btn btn-sm" style="padding:2px 7px;font-size:.72rem;border:1px solid #fca5a5;color:#dc2626;background:#fff" onclick="removeNonjecRow(this)">&#x2715;</button>';

        function _buildTypeOpts(sel) {{
          var types = {{'pull_test':'Pull Test','xsem':'X-SEM','other':'Other'}};
          return Object.entries(types).map(([k,v]) =>
            '<option value="' + k + '"' + (k===sel?' selected':'') + '>' + v + '</option>'
          ).join('');
        }}

        function addPrescreenRow() {{
          var empty = document.getElementById('ps-empty');
          if (empty) empty.remove();
          var tbody = document.getElementById('prescreenBody');
          var tr = document.createElement('tr');
          var defDur = _NONJEC_DURS['pull_test'] || 1;
          tr.innerHTML = (
            '<td style="padding:4px 6px"><select class="form-select form-select-sm ps-type" style="width:110px" onchange="onPsTypeChange(this)">' + _buildTypeOpts('pull_test') + '</select></td>'
            + '<td style="padding:4px 6px"><input type="text" class="form-control form-control-sm ps-name" placeholder="Custom name" style="width:130px" disabled></td>'
            + '<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm ps-dur" value="' + defDur + '" min="1" style="width:60px"></td>'
            + '<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm ps-n" value="0" min="0" style="width:72px"></td>'
            + '<td style="padding:4px 6px">' + _DEL_BTN + '</td>'
          );
          tbody.appendChild(tr);
        }}

        function addPostqualRow() {{
          var empty = document.getElementById('pq-empty');
          if (empty) empty.remove();
          var tbody = document.getElementById('postqualBody');
          var tr = document.createElement('tr');
          var defDur = _NONJEC_DURS['pull_test'] || 1;
          tr.innerHTML = (
            '<td style="padding:4px 6px"><select class="form-select form-select-sm pq-type" style="width:110px" onchange="onPqTypeChange(this)">' + _buildTypeOpts('pull_test') + '</select></td>'
            + '<td style="padding:4px 6px"><input type="text" class="form-control form-control-sm pq-name" placeholder="Custom name" style="width:120px" disabled></td>'
            + '<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm pq-dur" value="' + defDur + '" min="1" style="width:55px"></td>'
            + '<td style="padding:4px 6px"><select class="form-select form-select-sm pq-parent" style="min-width:155px">' + _JEDEC_TEST_OPTS + '</select></td>'
            + '<td style="padding:4px 6px"><input type="number" class="form-control form-control-sm pq-n" value="0" min="0" style="width:68px"></td>'
            + '<td style="padding:4px 6px">' + _DEL_BTN + '</td>'
          );
          tbody.appendChild(tr);
        }}

        function collectPrescreenData() {{
          var rows = document.querySelectorAll('#prescreenBody tr');
          var data = [];
          rows.forEach(function(tr) {{
            var typeEl = tr.querySelector('.ps-type');
            var nameEl = tr.querySelector('.ps-name');
            var durEl  = tr.querySelector('.ps-dur');
            var nEl    = tr.querySelector('.ps-n');
            if (!typeEl) return;
            data.push({{
              test_type: typeEl.value,
              custom_name: nameEl ? nameEl.value : '',
              duration_weeks: durEl ? (parseInt(durEl.value)||1) : 1,
              sample_count: nEl ? (parseInt(nEl.value)||0) : 0
            }});
          }});
          document.getElementById('prescreenJson').value = JSON.stringify(data);
        }}

        function collectPostqualData() {{
          var rows = document.querySelectorAll('#postqualBody tr');
          var data = [];
          rows.forEach(function(tr) {{
            var typeEl   = tr.querySelector('.pq-type');
            var nameEl   = tr.querySelector('.pq-name');
            var durEl    = tr.querySelector('.pq-dur');
            var parentEl = tr.querySelector('.pq-parent');
            var nEl      = tr.querySelector('.pq-n');
            if (!typeEl) return;
            data.push({{
              test_type: typeEl.value,
              custom_name: nameEl ? nameEl.value : '',
              duration_weeks: durEl ? (parseInt(durEl.value)||1) : 1,
              parent_stress_test_key: parentEl ? parentEl.value : '',
              sample_count: nEl ? (parseInt(nEl.value)||0) : 0
            }});
          }});
          document.getElementById('postqualJson').value = JSON.stringify(data);
        }}
        </script>"""

        body = f"""
        <h5 class="mb-3" style="font-weight:300">Sample Size Planner — {p['name']}</h5>
        <div class="row g-3">
          <!-- Left: Parameters + Result -->
          <div class="col-xl-3 col-lg-4">
            <div class="card" style="border:1px solid var(--df-border)">
              <div class="card-df"><h6 class="mb-0" style="font-size:.83rem">Parameters</h6></div>
              <div class="card-body p-3">
                <p class="text-muted mb-3" style="font-size:.76rem">
                  Per <strong>JESD47I §3.8</strong> — minimum units to demonstrate
                  a maximum defect rate at 90% confidence.
                </p>
                <form method="post" action="/projects/{pid}/sample-size">
                  <div class="mb-3">
                    <label class="form-label" style="font-size:.8rem">LTPD — Max % Defective</label>
                    <div class="input-group input-group-sm">
                      <input type="number" class="form-control" name="ltpd"
                             value="{ltpd}" min="0.01" max="99.99" step="0.01" required>
                      <span class="input-group-text">%</span>
                    </div>
                    <div class="form-text" style="font-size:.7rem">Used for pass/fail in Reporting</div>
                  </div>
                  <div class="mb-3">
                    <label class="form-label" style="font-size:.8rem">Confidence Level</label>
                    <div class="input-group input-group-sm">
                      <input type="number" class="form-control" name="confidence"
                             value="{confidence}" min="0.50" max="0.9999" step="0.01" required>
                      <span class="input-group-text">e.g. 0.90</span>
                    </div>
                    <div class="form-text" style="font-size:.7rem">Used for pass/fail in Reporting</div>
                  </div>
                  <div class="mb-3">
                    <label class="form-label" style="font-size:.8rem">Acceptance Number (C)</label>
                    <input type="number" class="form-control form-control-sm" name="failures"
                           value="{k}" min="0" step="1" required>
                    <div class="form-text" style="font-size:.7rem">0 = zero-failure plan</div>
                  </div>
                  <button type="submit" class="btn btn-sm btn-primary w-100">Calculate →</button>
                </form>
                <hr class="my-3">
                <p class="text-muted mb-0" style="font-size:.7rem">
                  LTPD 5% ≡ R = 95%. Confidence and LTPD
                  are applied when evaluating pass/fail in the Reporting tab.
                </p>
              </div>
            </div>
            {result_html}
          </div>
          <!-- Middle: Table A -->
          <div class="col-xl-5 col-lg-4">
            {table_a_html}
          </div>
          <!-- Right: Sample Allocation -->
          <div class="col-xl-4 col-lg-4">
            {per_test_panel}
          </div>
        </div>
        <div class="row g-3 mt-1">
          <div class="col-lg-6">
            {prescreen_panel}
          </div>
          <div class="col-lg-6">
            {postqual_panel}
          </div>
        </div>
        {nonjec_script}"""
        self.emit(body, f"Planner — {p['name']}", active="projects",
                  project=p, active_sub="sample-size")

# ── Project-scoped: Pass / Fail ───────────────────────────────────────────────

class ProjectPassFailHandler(Base):
    def _get_project_or_404(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404)
        return p

    def get(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        saved = _db.get_pass_fail(int(pid))
        n_val = saved.get("n", "")
        try:
            n_val = int(n_val) if n_val else ""
        except Exception:
            n_val = ""
        self._render(p, n=n_val, k=saved.get("failures",0),
                     c=saved.get("confidence",0.90), ltpd_inp=saved.get("ltpd_pct",5.0))

    def post(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        try:
            n        = int(self.get_argument("n", ""))
            k        = int(self.get_argument("failures", "0"))
            c        = float(self.get_argument("confidence", "0.90"))
            ltpd_inp = float(self.get_argument("ltpd_pct", "5.0"))
            if n < 1:     raise ValueError("Need at least 1 sample")
            if k > n:     raise ValueError("Failures cannot exceed samples tested")
            if not (0.50 <= c <= 0.9999): raise ValueError("Confidence must be 0.50–0.9999")
            if not (0.01 <= ltpd_inp <= 99.99): raise ValueError("Defective Rate must be 0.01–99.99%")
            _db.save_pass_fail(int(pid), n, k, c, ltpd_inp)
            r_req = 1.0 - ltpd_inp / 100.0
            passed, r_demo = _pf(n, k, c, r_req)
            sens = [(f, demonstrated_reliability(n, f, c)) for f in range(min(n+1, 9))]
            self._render(p, n=n, k=k, c=c, ltpd_inp=ltpd_inp,
                         passed=passed, r_demo=r_demo, sens=sens)
        except (ValueError, TypeError) as e:
            self._render(p, error=str(e))

    def _render(self, p, n="", k=0, c=0.90, ltpd_inp=5.0,
                passed=None, r_demo=None, sens=None, error=None):
        captured = {}
        def _capture(body, title="", active="", project=None, active_sub=""):
            captured["body"] = body
        orig_emit = self.emit
        self.emit = _capture
        PassFailHandler._render(self, n=n, k=k, c=c, ltpd_inp=ltpd_inp,
                                passed=passed, r_demo=r_demo, sens=sens, error=error)
        self.emit = orig_emit
        body = captured.get("body", "")
        self.emit(body, f"Pass/Fail — {p['name']}", active="projects",
                  project=p, active_sub="pass-fail")

# ── Project-scoped: Qual Report ───────────────────────────────────────────────

# ── Non-JEDEC results section (injected into project Qual Report form) ────────

def _build_nonjec_report_section(prescreen: list, postqual: list,
                                  saved: dict) -> str:
    """Return HTML for the Non-JEDEC Test Results card to embed in the report form.

    prescreen / postqual: list of entry dicts from DB.
    saved: dict keyed by (entry_id, source) from _db.get_nonjec_results().
    """

    def _entry_html(entry: dict, source: str) -> str:
        eid      = entry["id"]
        ttype    = entry.get("test_type", "pull_test")
        name     = _nonjec_display_name(entry)
        saved_r  = saved.get((eid, source), {})
        comments = saved_r.get("comments", "")
        pfx      = f"nonjec_{source}_{eid}"      # form field prefix

        # ── Saved thumbnail helper ────────────────────────────────────────────
        def _thumb(slot: int, label: str) -> str:
            img_b64 = saved_r.get(f"img_{slot}", "")
            preview = (f'<img src="{img_b64}" class="img-thumbnail mt-1"'
                       f' style="max-height:80px;max-width:120px;object-fit:cover"'
                       f' alt="{label}">') if img_b64 else ""
            return (f'<div><label class="form-label mb-1" style="font-size:.78rem">'
                    f'{label}</label>'
                    f'<input type="file" class="form-control form-control-sm"'
                    f' name="{pfx}_img{slot}" accept="image/*">'
                    f'{preview}</div>')

        if ttype == "pull_test":
            saved_mpa = saved_r.get("pull_strength_mpa", "")
            mpa_val   = f' value="{saved_mpa}"' if saved_mpa not in (None, "") else ""
            img_block = f"""
              <div class="row g-2 mb-2">
                <div class="col-md-6">{_thumb(1, "Surface A")}</div>
                <div class="col-md-6">{_thumb(2, "Surface B")}</div>
              </div>"""
            mpa_block = f"""
              <div class="mb-2">
                <label class="form-label mb-1" style="font-size:.78rem">Pull Strength (MPa)</label>
                <div class="input-group input-group-sm" style="max-width:200px">
                  <input type="number" class="form-control form-control-sm"
                         name="{pfx}_mpa" step="0.01" min="0" placeholder="e.g. 42.5"{mpa_val}>
                  <span class="input-group-text">MPa</span>
                </div>
              </div>"""
        else:
            # X-SEM or Other: up to 5 images
            n_imgs = 5
            img_cols = ""
            for i in range(1, n_imgs + 1):
                lbl = f"SEM Image {i}" if ttype == "xsem" else f"Image {i}"
                img_cols += f'<div class="col-md-4 col-lg-2">{_thumb(i, lbl)}</div>'
            img_block = f'<div class="row g-2 mb-2">{img_cols}</div>'
            mpa_block = ""

        comments_html = (
            f'<div>'
            f'<label class="form-label mb-1" style="font-size:.78rem">Comments <span class="text-muted">(optional)</span></label>'
            f'<textarea class="form-control form-control-sm" name="{pfx}_comments"'
            f' rows="2" placeholder="Optional notes…">{comments}</textarea>'
            f'</div>'
        )

        return (
            f'<div class="border rounded p-3 mb-3" style="background:#fafafa">'
            f'<h6 class="mb-3" style="font-size:.88rem;font-weight:600">{name}</h6>'
            f'{mpa_block}'
            f'{img_block}'
            f'{comments_html}'
            f'</div>'
        )

    # ── Build groups ──────────────────────────────────────────────────────────
    groups: list[tuple[str, list]] = []     # (group_label, list_of_(entry, source))

    # Pre-stress group: all prescreen + orphan postqual (no parent)
    pre_entries = [(e, "pre") for e in prescreen]
    orphans     = [(e, "pq") for e in postqual if not e.get("parent_stress_test_key", "").strip()]
    if pre_entries or orphans:
        groups.append(("Pre-Stress Tests", pre_entries + orphans))

    # Post-stress groups: one per parent stress key (in JEDEC test order)
    _stress_order = ["uhast", "tc", "tshock", "mshock", "vib", "hts"]
    pq_by_parent: dict[str, list] = {}
    for e in postqual:
        pk = e.get("parent_stress_test_key", "").strip()
        if pk:
            pq_by_parent.setdefault(pk, []).append(e)
    for sk in _stress_order:
        if sk in pq_by_parent:
            abbrev = _STRESS_ABBREV.get(sk, sk.upper())
            groups.append((f"Post-{abbrev} Tests",
                           [(e, "pq") for e in pq_by_parent[sk]]))
    # Any remaining keys not in stress_order
    for sk, entries_list in pq_by_parent.items():
        if sk not in _stress_order:
            abbrev = _STRESS_ABBREV.get(sk, sk.upper())
            groups.append((f"Post-{abbrev} Tests",
                           [(e, "pq") for e in entries_list]))

    if not groups:
        return ""   # no non-JEDEC tests configured

    group_html = ""
    for label, items in groups:
        if not items:
            continue
        entries_html = "".join(_entry_html(e, src) for e, src in items)
        group_html += (
            f'<h6 class="mb-3 mt-3" style="font-size:.82rem;text-transform:uppercase;'
            f'letter-spacing:.04em;color:var(--df-mid)">{label}</h6>'
            f'{entries_html}'
        )

    return (
        f'<div class="card mb-4">'
        f'<div class="card-df"><h6 class="mb-0">Non-JEDEC Test Results</h6></div>'
        f'<div class="card-body p-4">'
        f'{group_html}'
        f'</div>'
        f'</div>'
    )

# ── Helpers: parse uploaded image file → base64 data-URI ──────────────────────

def _read_upload_b64(handler, field_name: str) -> str:
    """Return base64 data-URI for an uploaded image field, or '' if absent."""
    try:
        files = handler.request.files.get(field_name)
        if not files:
            return ""
        f = files[0]
        if not f["body"]:
            return ""
        mime = f.get("content_type", "image/jpeg") or "image/jpeg"
        b64  = base64.b64encode(f["body"]).decode()
        return f"data:{mime};base64,{b64}"
    except Exception:
        return ""


class ProjectReportHandler(ReportHandler):
    def _get_project_or_404(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404)
        return p

    def get(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        _, s = self.sess()
        s["part_type"] = p["part_type"]
        # Inject saved planner sample counts so _render_form can use them as n defaults
        self._saved_n = _db.get_samples(int(pid))
        # Inject project LTPD + confidence so pass/fail in the form uses project settings
        _qual = _db.get_sample_size(int(pid))
        self._qual_ltpd       = _qual.get("ltpd", 5.0)
        self._qual_confidence = _qual.get("confidence", 0.90)
        # Inject saved test conditions so report includes Tests Performed section
        self._test_conditions = _db.get_test_conditions(int(pid))
        # Make the form POST back to the project-scoped URL (not the standalone /report)
        self._form_action = f"/projects/{pid}/report"
        captured = {}
        def _capture(body, title="", active="", project=None, active_sub=""):
            captured["body"] = body
        orig_emit = self.emit
        self.emit = _capture
        ReportHandler.get(self)
        self.emit = orig_emit
        body = captured.get("body", "")
        # ── Inject Non-JEDEC Test Results section before the submit button ────
        _prescreen  = _db.get_nonjec_prescreen(int(pid))
        _postqual   = _db.get_nonjec_postqual(int(pid))
        _nj_saved   = _db.get_nonjec_results(int(pid))
        _nj_section = _build_nonjec_report_section(_prescreen, _postqual, _nj_saved)
        _submit_marker = '<button type="submit" class="btn btn-primary btn-lg px-5">'
        if _nj_section and _submit_marker in body:
            body = body.replace(_submit_marker, _nj_section + _submit_marker, 1)
        self.emit(body, f"Qual Report — {p['name']}", active="projects",
                  project=p, active_sub="report")

    def post(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        _, s = self.sess()
        s["part_type"] = p["part_type"]
        # Inject project LTPD + confidence so ReportHandler.post uses them for pass/fail
        _qual = _db.get_sample_size(int(pid))
        self._qual_ltpd       = _qual.get("ltpd", 5.0)
        self._qual_confidence = _qual.get("confidence", 0.90)
        self._test_conditions = _db.get_test_conditions(int(pid))
        # Delegate to existing ReportHandler.post which builds report dict and stores in session
        captured = {}
        def _capture(body, title="", active="", project=None, active_sub=""):
            captured["body"] = body
        orig_emit = self.emit
        self.emit = _capture
        ReportHandler.post(self)
        self.emit = orig_emit
        # Save whatever samples were parsed back to DB
        report = s.get("last_report")
        if report and isinstance(report, dict):
            samples = report.get("samples", {})
            if samples:
                _db.save_samples(int(pid), samples)
                # Auto-save CSAM images into the gallery
                _img_stage_map = [
                    ("img_bpc",  "pre_cond"),
                    ("img_apc",  "post_cond"),
                    ("img_atst", "post_test"),
                ]
                for test_key, sample_list in samples.items():
                    if not isinstance(sample_list, list):
                        continue
                    for sample in sample_list:
                        sample_id = sample.get("id", "")
                        for field, stage in _img_stage_map:
                            img_data = sample.get(field, "")
                            if img_data:
                                _db.upsert_csam_image(
                                    int(pid), sample_id, test_key, stage, img_data
                                )
        # ── Save non-JEDEC test results ───────────────────────────────────────
        _prescreen = _db.get_nonjec_prescreen(int(pid))
        _postqual  = _db.get_nonjec_postqual(int(pid))
        all_entries = (
            [(e, "pre") for e in _prescreen] +
            [(e, "pq")  for e in _postqual]
        )
        for entry, src in all_entries:
            eid  = entry["id"]
            pfx  = f"nonjec_{src}_{eid}"
            ttype = entry.get("test_type", "pull_test")

            # Pull strength (pull_test only)
            mpa_raw = self.get_argument(f"{pfx}_mpa", "").strip()
            try:    mpa_val = float(mpa_raw) if mpa_raw else None
            except: mpa_val = None

            # Comments
            comments = self.get_argument(f"{pfx}_comments", "").strip()

            # Images: 2 for pull_test, 5 for others
            n_imgs = 2 if ttype == "pull_test" else 5
            images = []
            for i in range(1, n_imgs + 1):
                img = _read_upload_b64(self, f"{pfx}_img{i}")
                if not img:
                    # Keep existing image if no new upload
                    img = _db.get_nonjec_results(int(pid)).get((eid, src), {}).get(f"img_{i}", "")
                images.append(img)

            _db.save_nonjec_result(int(pid), eid, src,
                                   pull_strength_mpa=mpa_val,
                                   comments=comments,
                                   images=images)

        body = captured.get("body", "")
        # Re-inject non-JEDEC section with saved values so it re-renders correctly
        _nj_saved   = _db.get_nonjec_results(int(pid))
        _nj_section = _build_nonjec_report_section(_prescreen, _postqual, _nj_saved)
        _submit_marker = '<button type="submit" class="btn btn-primary btn-lg px-5">'
        if _nj_section and _submit_marker in body:
            body = body.replace(_submit_marker, _nj_section + _submit_marker, 1)
        self.emit(body, f"Qual Report — {p['name']}", active="projects",
                  project=p, active_sub="report")

class ProjectReportPdfHandler(Base):
    def get(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404); return
        _, s = self.sess()
        s["part_type"] = p["part_type"]
        # Delegate entirely to existing PDF handler
        ReportPdfHandler.get(self)


# ── Project-scoped: CSAM Image Repository ─────────────────────────────────────

_CSAM_TEST_LABELS = {
    "uhast": "uHAST", "tc": "TC", "tshock": "T-Shock", "mshock": "M-Shock",
    "vib": "Vibration", "pc": "Power Cycling", "hts": "HTS",
    "shadow_moire": "Shadow Moiré", "htol": "HTOL", "elfr": "ELFR",
    "thb": "THB", "esd_cdm": "ESD CDM", "esd_hbm": "ESD HBM",
    "latchup": "Latch-up", "ptc": "PTC",
}
_CSAM_STAGES = [
    ("pre_bond",  "Pre-Bond"),
    ("post_bond", "Post-Bond"),
    ("post_assy", "Post-Assy"),
    ("pre_cond",  "Pre-Conditioning"),
    ("post_cond", "Post-Conditioning"),
    ("post_test", "Post-Test"),
    ("other",     "Other"),
]

class ProjectCsamHandler(Base):
    def _get_or_404(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404)
        return p

    def get(self, pid):
        p = self._get_or_404(pid)
        if not p: return
        images = _db.list_csam_images(int(pid))

        # Group by test_key → sample_id → stage order
        from collections import OrderedDict
        groups = OrderedDict()  # test_key → {sample_id → [img, ...]}
        for img in images:
            tk = img["test_key"] or "—"
            si = img["sample_id"] or "—"
            if tk not in groups:
                groups[tk] = OrderedDict()
            if si not in groups[tk]:
                groups[tk][si] = []
            groups[tk][si].append(img)

        _stage_map = dict(_CSAM_STAGES)
        gallery_html = ""
        if not images:
            gallery_html = (
                '<div class="text-center text-muted py-5" style="font-size:.9rem">'
                '<i class="bi bi-image fs-2 d-block mb-2"></i>'
                '<div>No images yet.</div>'
                '<div style="font-size:.8rem;margin-top:.5rem">Images appear here automatically '
                'when you upload them in the Reporting section.</div></div>'
            )
        else:
            for tk, samples in groups.items():
                tname = _CSAM_TEST_LABELS.get(tk, tk)
                gallery_html += (
                    f'<h6 class="mt-4 mb-2" style="font-size:.83rem;font-weight:700;'
                    f'text-transform:uppercase;letter-spacing:.05em;color:#4b5563">'
                    f'{tname}</h6>'
                )
                for si, imgs in samples.items():
                    gallery_html += (
                        f'<div style="font-size:.78rem;font-weight:600;color:#374151;'
                        f'margin-bottom:.4rem">{si}</div>'
                        f'<div class="row g-2 mb-3">'
                    )
                    for img in imgs:
                        stage_label = _stage_map.get(img["stage"], img["stage"] or "—")
                        iid = img["id"]
                        gallery_html += f"""
                        <div class="col-6 col-md-4 col-lg-3">
                          <div class="card shadow-sm h-100" style="border:1px solid #e5e7eb;overflow:hidden">
                            <a href="/projects/{pid}/csam/{iid}" target="_blank" class="d-block"
                               style="background:#f9fafb;text-align:center;padding:8px">
                              <img src="/projects/{pid}/csam/{iid}/thumb"
                                   style="max-height:130px;max-width:100%;object-fit:contain;border-radius:4px"
                                   alt="{stage_label}" loading="lazy">
                            </a>
                            <div class="card-body p-2">
                              <div style="font-size:.72rem;font-weight:600;color:#374151">{stage_label}</div>
                              {f'<div style="font-size:.68rem;color:#9ca3af;margin-top:2px">{img["notes"]}</div>' if img.get("notes") else ""}
                              <div style="font-size:.66rem;color:#d1d5db;margin-top:3px">{img["created_at"][:10]}</div>
                            </div>
                            <div class="card-footer p-2 d-flex gap-1">
                              <a href="/projects/{pid}/csam/{iid}" target="_blank"
                                 class="btn btn-sm flex-fill" style="font-size:.7rem;border:1px solid #e5e7eb">
                                <i class="bi bi-zoom-in"></i> View
                              </a>
                              <form method="post" action="/projects/{pid}/csam/{iid}/delete"
                                    class="d-inline flex-fill"
                                    onsubmit="return confirm('Delete this image?')">
                                <button class="btn btn-sm w-100" style="font-size:.7rem;
                                  border:1px solid #fca5a5;color:#dc2626;background:#fff">
                                  <i class="bi bi-trash"></i>
                                </button>
                              </form>
                            </div>
                          </div>
                        </div>"""
                    gallery_html += "</div>"

        body = f"""
        <h5 class="mb-1" style="font-weight:300">CSAM Gallery — {p['name']}</h5>
        <p class="text-muted mb-4" style="font-size:.83rem">
          Images are saved automatically when you submit the Reporting form.
          They are grouped by test, sample ID, and imaging stage.
        </p>

        <div class="card" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex align-items-center justify-content-between">
            <h6 class="mb-0">Images</h6>
            <span class="badge bg-secondary">{len(images)} image{"s" if len(images)!=1 else ""}</span>
          </div>
          <div class="card-body p-3">
            {gallery_html}
          </div>
        </div>"""
        self.emit(body, f"CSAM Gallery — {p['name']}", active="projects",
                  project=p, active_sub="csam")


class ProjectCsamImageHandler(Base):
    """Serve a single CSAM image (full-size) or delete it."""

    def get(self, pid, iid):
        img = _db.get_csam_image(int(iid))
        if not img:
            self.send_error(404); return
        # Serve the image as a data URI in a minimal HTML page
        fname = img.get("filename") or "CSAM Image"
        data  = img.get("image_data", "")
        self.set_header("Content-Type", "text/html; charset=utf-8")
        self.write(f"""<!doctype html><html><head><title>{fname}</title>
        <style>body{{margin:0;background:#111;display:flex;align-items:center;
        justify-content:center;min-height:100vh}}
        img{{max-width:98vw;max-height:98vh;object-fit:contain}}</style>
        </head><body><img src="{data}" alt="{fname}"></body></html>""")


class ProjectCsamThumbHandler(Base):
    """Serve thumbnail (same image, browser caches it)."""

    def get(self, pid, iid):
        img = _db.get_csam_image(int(iid))
        if not img:
            self.send_error(404); return
        data = img.get("image_data", "")
        # Extract mime and raw bytes from data URI
        if data.startswith("data:"):
            try:
                header, b64data = data.split(",", 1)
                mime = header.split(";")[0].split(":")[1]
                raw  = base64.b64decode(b64data)
                self.set_header("Content-Type", mime)
                self.set_header("Cache-Control", "public, max-age=86400")
                self.write(raw)
                return
            except Exception:
                pass
        self.send_error(404)


class ProjectCsamDeleteHandler(Base):
    def post(self, pid, iid):
        _db.delete_csam_image(int(iid))
        self.redirect(f"/projects/{pid}/csam")


# ── Non-JEDEC test type registry ──────────────────────────────────────────────
_NONJEC_TYPES = {
    "pull_test": {"label": "Pull Test", "default_dur": 1},
    "xsem":      {"label": "X-SEM",     "default_dur": 1},
    "other":     {"label": "Other",      "default_dur": None},  # duration required
}

# Abbreviated stress test names used for GANTT labels ("Post-TC Pull Test", etc.)
_STRESS_ABBREV: dict[str, str] = {
    "precond": "Precond",
    "uhast":   "uHAST",
    "tc":      "TC",
    "tshock":  "T-Shock",
    "mshock":  "M-Shock",
    "vib":     "Vib",
    "hts":     "HTS",
}

def _nonjec_display_name(entry: dict) -> str:
    ttype = entry.get("test_type", "pull_test")
    if ttype == "other":
        return entry.get("custom_name", "").strip() or "Other Test"
    return _NONJEC_TYPES.get(ttype, {}).get("label", ttype)

def _nonjec_duration(entry: dict) -> int:
    ttype = entry.get("test_type", "pull_test")
    if ttype == "other":
        return max(1, int(entry.get("duration_weeks", 1)))
    return _NONJEC_TYPES.get(ttype, {}).get("default_dur", 1)


# ── Dynamic JEDEC qualification task template ─────────────────────────────────
# Durations are computed from sample counts stored in project_samples.
# Tasks run sequentially; durations round up to whole weeks (min 1).

def _compute_seeded_tasks(sample_counts: dict,
                           prescreen: list | None = None,
                           postqual:  list | None = None,
                           part_type: str = "ttv") -> list[dict]:
    """Build the qual plan with durations derived from sample counts.

    For TTV: includes SCD prep steps and post-stress testing (functional/thermal).
    For Die: skips TTV-specific prep and post-stress testing tasks (CSAM only).
    """
    is_ttv = (part_type != "die")
    def n(key):
        v = sample_counts.get(key, 0)
        return max(1, int(v)) if v else 1

    total = sum(
        int(v) for v in sample_counts.values()
        if isinstance(v, (int, float)) and v > 0
    ) or 1

    def cw(days):
        return max(1, math.ceil(days / 7))

    # ── Phase 1: sequential preparation steps ─────────────────────────────
    result, sw = [], 1
    prep = [
        ("SCD Surface Prep",
            "Preparation", "", 2),
        ("SCD Bonding + CSAM",
            "Preparation", "", cw(total / 4)),   # 1 day per 4 samples
    ]
    if is_ttv:
        # TTV Calibration: 10 TTVs/week (2/working day); all units must be calibrated
        ttv_cal_dur = max(1, math.ceil(total / 10))
        prep.append(("TTV Calibration", "Preparation", "", ttv_cal_dur))
    for name, cat, key, dur in prep:
        result.append({"task_name": name, "category": cat, "test_key": key,
                        "start_week": sw, "duration": dur})
        sw += dur - 1  # next prep starts on the last week of this one (≥ X overlap rule)

    # ── Phase 2: stress + post-stress pairs (all pairs start same week) ───
    # Each pair: (stress_name, stress_key, stress_dur,
    #             post_csam_name, post_test_name, test_key, post_test_dur)
    # Post steps split: CSAM (1 wk) → Testing (1 day/sample → weeks)
    stress_start = sw
    # Preconditioning gates all standard stress tests
    result.append({
        "task_name":  "Preconditioning",
        "category":   "Stress",
        "test_key":   "precond",
        "start_week": stress_start,
        "duration":   2,
        "n_mode":     "total",
    })

    # ── Non-JEDEC: Pre-screen tasks go right after Preconditioning ────────────
    # Labeled "Pre-{test_name}", all start at the same week as Preconditioning.
    # Post-qual entries with no parent are also treated as pre-screen.
    def _append_nonjec_prelike(entry: dict) -> None:
        base = _nonjec_display_name(entry)
        result.append({
            "task_name":  f"Pre-{base}",
            "category":   "Non-JEDEC Test",
            "test_key":   entry.get("test_type", "pull_test"),
            "start_week": stress_start,
            "duration":   _nonjec_duration(entry),
            "n_mode":     "custom",
            "n_custom":   int(entry.get("sample_count", 0)) or None,
        })

    for entry in (prescreen or []):
        _append_nonjec_prelike(entry)

    # Orphan post-qual entries (no parent stress selected) behave like pre-screen
    orphan_postqual = [e for e in (postqual or [])
                       if not e.get("parent_stress_test_key", "").strip()]
    for entry in orphan_postqual:
        _append_nonjec_prelike(entry)

    # ── Build a lookup: stress_key → list of post-qual entries with that parent ─
    parented_postqual: dict[str, list] = {}
    for e in (postqual or []):
        pk = e.get("parent_stress_test_key", "").strip()
        if pk:
            parented_postqual.setdefault(pk, []).append(e)

    pairs = [
        ("uHAST",      "uhast",  1, "Post-uHAST CSAM",   "Post-uHAST Testing",   "uhast",  cw(n("uhast"))),
        ("TC",         "tc",     3, "Post-TC CSAM",      "Post-TC Testing",      "tc",     cw(n("tc"))),
        ("T-Shock",    "tshock", 1, "Post-T-Shock CSAM", "Post-T-Shock Testing", "tshock", cw(n("tshock"))),
        ("M-Shock",    "mshock", 1, "Post-M-Shock CSAM", "Post-M-Shock Testing", "mshock", cw(n("mshock"))),
        ("Vibration",  "vib",    1, "Post-Vib CSAM",     "Post-Vib Testing",     "vib",    cw(n("vib"))),
        ("HTS",        "hts",    6, "Post-HTS CSAM",     "Post-HTS Testing",     "hts",    cw(n("hts"))),
    ]
    # Power Cycling and Shadow Moiré apply to Active and TTV, not Die
    if is_ttv:
        pairs.append(
            ("Pwr Cycling",  "pc",           2, "Post-PC CSAM",    "Post-PC Testing",    "pc",           cw(n("pc")))
        )
        pairs.append(
            # Shadow Moiré: characterization only — no CSAM or functional post-step
            ("Shadow Moiré", "shadow_moire", 1, "",                "",                   "shadow_moire", 0)
        )

    # Standard stress tests may start on the last week of Preconditioning (≥ X overlap rule)
    # Preconditioning duration = 2, so its last week is stress_start + 2 - 1 = stress_start + 1
    other_stress_start = stress_start + 1

    for (sname, skey, sdur, pcsam_name, ptest_name, tkey, ptest_dur) in pairs:
        stress_idx = len(result)          # remember position of this stress task
        result.append({"task_name": sname, "category": "Stress", "test_key": skey,
                        "start_week": other_stress_start, "duration": sdur, "n_mode": "test"})
        post_sw = other_stress_start + sdur - 1  # analysis may begin on the last week of stress
        last_analysis_end = post_sw                # tracks where the last child task ends
        if pcsam_name:
            result.append({"task_name": pcsam_name, "category": "Analysis", "test_key": tkey,
                            "start_week": post_sw, "duration": 1,
                            "n_mode": "test", "_parent_idx": stress_idx})
            post_sw += 1
            last_analysis_end = post_sw
        if ptest_name and is_ttv:
            # Post-stress testing (functional/thermal) is TTV-specific; omit for Die
            result.append({"task_name": ptest_name, "category": "Analysis", "test_key": tkey,
                            "start_week": post_sw, "duration": ptest_dur,
                            "n_mode": "test", "_parent_idx": stress_idx})
            last_analysis_end = post_sw + ptest_dur

        # ── Non-JEDEC: Post-qual tasks parented to this stress test ───────────
        # Inserted immediately after the last analysis row; start after it ends.
        abbrev = _STRESS_ABBREV.get(skey, skey.upper())
        for entry in parented_postqual.get(skey, []):
            base = _nonjec_display_name(entry)
            result.append({
                "task_name":  f"Post-{abbrev} {base}",
                "category":   "Non-JEDEC Test",
                "test_key":   entry.get("test_type", "pull_test"),
                "start_week": last_analysis_end,
                "duration":   _nonjec_duration(entry),
                "n_mode":     "custom",
                "n_custom":   int(entry.get("sample_count", 0)) or None,
                "_parent_idx": stress_idx,
            })

    return result

_GANTT_STATUS_COLORS = {
    "not_started": ("#9ca3af", "#6b7280"),  # (bar bg, text)
    "in_progress":  ("#f59e0b", "#92400e"),
    "complete":     ("#22c55e", "#14532d"),
    "blocked":      ("#ef4444", "#7f1d1d"),
    "na":           ("#e5e7eb", "#9ca3af"),
}
_GANTT_STATUS_LABELS = {
    "not_started": "Not Started",
    "in_progress":  "In Progress",
    "complete":     "Complete",
    "blocked":      "Blocked",
    "na":           "N/A",
}

# ── Project-scoped: GANTT Tracker ─────────────────────────────────────────────

class ProjectTrackerHandler(Base):
    def _get_project_or_404(self, pid):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404)
        return p

    def get(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        tasks      = _db.list_gantt_tasks(p["id"])
        has_hist   = _db.has_gantt_history(p["id"])
        flash      = self.get_argument("flash", "")
        body       = self._render_gantt(p, tasks, has_history=has_hist, flash=flash)
        self.emit(body, f"Schedule — {p['name']}", active="projects",
                  project=p, active_sub="tracker")

    def post(self, pid):
        p = self._get_project_or_404(pid)
        if not p: return
        action = self.get_argument("action", "")

        if action == "seed":
            sample_counts = _db.get_samples(p["id"])
            has_any = any(
                isinstance(v, (int, float)) and int(v) > 0
                for v in sample_counts.values()
            )
            if not sample_counts or not has_any:
                self.redirect(
                    f"/projects/{p['id']}/tracker"
                    "?flash=no_samples"
                )
                return
            _prescreen = _db.get_nonjec_prescreen(p["id"])
            _postqual  = _db.get_nonjec_postqual(p["id"])
            defaults = _compute_seeded_tasks(sample_counts, _prescreen, _postqual,
                                              part_type=p.get("part_type", "ttv"))
            _db.bulk_add_gantt_tasks(p["id"], defaults)
        elif action == "clear":
            existing = _db.list_gantt_tasks(p["id"])
            if existing:
                _db.push_gantt_history(p["id"], "clear", existing)
            _db.clear_gantt_tasks(p["id"])
        elif action == "set_start":
            sd = self.get_argument("start_date", "").strip()
            if sd:
                _db.save_meta(p["id"], gantt_start_date=sd)
        elif action == "add":
            _n_mode = self.get_argument("n_mode", "auto").strip() or "auto"
            _n_cust_raw = self.get_argument("n_custom", "").strip()
            _n_cust = int(_n_cust_raw) if _n_cust_raw.isdigit() else None
            _par_raw2 = self.get_argument("parent_task_id", "").strip()
            _par2 = int(_par_raw2) if _par_raw2.isdigit() else None
            _add_anchor   = _get_gantt_anchor(p["id"])
            _iso_sw_raw   = self.get_argument("start_week", "").strip()
            _iso_sw       = int(_iso_sw_raw) if _iso_sw_raw.isdigit() else _add_anchor.isocalendar()[1]
            _rel_sw       = _iso_week_to_relative(_add_anchor, _iso_sw)
            _db.add_gantt_task(
                p["id"],
                task_name  = self.get_argument("task_name", "New Task"),
                category   = self.get_argument("category", ""),
                start_week = _rel_sw,
                duration   = int(self.get_argument("duration", "1")),
                status     = self.get_argument("status", "not_started"),
                n_mode     = _n_mode,
                n_custom   = _n_cust,
                parent_task_id = _par2,
            )
        elif action == "undo":
            entry = _db.pop_gantt_history(p["id"])
            if entry:
                atype    = entry["action_type"]
                snapshot = entry["snapshot"]
                if atype == "delete_task":
                    t = snapshot
                    _db.add_gantt_task(
                        p["id"], task_name=t["task_name"], category=t.get("category",""),
                        test_key=t.get("test_key",""),
                        start_week=t["start_week"], duration=t["duration"],
                        status=t["status"],
                        n_mode=t.get("n_mode","auto") or "auto",
                        n_custom=t.get("n_custom"),
                        parent_task_id=t.get("parent_task_id"),
                    )
                elif atype == "edit_task":
                    t = snapshot
                    _db.update_gantt_task(
                        t["id"], p["id"],
                        task_name=t["task_name"], category=t.get("category",""),
                        test_key=t.get("test_key",""),
                        start_week=t["start_week"], duration=t["duration"],
                        status=t["status"],
                        n_mode=t.get("n_mode","auto") or "auto",
                        n_custom=t.get("n_custom"),
                    )
                elif atype == "clear":
                    _db.clear_gantt_tasks(p["id"])
                    for t in snapshot:
                        _db.add_gantt_task(
                            p["id"], task_name=t["task_name"], category=t.get("category",""),
                            test_key=t.get("test_key",""),
                            start_week=t["start_week"], duration=t["duration"],
                            status=t["status"],
                            n_mode=t.get("n_mode","auto") or "auto",
                            n_custom=t.get("n_custom"),
                            parent_task_id=t.get("parent_task_id"),
                        )
        elif action == "test_start":
            from datetime import datetime as _dt
            test_key = self.get_argument("test_key", "").strip()
            if test_key:
                conds    = _db.get_test_conditions(p["id"])
                cond_key = conds.get(test_key, "")
                dur      = _CONDITION_HOURS.get(test_key, {}).get(cond_key, None)
                _db.upsert_test_tracker(
                    p["id"], test_key,
                    status="active",
                    started_at=_dt.utcnow().isoformat(timespec="seconds"),
                    duration_hours=dur,
                )
        elif action == "test_complete":
            from datetime import datetime as _dt
            test_key = self.get_argument("test_key", "").strip()
            if test_key:
                _db.upsert_test_tracker(
                    p["id"], test_key,
                    status="complete",
                    completed_at=_dt.utcnow().isoformat(timespec="seconds"),
                )
        elif action == "test_uncomplete":
            test_key = self.get_argument("test_key", "").strip()
            if test_key:
                _db.upsert_test_tracker(
                    p["id"], test_key,
                    status="active",
                    completed_at=None,
                )
        elif action == "test_cancel":
            test_key = self.get_argument("test_key", "").strip()
            if test_key:
                _db.upsert_test_tracker(
                    p["id"], test_key,
                    status="pending",
                    started_at=None,
                    completed_at=None,
                    duration_hours=None,
                )
        elif action == "bulk_edit":
            import json as _json
            try:
                changes = _json.loads(self.get_argument("changes", "[]"))
                if changes:
                    existing = _db.list_gantt_tasks(p["id"])
                    _db.push_gantt_history(p["id"], "bulk_edit", existing)
                    for ch in changes:
                        if ch.get("duration", 0) > 0:
                            _db.update_gantt_task(
                                int(ch["id"]), p["id"],
                                start_week=int(ch["start_week"]),
                                duration=int(ch["duration"]),
                            )
            except Exception:
                pass
        elif action == "reorder":
            import json as _json
            try:
                ordered_ids = _json.loads(self.get_argument("order", "[]"))
                if ordered_ids:
                    _db.reorder_gantt_tasks(p["id"], [int(i) for i in ordered_ids])
            except Exception:
                pass
            # Respond with 204 (no redirect — called via fetch)
            self.set_status(204)
            return
        self.redirect(f"/projects/{p['id']}/tracker")

    def _render_gantt(self, p: dict, tasks: list, has_history: bool = False, flash: str = "") -> str:
        from datetime import date, timedelta, datetime as _dt
        pid = p["id"]

        # ── Calendar anchor ───────────────────────────────────────────────
        meta = _db.get_meta(pid)
        start_date_str = (meta.get("gantt_start_date") or "").strip()
        try:
            anchor = date.fromisoformat(start_date_str)
        except ValueError:
            anchor = date.today()
            # align to Monday of current week
            anchor = anchor - timedelta(days=anchor.weekday())

        today = date.today()
        # Which week number is today (1-indexed from anchor)?
        days_since = (today - anchor).days
        current_week = max(1, days_since // 7 + 1) if days_since >= 0 else None
        today_day_idx = days_since if days_since >= 0 else -1

        # ISO-week values passed to JS for form display/conversion
        _epoch = date(1970, 1, 1)
        anchor_epoch_ms  = int((anchor - _epoch).total_seconds() * 1000)
        current_iso_week = today.isocalendar()[1]
        anchor_iso_week  = anchor.isocalendar()[1]
        anchor_year      = anchor.isocalendar()[0]

        # ── Chart width ────────────────────────────────────────────────────
        if tasks:
            max_week = max(t["start_week"] + t["duration"] - 1 for t in tasks)
            n_weeks  = max(max_week + 2, 26)
        else:
            n_weeks = 26

        # ── Per-test sample counts ─────────────────────────────────────────
        sample_counts = _db.get_samples(pid)
        # Values may be plain ints (from planner) or lists (from report); normalise
        def _extract_n(raw):
            if isinstance(raw, list):
                return len(raw) if raw else 0
            try:
                v = int(raw)
                return v if v > 0 else 0
            except (TypeError, ValueError):
                return 0
        total_n = sum(_extract_n(v) for v in sample_counts.values()) if sample_counts else 0

        # ── Status badge helper ────────────────────────────────────────────
        def status_badge(s):
            bg, fg = _GANTT_STATUS_COLORS.get(s, ("#9ca3af", "#6b7280"))
            label  = _GANTT_STATUS_LABELS.get(s, s)
            return (f'<span style="background:{bg};color:{fg};padding:2px 8px;'
                    f'border-radius:4px;font-size:.72rem;font-weight:600">{label}</span>')

        # ── Category colors ────────────────────────────────────────────────
        _cat_colors = ["#3b82f6","#8b5cf6","#ec4899","#f97316","#14b8a6","#64748b"]
        cats      = list(dict.fromkeys(t["category"] for t in tasks if t["category"]))
        cat_color = {c: _cat_colors[i % len(_cat_colors)] for i, c in enumerate(cats)}

        # ── Week → date helper ────────────────────────────────────────────
        def week_date(w):
            return anchor + timedelta(weeks=w - 1)

        # ── Month/year header rows ─────────────────────────────────────────
        # Group consecutive weeks by month label "Jan 2026"
        month_spans = []  # list of (label, colspan)
        cur_label, span = None, 0
        for w in range(1, n_weeks + 1):
            d = week_date(w)
            lbl = d.strftime("%b %Y")
            if lbl == cur_label:
                span += 1
            else:
                if cur_label:
                    month_spans.append((cur_label, span))
                cur_label, span = lbl, 1
        if cur_label:
            month_spans.append((cur_label, span))

        month_headers = ""
        for lbl, cs in month_spans:
            month_headers += (
                f'<th colspan="{cs}" style="padding:3px 6px;text-align:center;'
                f'font-size:.72rem;color:#374151;font-weight:700;'
                f'border-left:2px solid #d1d5db;white-space:nowrap">{lbl}</th>'
            )

        # ── Week-number header row ─────────────────────────────────────────
        wk_headers = ""
        for w in range(1, n_weeks + 1):
            d     = week_date(w)
            label = f"W{d.isocalendar()[1]}"   # ISO week number
            is_now = (current_week == w)
            bg   = "#fef9c3" if is_now else "transparent"
            border = "border-left:2px solid #f59e0b;" if is_now else "border-left:1px solid #e5e7eb;"
            wk_headers += (
                f'<th style="min-width:34px;padding:3px 2px;text-align:center;'
                f'font-size:.68rem;color:{"#92400e" if is_now else "#9ca3af"};'
                f'font-weight:{"700" if is_now else "400"};background:{bg};{border}">'
                f'{label}</th>'
            )

        # ── Constraint computation ──────────────────────────────────────────
        import json as _json_
        from collections import defaultdict as _dd_
        _prep_sorted  = [t for t in tasks if (t.get("category") or "") == "Preparation"]
        _stress_gnt   = [t for t in tasks if (t.get("category") or "") == "Stress"]
        _analysis_gnt = [t for t in tasks if (t.get("category") or "") == "Analysis"]

        _prep_chain_end = max(
            (t["start_week"] + t["duration"] - 1 for t in _prep_sorted), default=0
        )
        _stress_end_by_id = {t["id"]: t["start_week"] + t["duration"] - 1 for t in _stress_gnt}

        # Preconditioning = Stress tasks with test_key "precond";
        # all other Stress tasks are gated by Preconditioning end.
        _precond_ids = {t["id"] for t in _stress_gnt if (t.get("test_key") or "") == "precond"}
        _precond_end = max(
            (_stress_end_by_id[i] for i in _precond_ids if i in _stress_end_by_id),
            default=_prep_chain_end
        )

        _ab = _dd_(list)
        for _at in _analysis_gnt:
            _ap = _at.get("parent_task_id")
            if _ap: _ab[_ap].append(_at)
        _reporting_gate = (
            min(max(_at["start_week"] + _at["duration"] - 1 for _at in _g) for _g in _ab.values())
            if _ab else _prep_chain_end
        )

        _task_data_py = {}
        for t in tasks:
            _cat = (t.get("category") or "").strip()
            _tid = t["id"]
            _bbg, _ = _GANTT_STATUS_COLORS.get(t["status"], ("#9ca3af", "#6b7280"))
            _ent = {
                "category":       _cat,
                "test_key":       (t.get("test_key") or ""),
                "start_week":     t["start_week"],
                "duration":       t["duration"],
                "color":          _bbg,
                "parent_task_id": t.get("parent_task_id"),
                "locked_start":   False,
                "min_start":      1,
                "prep_idx":       -1,
            }
            if _cat == "Preparation":
                _ix = next((i for i, p in enumerate(_prep_sorted) if p["id"] == _tid), 0)
                _ent["locked_start"] = True
                _ent["prep_idx"]     = _ix
                _ent["min_start"]    = (
                    1 if _ix == 0
                    else _prep_sorted[_ix-1]["start_week"] + _prep_sorted[_ix-1]["duration"] - 1
                )  # ≥ X rule: can start on the last week of the previous prep task
            elif _cat == "Stress":
                # Preconditioning gated by prep; all others gated by Preconditioning end
                # Overlap on the final week of the gating step is allowed
                if _tid in _precond_ids:
                    _ent["min_start"] = _prep_chain_end
                else:
                    _ent["min_start"] = _precond_end
            elif _cat == "Analysis":
                _pid2 = t.get("parent_task_id")
                _ent["min_start"] = (_stress_end_by_id.get(_pid2, _prep_chain_end) if _pid2 else _prep_chain_end)
            elif _cat == "Non-JEDEC Test":
                _pid2 = t.get("parent_task_id")
                if _pid2:
                    # Post-qual: may start on the same week the parent stress task ends
                    _ent["min_start"] = _stress_end_by_id.get(_pid2, _prep_chain_end)
                else:
                    # Pre-screen: gated by prep chain end
                    _ent["min_start"] = _prep_chain_end
            elif _cat == "Reporting":
                _ent["min_start"] = _reporting_gate
            _task_data_py[_tid] = _ent

        _task_data_js        = _json_.dumps({str(k): v for k, v in _task_data_py.items()})
        _prep_order_js       = _json_.dumps([t["id"] for t in _prep_sorted])
        _stress_tasks_js     = _json_.dumps([{"id": t["id"], "name": t["task_name"]} for t in _stress_gnt])
        _stress_end_by_id_js = _json_.dumps({str(k): v for k, v in _stress_end_by_id.items()})
        _stress_parent_opts = '<option value="">— select parent stress —</option>' + "".join(
            f'<option value="{t["id"]}">{t["task_name"]}</option>' for t in _stress_gnt
        )

        # Stress tests by key for dropdown population
        _proj_stress_tests = {k: v["name"] for k, v in applicable_tests(p.get("part_type", "active")).items()}
        _proj_stress_tests_js = _json_.dumps(_proj_stress_tests)

        # ── Task sidebar rows ──────────────────────────────────────────────
        if tasks:
            task_rows = ""
            for t in tasks:
                tid_      = t["id"]
                cc        = cat_color.get(t["category"], "#64748b")
                safe_name = t["task_name"].replace("'", "\\'")
                cat_hint  = ("· " + t["category"]) if t["category"] else ""
                tkey      = t.get("test_key", "") or ""
                n_mode_t  = t.get("n_mode", "auto") or "auto"
                n_cust_t  = t.get("n_custom", None)
                cat_t     = (t.get("category") or "").strip()
                # Resolve 'auto' using category-based defaults
                _CAT_TOTAL = {"Preparation", "Reporting"}
                if n_mode_t == "auto":
                    n_mode_t = "total" if cat_t in _CAT_TOTAL else "test"
                if n_mode_t == "total":
                    n_samp = str(total_n) if total_n > 0 else "TBD"
                elif n_mode_t == "test":
                    raw   = sample_counts.get(tkey, None) if tkey else None
                    n_val = _extract_n(raw) if raw is not None else 0
                    n_samp = str(n_val) if n_val > 0 else "TBD"
                elif n_mode_t == "custom":
                    try:
                        cv = int(n_cust_t)
                        n_samp = str(cv) if cv > 0 else "TBD"
                    except (TypeError, ValueError):
                        n_samp = "TBD"
                else:
                    n_samp = "TBD"
                # Always show the n= pill; grey it out when TBD
                if n_samp == "TBD":
                    samp_pill = (
                        f'<span style="background:#f3f4f6;color:#9ca3af;border:1px solid #e5e7eb;'
                        f'border-radius:10px;font-size:.68rem;padding:1px 6px;'
                        f'font-weight:600;white-space:nowrap">n=TBD</span>'
                    )
                else:
                    samp_pill = (
                        f'<span style="background:#f0fdf4;color:#15803d;border:1px solid #bbf7d0;'
                        f'border-radius:10px;font-size:.68rem;padding:1px 6px;'
                        f'font-weight:600;white-space:nowrap">n={n_samp}</span>'
                    )
                # Second line: category dot + n= pill (always visible below the name)
                meta_line = ""
                if cat_hint or samp_pill:
                    sep = " &nbsp;·&nbsp; " if cat_hint and samp_pill else ""
                    meta_line = (f'<span style="font-size:.7rem;color:#9ca3af;display:block;'
                                 f'padding-left:14px;margin-top:1px">'
                                 f'{cat_hint}{sep}{samp_pill}</span>')
                task_rows += (
                    f'<tr id="tr-{tid_}" onclick="ganttSelectRow({tid_})" style="cursor:default">'
                    f'<td style="padding:6px 8px;max-width:220px" title="{t["task_name"]}">'
                    f'<span class="gantt-drag-handle" title="Drag to reorder" '
                    f'style="cursor:grab;color:#d1d5db;margin-right:4px;font-size:.85rem;'
                    f'vertical-align:middle;user-select:none">⠿</span>'
                    f'<span style="display:inline-block;width:9px;height:9px;'
                    f'border-radius:50%;background:{cc};margin-right:5px;flex-shrink:0;vertical-align:middle"></span>'
                    f'<span style="font-size:.82rem;white-space:nowrap;overflow:hidden;'
                    f'text-overflow:ellipsis;max-width:175px;display:inline-block;vertical-align:middle">'
                    f'{t["task_name"]}</span>'
                    f'{meta_line}'
                    f'</td>'
                    f'<td style="padding:6px 4px;text-align:center;white-space:nowrap">{status_badge(t["status"])}</td>'
                    f'<td style="padding:6px 4px;text-align:center;white-space:nowrap">'
                    f'<button class="btn btn-sm" style="padding:2px 7px;font-size:.73rem;'
                    f'border:1px solid #e0e0e0;border-radius:4px;background:#fff" '
                    f"onclick=\"event.stopPropagation();"
                    f"openEditModal({tid_},'{safe_name}','{t['category']}',"
                    f"{t['start_week']},{t['duration']},'{t['status']}',"
                    f"'{t.get('n_mode','auto') or 'auto'}',"
                    f"{t['n_custom'] if t.get('n_custom') is not None else 'null'},"
                    f"{t.get('parent_task_id') or 'null'},"
                    f"'{t.get('test_key','') or ''}')\""
                    f'>Edit</button> '
                    f'<form method="post" action="/projects/{pid}/tracker/task/{tid_}/delete"'
                    f' class="d-inline" onsubmit="return confirm(\'Delete task?\')">'
                    f'<button class="btn btn-sm" style="padding:2px 7px;font-size:.73rem;'
                    f'border:1px solid #fca5a5;border-radius:4px;background:#fff;color:#ef4444"'
                    f' onclick="event.stopPropagation()"'
                    f'>Del</button></form>'
                    f'</td></tr>'
                )
        else:
            task_rows = ('<tr><td colspan="3" class="text-center text-muted py-4"'
                         ' style="font-size:.85rem">No tasks yet.</td></tr>')

        # ── GANTT bar rows ─────────────────────────────────────────────────
        gantt_rows = ""
        for t in tasks:
            cells = ""
            for w in range(1, n_weeks + 1):
                in_range = t["start_week"] <= w <= t["start_week"] + t["duration"] - 1
                is_start = w == t["start_week"]
                is_end   = w == t["start_week"] + t["duration"] - 1
                is_now   = (current_week == w)
                now_attr = ' data-now="1"' if is_now else ""
                if in_range:
                    bg, _ = _GANTT_STATUS_COLORS.get(t["status"], ("#9ca3af","#6b7280"))
                    br = ("border-radius:4px;" if (is_start and is_end)
                          else "border-radius:4px 0 0 4px;" if is_start
                          else "border-radius:0 4px 4px 0;" if is_end else "")
                    now_border = "border-left:2px solid #f59e0b;" if is_now else ""
                    cells += (f'<td data-tid="{t["id"]}" data-week="{w}"{now_attr}'
                               f' style="background:{bg};{br}{now_border}'
                               f'padding:0;height:26px"></td>')
                else:
                    bg_cell = "#fef9c3" if is_now else "#f9fafb"
                    bl      = "border-left:2px solid #f59e0b;" if is_now else (
                              "border-left:2px solid #d1d5db;" if w % 4 == 1 else
                              "border-left:1px solid #e5e7eb;")
                    cells += (f'<td data-tid="{t["id"]}" data-week="{w}"{now_attr}'
                               f' style="background:{bg_cell};{bl}padding:0;height:26px"></td>')
            gantt_rows += f'<tr data-gantt-row="{t["id"]}">{cells}</tr>'

        if not tasks:
            gantt_rows = (f'<tr><td colspan="{n_weeks}" class="text-center text-muted py-4"'
                          f' style="font-size:.85rem">Add tasks to see the chart.</td></tr>')

        # ── Options HTML ───────────────────────────────────────────────────
        status_opts_html = "".join(
            f'<option value="{k}">{v}</option>'
            for k, v in _GANTT_STATUS_LABELS.items()
        )

        # ── Legend ─────────────────────────────────────────────────────────
        legend = "".join(
            f'<span style="display:inline-flex;align-items:center;margin-right:12px;font-size:.75rem">'
            f'<span style="width:12px;height:12px;border-radius:3px;background:{bg};'
            f'display:inline-block;margin-right:4px"></span>{_GANTT_STATUS_LABELS[s]}</span>'
            for s, (bg, _fg) in _GANTT_STATUS_COLORS.items()
        )
        cat_legend = "".join(
            f'<span style="display:inline-flex;align-items:center;margin-right:12px;font-size:.75rem">'
            f'<span style="width:10px;height:10px;border-radius:50%;background:{cc};'
            f'display:inline-block;margin-right:4px"></span>{c}</span>'
            for c, cc in cat_color.items()
        )

        # ── Buttons ────────────────────────────────────────────────────────
        undo_btn = (
            f'<form method="post" action="/projects/{pid}/tracker" class="d-inline" id="undoForm">'
            f'<input type="hidden" name="action" value="undo">'
            f'<button type="submit" class="btn btn-sm ms-1" id="undoBtn" '
            f'style="border:1px solid #93c5fd;color:#1d4ed8;background:#eff6ff;font-size:.8rem">'
            f'<i class="bi bi-arrow-counterclockwise me-1"></i>Undo <kbd style="font-size:.68rem;'
            f'background:#dbeafe;border:1px solid #93c5fd;border-radius:3px;padding:0 4px">⌘Z</kbd>'
            f'</button></form>'
        ) if has_history else ""

        seed_btn = "" if tasks else (
            f'<form method="post" action="/projects/{pid}/tracker" class="d-inline ms-2">'
            f'<input type="hidden" name="action" value="seed">'
            f'<button class="btn btn-sm" style="background:var(--df-accent);color:#fff;'
            f'border:none;font-size:.8rem"><i class="bi bi-magic me-1"></i>Seed from JEDEC Template'
            f'</button></form>'
        )
        clear_btn = "" if not tasks else (
            f'<form method="post" action="/projects/{pid}/tracker" class="d-inline ms-2"'
            f' onsubmit="return confirm(\'Clear all tasks from this schedule? This cannot be undone.\')">'
            f'<input type="hidden" name="action" value="clear">'
            f'<button class="btn btn-sm" style="border:1px solid #fca5a5;color:#ef4444;'
            f'background:#fff;font-size:.8rem"><i class="bi bi-trash me-1"></i>Clear Schedule'
            f'</button></form>'
        )

        # ── Test status overview (stress tests) ───────────────────────────
        tracker_data  = _db.get_test_tracker(pid)
        saved_conds_ov = _db.get_test_conditions(pid)
        # Collect unique Stress-category tasks from GANTT (maintain order)
        seen_tkeys: set = set()
        stress_tasks: list = []
        for t in tasks:
            if (t.get("category") or "").strip() != "Stress":
                continue
            tkey = (t.get("test_key") or "").strip()
            if tkey and tkey not in seen_tkeys:
                seen_tkeys.add(tkey)
                stress_tasks.append(t)
            elif not tkey:
                # Show Stress tasks without a test_key (custom tasks) by task id
                stress_tasks.append(t)

        ov_cards = ""
        for t in stress_tasks:
            tkey       = t["test_key"]
            test_info  = TESTS.get(tkey, {})
            test_name  = test_info.get("name", t["task_name"])
            cond_key   = saved_conds_ov.get(tkey, "")
            cond_opts  = _TEST_CONDITION_OPTIONS.get(tkey, [])
            cond_label = next((lbl for k, lbl in cond_opts if k == cond_key), "")
            tr         = tracker_data.get(tkey, {})
            tr_status  = tr.get("status", "pending")
            started_at    = tr.get("started_at",    "") or ""
            completed_at  = tr.get("completed_at",  "") or ""
            dur_hours     = tr.get("duration_hours", None)

            task_start_week = t["start_week"]
            is_reachable    = (current_week is not None and
                               current_week >= task_start_week)

            def _fmt_date(s):
                try:
                    return date.fromisoformat(s[:10]).strftime("%-d %b %Y")
                except Exception:
                    return s[:10] if s else "—"

            if tr_status == "complete":
                hdr_bg, hdr_fg  = "#d1fae5", "#065f46"
                hdr_text        = "✓ Complete"
                status_body     = (
                    f'<div style="font-size:.72rem;color:#374151;margin-top:4px">'
                    f'<div>Started: <strong>{_fmt_date(started_at)}</strong></div>'
                    f'<div>Completed: <strong>{_fmt_date(completed_at)}</strong></div>'
                    f'</div>'
                )
                action_html     = (
                    f'<form method="post" action="/projects/{pid}/tracker" class="mt-2">'
                    f'<input type="hidden" name="action" value="test_uncomplete">'
                    f'<input type="hidden" name="test_key" value="{tkey}">'
                    f'<button type="submit" class="btn btn-sm w-100"'
                    f' style="background:#fff;color:#6b7280;border:1px solid #d1d5db;'
                    f'font-size:.7rem;padding:2px 0">↩ Undo Complete</button>'
                    f'</form>'
                )
            elif tr_status == "active":
                hdr_bg, hdr_fg = "#fff7ed", "#9a3412"
                hdr_text       = "▶ In Progress"
                if started_at and dur_hours:
                    try:
                        elapsed_h = (_dt.utcnow() -
                                     _dt.fromisoformat(started_at)).total_seconds() / 3600
                    except Exception:
                        elapsed_h = 0.0
                    pct       = min(100.0, elapsed_h / dur_hours * 100.0)
                    bar_color = "#16a34a" if pct >= 100.0 else "#f97316"
                    status_body = (
                        f'<div style="background:#e5e7eb;border-radius:4px;height:8px;'
                        f'margin-top:6px;overflow:hidden">'
                        f'<div class="test-progress-bar" '
                        f'data-started="{started_at}" data-duration="{dur_hours}"'
                        f' style="width:{pct:.1f}%;background:{bar_color};height:100%;'
                        f'border-radius:4px;transition:width .5s"></div>'
                        f'</div>'
                        f'<div class="test-progress-txt" style="font-size:.7rem;color:#6b7280;margin-top:2px">'
                        f'{elapsed_h:.1f}h / {dur_hours:.0f}h</div>'
                    )
                else:
                    status_body = '<div style="font-size:.72rem;color:#6b7280;margin-top:4px">In progress</div>'
                action_html = (
                    f'<form method="post" action="/projects/{pid}/tracker" class="mt-2 mb-1">'
                    f'<input type="hidden" name="action" value="test_complete">'
                    f'<input type="hidden" name="test_key" value="{tkey}">'
                    f'<button type="submit" class="btn btn-sm w-100"'
                    f' style="background:#16a34a;color:#fff;border:none;'
                    f'font-size:.72rem;padding:3px 0">Mark Complete</button>'
                    f'</form>'
                    f'<form method="post" action="/projects/{pid}/tracker">'
                    f'<input type="hidden" name="action" value="test_cancel">'
                    f'<input type="hidden" name="test_key" value="{tkey}">'
                    f'<button type="submit" class="btn btn-sm w-100"'
                    f' style="background:#fff;color:#dc2626;border:1px solid #fca5a5;'
                    f'font-size:.72rem;padding:3px 0">Cancel</button>'
                    f'</form>'
                )
            else:
                _sd         = week_date(task_start_week)
                _iso_wk     = _sd.isocalendar()[1]
                _sched_date = _sd.strftime("%-d %b")
                hdr_bg, hdr_fg = "#dbeafe", "#1e40af"
                hdr_text       = "Pending"
                status_body    = (
                    f'<div style="font-size:.72rem;color:#6b7280;margin-top:4px">'
                    f'Scheduled: W{_iso_wk} ({_sched_date})</div>'
                )
                action_html    = (
                    f'<form method="post" action="/projects/{pid}/tracker" class="mt-2">'
                    f'<input type="hidden" name="action" value="test_start">'
                    f'<input type="hidden" name="test_key" value="{tkey}">'
                    f'<button type="submit" class="btn btn-sm w-100"'
                    f' style="background:var(--df-accent);color:#fff;border:none;'
                    f'font-size:.72rem;padding:3px 0">Start Test</button>'
                    f'</form>'
                )

            cond_display = (
                f'<div style="font-size:.7rem;color:#6b7280;margin-top:2px">{cond_label}</div>'
                if cond_label else ""
            )
            ov_cards += (
                f'<div style="min-width:195px;max-width:195px;border:1px solid var(--df-border);'
                f'border-radius:8px;overflow:hidden;flex-shrink:0">'
                f'<div style="background:{hdr_bg};color:{hdr_fg};padding:5px 10px;'
                f'font-size:.72rem;font-weight:600">{hdr_text}</div>'
                f'<div style="padding:8px 10px">'
                f'<div style="font-size:.82rem;font-weight:600;color:var(--df-navy)">{test_name}</div>'
                f'{cond_display}'
                f'{status_body}'
                f'{action_html}'
                f'</div></div>'
            )

        overview_html = ""
        if ov_cards:
            overview_html = f"""
        <div class="card mb-3" style="border:1px solid var(--df-border)">
          <div class="card-df d-flex align-items-center gap-2">
            <h6 class="mb-0" style="font-size:.82rem">Test Status Overview</h6>
            <span class="text-white-50" style="font-size:.7rem">Stress tests only</span>
          </div>
          <div class="card-body p-3">
            <div style="display:flex;gap:10px;overflow-x:auto;padding-bottom:4px">
              {ov_cards}
            </div>
          </div>
        </div>"""

        # ── Start-date picker ──────────────────────────────────────────────
        date_picker = (
            f'<form method="post" action="/projects/{pid}/tracker" '
            f'class="d-inline-flex align-items-center ms-3 gap-2">'
            f'<input type="hidden" name="action" value="set_start">'
            f'<label style="font-size:.75rem;color:var(--df-grey);white-space:nowrap">Start date</label>'
            f'<input type="date" name="start_date" value="{anchor.isoformat()}" '
            f'class="form-control form-control-sm" style="width:145px;font-size:.8rem">'
            f'<button type="submit" class="btn btn-sm" '
            f'style="border:1px solid var(--df-border);font-size:.78rem;white-space:nowrap">Set</button>'
            f'</form>'
        )

        return f"""
        <!-- Edit-task modal -->
        <div class="modal fade" id="editModal" tabindex="-1">
          <div class="modal-dialog">
            <div class="modal-content">
              <form method="post" id="editForm">
                <div class="modal-header">
                  <h6 class="modal-title mb-0">Edit Task</h6>
                  <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body">
                  <div class="mb-3">
                    <label class="form-label" style="font-size:.83rem">Category</label>
                    <select class="form-select form-select-sm" name="category" id="e_cat"
                            onchange="onEditCatChange()">
                      <option value="">— select —</option>
                      <option value="Preparation">Preparation</option>
                      <option value="Stress">Stress</option>
                      <option value="Analysis">Analysis</option>
                      <option value="Reporting">Reporting</option>
                      <option value="Non-JEDEC Test">Non-JEDEC Test</option>
                    </select>
                  </div>
                  <div class="mb-3" id="e_name_div">
                    <label class="form-label" style="font-size:.83rem">Task Name</label>
                    <input type="text" class="form-control form-control-sm" name="task_name" id="e_name" required>
                  </div>
                  <div class="mb-3" id="e_stress_div" style="display:none">
                    <label class="form-label" style="font-size:.83rem">Stress Test</label>
                    <select class="form-select form-select-sm" id="e_stress_sel" onchange="onEditStressSelChange()">
                      <option value="">— select test —</option>
                      <option value="_other">Other Test...</option>
                    </select>
                  </div>
                  <input type="hidden" name="test_key" id="e_test_key">
                  <div class="row g-2">
                    <div class="col">
                      <label class="form-label" style="font-size:.83rem">Start Wk (ISO)</label>
                      <input type="number" class="form-control form-control-sm" name="start_week" id="e_sw"
                             min="{current_iso_week}" max="53">
                    </div>
                    <div class="col">
                      <label class="form-label" style="font-size:.83rem">Duration (weeks)</label>
                      <input type="number" class="form-control form-control-sm" name="duration" id="e_dur"
                             min="1" max="104">
                    </div>
                  </div>
                  <div class="mt-3">
                    <label class="form-label" style="font-size:.83rem">Status</label>
                    <select class="form-select form-select-sm" name="status" id="e_status">
                      {status_opts_html}
                    </select>
                  </div>
                  <div class="mt-3">
                    <label class="form-label" style="font-size:.83rem">Sample Size (n)</label>
                    <select class="form-select form-select-sm" name="n_mode" id="e_nmode"
                            onchange="toggleNcust()">
                      <option value="auto">Auto — test key → per-test, else total</option>
                      <option value="total">Total — all tests combined</option>
                      <option value="test">This test only</option>
                      <option value="custom">Custom number</option>
                    </select>
                  </div>
                  <div class="mt-2" id="e_ncust_div" style="display:none">
                    <label class="form-label" style="font-size:.8rem">Custom n</label>
                    <input type="number" class="form-control form-control-sm" name="n_custom"
                           id="e_ncust" min="1" placeholder="e.g. 45">
                  </div>
                  <div class="mt-3" id="e_parent_div" style="display:none">
                    <label class="form-label" style="font-size:.83rem">Parent Stress Task</label>
                    <select class="form-select form-select-sm" name="parent_task_id" id="e_parent_sel">
                      <option value="">— select —</option>
                    </select>
                  </div>
                </div>
                <div class="modal-footer">
                  <button type="button" class="btn btn-sm btn-outline-secondary" data-bs-dismiss="modal">Cancel</button>
                  <button type="submit" class="btn btn-sm"
                    style="background:var(--df-accent);color:#fff;border:none">Save</button>
                </div>
              </form>
            </div>
          </div>
        </div>

        {
            '''<div class="alert alert-warning d-flex align-items-center gap-2 mb-3 py-2 px-3" role="alert" style="font-size:.85rem">
          <i class="bi bi-exclamation-triangle-fill flex-shrink-0"></i>
          <span>No sample counts found. Please add sample counts in the
          <a href="/projects/''' + str(pid) + '''/samples" class="alert-link">Sample Planner</a>
          before seeding the schedule.</span>
        </div>''' if flash == "no_samples" else ""
        }
        <div class="d-flex align-items-center justify-content-between mb-3 flex-wrap gap-2">
          <h5 class="mb-0" style="font-weight:300">Schedule — {p['name']}</h5>
          <div class="d-flex align-items-center flex-wrap gap-1">
            {date_picker}
            {undo_btn}
            {seed_btn}
            {clear_btn}
            <button id="ganttEditBtn" class="btn btn-sm ms-1" onclick="ganttEnterEdit()"
              style="border:1px solid #c4b5fd;color:#6d28d9;background:#fff;font-size:.8rem">
              <i class="bi bi-pencil-square me-1"></i>Edit
            </button>
            <button id="ganttSaveBtn" class="btn btn-sm ms-1" onclick="ganttSave()"
              style="display:none;background:#16a34a;color:#fff;border:none;font-size:.8rem">
              <i class="bi bi-check2 me-1"></i>Save
            </button>
            <button id="ganttCancelBtn" class="btn btn-sm ms-1" onclick="ganttCancelEdit()"
              style="display:none;border:1px solid #fca5a5;color:#dc2626;background:#fff;font-size:.8rem">
              <i class="bi bi-x-lg me-1"></i>Cancel Edit
            </button>
            <button id="ganttDelSelBtn" class="btn btn-sm ms-1" onclick="ganttDeleteSel()"
              style="display:none;border:1px solid #fca5a5;color:#dc2626;background:#fff;font-size:.8rem">
              <i class="bi bi-eraser me-1"></i>Delete
            </button>
            <button class="btn btn-sm ms-1" style="border:1px solid var(--df-border);font-size:.8rem"
              data-bs-toggle="collapse" data-bs-target="#addTaskPanel">
              <i class="bi bi-plus-lg me-1"></i>Add Task
            </button>
            <a href="/projects/{pid}/tracker/xlsx"
               class="btn btn-sm ms-1" style="border:1px solid var(--df-border);font-size:.8rem"
               download>
              <i class="bi bi-file-earmark-spreadsheet me-1"></i>Export Excel
            </a>
            <div class="btn-group btn-group-sm ms-2" id="ganttModeToggle" role="group">
              <button type="button" id="btnWeekMode"
                onclick="switchGanttMode('week')"
                class="btn btn-outline-secondary active"
                style="font-size:.78rem;padding:3px 8px">
                <i class="bi bi-calendar3-week me-1"></i>Week
              </button>
              <button type="button" id="btnDayMode"
                onclick="switchGanttMode('day')"
                class="btn btn-outline-secondary"
                style="font-size:.78rem;padding:3px 8px">
                <i class="bi bi-calendar3 me-1"></i>Day
              </button>
            </div>
          </div>
        </div>
        <!-- Edit mode hint bar -->
        <div id="editHint" style="display:none;font-size:.78rem;color:#6d28d9;
          background:#f5f3ff;border:1px solid #c4b5fd;border-radius:6px;
          padding:6px 12px;margin-bottom:8px">
          <i class="bi bi-pencil-square me-1"></i>
          <strong>Edit mode:</strong> Click a task row to select it.
          Drag <strong>right</strong> to add weeks · Drag <strong>left</strong> to mark for deletion · Press <kbd>Delete</kbd> to remove marked cells.
          Preparation steps cascade automatically. Click <strong>Save</strong> when done.
        </div>

        <!-- Add task panel -->
        <div class="collapse mb-3" id="addTaskPanel">
          <div class="card card-body p-3" style="border:1px solid var(--df-border)">
            <form method="post" action="/projects/{pid}/tracker" class="row g-2 align-items-end" onsubmit="syncAddTaskName()">
              <input type="hidden" name="action" value="add">
              <div class="col-6 col-md-2">
                <label class="form-label mb-1" style="font-size:.78rem">Category</label>
                <select class="form-select form-select-sm" name="category" id="add_cat"
                        onchange="onAddCatChange()">
                  <option value="">— select —</option>
                  <option value="Preparation">Preparation</option>
                  <option value="Stress">Stress</option>
                  <option value="Analysis">Analysis</option>
                  <option value="Reporting">Reporting</option>
                  <option value="Non-JEDEC Test">Non-JEDEC Test</option>
                </select>
              </div>
              <div class="col-12 col-md-4" id="add_name_div">
                <label class="form-label mb-1" style="font-size:.78rem">Task Name</label>
                <input type="text" class="form-control form-control-sm" id="add_name_text"
                       placeholder="e.g. Temperature Cycling" required>
              </div>
              <div class="col-12 col-md-4" id="add_stress_div" style="display:none">
                <label class="form-label mb-1" style="font-size:.78rem">Stress Test</label>
                <select class="form-select form-select-sm" id="add_stress_sel" onchange="onAddStressSelChange()">
                  <option value="">— select test —</option>
                  <option value="_other">Other Test...</option>
                </select>
              </div>
              <input type="hidden" name="task_name" id="add_task_name_val">
              <input type="hidden" name="test_key" id="add_test_key">
              <div class="col-6 col-md-1">
                <label class="form-label mb-1" style="font-size:.78rem">Wk # (ISO)</label>
                <input type="number" class="form-control form-control-sm" name="start_week"
                       id="add_sw" value="{current_iso_week}" min="{current_iso_week}">
              </div>
              <div class="col-6 col-md-1">
                <label class="form-label mb-1" style="font-size:.78rem">Duration</label>
                <input type="number" class="form-control form-control-sm" name="duration" value="1" min="1">
              </div>
              <div class="col-6 col-md-2">
                <label class="form-label mb-1" style="font-size:.78rem">Status</label>
                <select class="form-select form-select-sm" name="status">
                  {status_opts_html}
                </select>
              </div>
              <div class="col-12 col-md-3">
                <label class="form-label mb-1" style="font-size:.78rem">Sample Size (n)</label>
                <select class="form-select form-select-sm" name="n_mode"
                        onchange="toggleAddNcust(this)">
                  <option value="auto">Auto</option>
                  <option value="total">Total (all tests)</option>
                  <option value="test">This test only</option>
                  <option value="custom">Custom number</option>
                </select>
              </div>
              <div class="col-6 col-md-1" id="add_ncust_div" style="display:none">
                <label class="form-label mb-1" style="font-size:.78rem">Custom n</label>
                <input type="number" class="form-control form-control-sm" name="n_custom"
                       id="add_ncust" min="1" placeholder="—">
              </div>
              <div class="col-12 col-md-3" id="add_parent_div" style="display:none">
                <label class="form-label mb-1" style="font-size:.78rem">Parent Stress Task</label>
                <select class="form-select form-select-sm" name="parent_task_id" id="add_parent_sel"
                        onchange="onAddParentChange()">
                  {_stress_parent_opts}
                </select>
              </div>
              <div class="col-12 col-md-2">
                <button class="btn btn-sm w-100"
                  style="background:var(--df-accent);color:#fff;border:none">Add Task</button>
              </div>
            </form>
          </div>
        </div>

        <!-- Bulk save form (hidden) -->
        <form id="ganttBulkForm" method="post" action="/projects/{pid}/tracker" style="display:none">
          <input type="hidden" name="action" value="bulk_edit">
          <input type="hidden" name="changes" id="ganttBulkChanges">
        </form>

        {overview_html}

        <!-- Main layout -->
        <div class="card shadow-sm mb-3" style="overflow:hidden" id="ganttCard">
          <div style="display:flex;overflow-x:auto;min-width:0">

            <!-- Sidebar -->
            <div style="min-width:460px;max-width:460px;border-right:2px solid #e5e7eb;flex-shrink:0">
              <table class="table table-sm mb-0" style="border-radius:0;table-layout:fixed">
                <thead>
                  <tr style="background:#f3f4f6">
                    <th style="padding:8px;font-size:.73rem;color:#6b7280;font-weight:600;width:65%">TASK</th>
                    <th style="padding:8px;font-size:.73rem;color:#6b7280;font-weight:600;text-align:center">STATUS</th>
                    <th style="padding:8px;font-size:.73rem;color:#6b7280;font-weight:600;width:80px"></th>
                  </tr>
                </thead>
                <tbody id="ganttSidebody">{task_rows}</tbody>
              </table>
            </div>

            <!-- Chart -->
            <div id="ganttScroll" style="overflow-x:auto;flex:1 1 auto">
              <table id="ganttChartTable" style="border-collapse:collapse;height:100%;user-select:none">
                <thead>
                  <tr style="background:#f3f4f6">{month_headers}</tr>
                  <tr style="background:#f9fafb">{wk_headers}</tr>
                </thead>
                <tbody id="ganttChartBody">{gantt_rows}</tbody>
              </table>
            </div>

          </div>
        </div>

        <!-- Legends -->
        <div class="d-flex flex-wrap gap-3 align-items-center" style="font-size:.75rem;color:#6b7280">
          <div><strong>Status:</strong> {legend}</div>
          {"" if not cat_legend else f"<div><strong>Category:</strong> {cat_legend}</div>"}
        </div>

        <script src="https://cdnjs.cloudflare.com/ajax/libs/Sortable/1.15.2/Sortable.min.js"></script>
        <script>
        // ── Server data ───────────────────────────────────────────────────────────
        const TASK_DATA        = {_task_data_js};
        const PREP_ORDER       = {_prep_order_js};
        const STRESS_TASKS     = {_stress_tasks_js};
        const STRESS_END_BY_ID = {_stress_end_by_id_js};
        const PROJ_STRESS_TESTS = {_proj_stress_tests_js};
        let   prepChainEnd     = {_prep_chain_end};
        const reportingGate    = {_reporting_gate};

        // ── ISO week helpers ──────────────────────────────────────────────────────
        // Anchor date as JS milliseconds (Monday of project start week)
        const ANCHOR_EPOCH_MS  = {anchor_epoch_ms};
        const ANCHOR_ISO_WEEK  = {anchor_iso_week};
        const ANCHOR_YEAR      = {anchor_year};
        const CURRENT_ISO_WEEK = {current_iso_week};
        const N_WEEKS          = {n_weeks};
        const TODAY_DAY_IDX    = {today_day_idx};

        // Convert a 1-based relative week (from anchor) to an ISO week number.
        function relToIso(rel) {{
          const ms  = ANCHOR_EPOCH_MS + (rel - 1) * 7 * 86400000;
          const d   = new Date(ms);
          // ISO week calculation (standard algorithm)
          const tmp = new Date(Date.UTC(d.getUTCFullYear(), d.getUTCMonth(), d.getUTCDate()));
          tmp.setUTCDate(tmp.getUTCDate() + 4 - (tmp.getUTCDay() || 7));
          const yearStart = new Date(Date.UTC(tmp.getUTCFullYear(), 0, 1));
          return Math.ceil((((tmp - yearStart) / 86400000) + 1) / 7);
        }}

        // Convert an ISO week number back to a 1-based relative week from anchor.
        // Year inference: if isoWk >= ANCHOR_ISO_WEEK use anchor year, else anchor year+1.
        function isoToRel(isoWk) {{
          const year = (isoWk >= ANCHOR_ISO_WEEK) ? ANCHOR_YEAR : ANCHOR_YEAR + 1;
          // Monday of ISO week: Jan 4 is always in week 1
          const jan4    = new Date(Date.UTC(year, 0, 4));
          const jan4dow = (jan4.getUTCDay() + 6) % 7;  // 0 = Mon
          const targetMs = jan4.getTime() - jan4dow * 86400000 + (isoWk - 1) * 7 * 86400000;
          const diffDays  = Math.round((targetMs - ANCHOR_EPOCH_MS) / 86400000);
          return Math.max(1, Math.floor(diffDays / 7) + 1);
        }}

        // ── Week / Day mode toggle ────────────────────────────────────────────────
        let ganttMode  = localStorage.getItem('ganttMode') || 'week';
        let _origThead = null;
        let _origTbody = null;
        let _taskOrder = [];   // ordered task IDs captured from server-rendered chart body

        // Per-task day fill sets (parallel to filledWeeks; days are 0-indexed from anchor Mon)
        const filledDays = {{}};
        // Days marked for deletion in day mode (parallel to delSel)
        const delSelDays = new Set();

        // ── Conversion helpers ────────────────────────────────────────────────────
        function filledWeeksToDays(tid) {{
          const fw = filledWeeks[String(tid)];
          if (!fw || fw.size === 0) return new Set();
          const fd = new Set();
          fw.forEach(w => {{
            for (let d = (w - 1) * 7; d <= (w - 1) * 7 + 6; d++) fd.add(d);
          }});
          return fd;
        }}
        function filledDaysToWeeks(tid) {{
          const fd = filledDays[String(tid)];
          if (!fd || fd.size === 0) return new Set();
          const fw = new Set();
          fd.forEach(d => fw.add(Math.floor(d / 7) + 1));
          return fw;
        }}

        function scrollToNow() {{
          const scroller = document.getElementById('ganttScroll');
          if (!scroller) return;
          const nowEl = scroller.querySelector("td[data-now='1']") ||
                        scroller.querySelector("th[data-day='" + TODAY_DAY_IDX + "']");
          if (nowEl) {{
            scroller.scrollLeft = nowEl.offsetLeft - scroller.clientWidth / 2 + nowEl.offsetWidth / 2;
          }}
        }}

        function switchGanttMode(mode) {{
          // Sync state between modes for dirty tasks
          if (mode === 'day' && ganttMode === 'week') {{
            dirtyTids.forEach(tid => {{
              filledDays[String(tid)] = filledWeeksToDays(String(tid));
            }});
            // Also seed filledDays for all tasks on first switch
            Object.keys(TASK_DATA).forEach(tid => {{
              if (!filledDays[tid]) filledDays[tid] = filledWeeksToDays(tid);
            }});
          }} else if (mode === 'week' && ganttMode === 'day') {{
            dirtyTids.forEach(tid => {{
              filledWeeks[String(tid)] = filledDaysToWeeks(String(tid));
            }});
            delSelDays.clear();
          }}
          ganttMode = mode;
          localStorage.setItem('ganttMode', mode);
          document.getElementById('btnWeekMode').classList.toggle('active', mode === 'week');
          document.getElementById('btnDayMode').classList.toggle('active', mode === 'day');
          const table = document.getElementById('ganttChartTable');
          if (!table) return;
          if (mode === 'week') {{
            if (_origThead) table.querySelector('thead').innerHTML = _origThead;
            if (_origTbody) table.querySelector('tbody').innerHTML = _origTbody;
            // Re-render edits in week mode
            if (editActive) ganttRenderAll();
          }} else {{
            buildDayChart(table);
            if (editActive) ganttRenderAll();
          }}
          scrollToNow();
        }}

        function buildDayChart(table) {{
          const nDays     = N_WEEKS * 7;
          const DAY_NAMES = ['M','T','W','T','F','S','S'];

          // ── Month header ──────────────────────────────────────────────────────
          const monthGroups = [];
          let curLabel = null, curSpan = 0;
          for (let d = 0; d < nDays; d++) {{
            const ms  = ANCHOR_EPOCH_MS + d * 86400000;
            const lbl = new Date(ms).toLocaleString('en-US', {{month:'short', year:'numeric', timeZone:'UTC'}});
            if (lbl === curLabel) {{ curSpan++; }}
            else {{
              if (curLabel) monthGroups.push({{lbl: curLabel, span: curSpan}});
              curLabel = lbl; curSpan = 1;
            }}
          }}
          if (curLabel) monthGroups.push({{lbl: curLabel, span: curSpan}});

          const monthHdr = monthGroups.map(g =>
            '<th colspan="' + g.span + '" style="padding:3px 6px;text-align:center;' +
            'font-size:.72rem;color:#374151;font-weight:700;border-left:2px solid #d1d5db;' +
            'white-space:nowrap">' + g.lbl + '</th>'
          ).join('');

          // ── Day header ────────────────────────────────────────────────────────
          let dayHdr = '';
          for (let d = 0; d < nDays; d++) {{
            const dow     = d % 7;   // 0=Mon because anchor is always a Monday
            const isWknd  = dow >= 5;
            const isToday = (d === TODAY_DAY_IDX);
            const dayNum  = new Date(ANCHOR_EPOCH_MS + d * 86400000).getUTCDate();
            const bg      = isToday ? '#fef9c3' : isWknd ? '#ebebeb' : 'transparent';
            const col     = isToday ? '#92400e' : '#9ca3af';
            const fw      = isToday ? '700' : (dow === 0 ? '600' : '400');
            const bl      = dow === 0 ? 'border-left:2px solid #d1d5db;' : 'border-left:1px solid #e5e7eb;';
            const label   = (dow === 0) ? String(dayNum) : DAY_NAMES[dow];
            dayHdr += '<th data-day="' + d + '" style="min-width:22px;padding:2px 1px;text-align:center;' +
              'font-size:.62rem;color:' + col + ';font-weight:' + fw + ';background:' + bg + ';' + bl + '">' +
              label + '</th>';
          }}

          // ── Body rows (use _taskOrder captured from original server render) ───
          let bodyHtml = '';
          const tids = _taskOrder.length > 0 ? _taskOrder : Object.keys(TASK_DATA);
          tids.forEach(function(tid) {{
            const d = TASK_DATA[tid];
            if (!d) return;
            // Use filledDays if available (post-edit), else derive from filledWeeks
            const fd = filledDays[tid] || filledWeeksToDays(tid);
            filledDays[tid] = fd;
            const barColor = d.color;
            let cells = '';
            let prevFill = false;
            for (let day = 0; day < nDays; day++) {{
              const inRange  = fd.has(day);
              const nextFill = fd.has(day + 1);
              const dow      = day % 7;
              const isWknd   = dow >= 5;
              const isToday  = (day === TODAY_DAY_IDX);
              const isStart  = inRange && !fd.has(day - 1);
              const isEnd    = inRange && !nextFill;
              const bl       = isToday  ? 'border-left:2px solid #f59e0b;'
                             : dow === 0 ? 'border-left:2px solid #d1d5db;'
                             : 'border-left:1px solid #e5e7eb;';
              if (inRange) {{
                const br = (isStart && isEnd) ? 'border-radius:4px;'
                         : isStart ? 'border-radius:4px 0 0 4px;'
                         : isEnd   ? 'border-radius:0 4px 4px 0;' : '';
                cells += '<td data-tid="' + tid + '" data-day="' + day + '"' +
                  ' style="background:' + barColor + ';' + br + bl + 'padding:0;height:26px;min-width:22px"></td>';
              }} else {{
                const bgCell = isToday ? '#fef9c3' : isWknd ? '#ebebeb' : '#f9fafb';
                cells += '<td data-tid="' + tid + '" data-day="' + day + '"' +
                  ' style="background:' + bgCell + ';' + bl + 'padding:0;height:26px;min-width:22px"></td>';
              }}
              prevFill = inRange;
            }}
            bodyHtml += '<tr data-gantt-row="' + tid + '">' + cells + '</tr>';
          }});

          table.querySelector('thead').innerHTML =
            '<tr style="background:#f3f4f6">' + monthHdr + '</tr>' +
            '<tr style="background:#f9fafb">' + dayHdr   + '</tr>';
          table.querySelector('tbody').innerHTML = bodyHtml;
        }}

        // Initialise on load
        document.addEventListener('DOMContentLoaded', function() {{
          const table = document.getElementById('ganttChartTable');
          if (table) {{
            _origThead = table.querySelector('thead').innerHTML;
            _origTbody = table.querySelector('tbody').innerHTML;
            // Capture task order from server-rendered chart body
            _taskOrder = Array.from(
              table.querySelectorAll('tbody tr[data-gantt-row]')
            ).map(r => r.dataset.ganttRow);
          }}
          const btnW = document.getElementById('btnWeekMode');
          const btnD = document.getElementById('btnDayMode');
          if (ganttMode === 'day') {{
            if (btnW) btnW.classList.remove('active');
            if (btnD) btnD.classList.add('active');
            if (table) buildDayChart(table);
          }} else {{
            if (btnW) btnW.classList.add('active');
            if (btnD) btnD.classList.remove('active');
            // Re-render bars from JS state (filledWeeks may have been corrected
            // by fixLoadedConstraints before this handler ran).
            ganttRenderAll();
          }}
          dirtyTids.clear();   // constraint auto-fixes are not "unsaved user edits"
          scrollToNow();
        }});

        // ── State ─────────────────────────────────────────────────────────────────
        const filledWeeks  = {{}};
        const dirtyTids    = new Set();
        let   selTid       = null;
        let   drag         = null;
        const delSel       = new Set();
        let   editActive   = false;
        // adminUnlocked is a global var set by the base template
        const undoStack    = [];           // client-side undo history
        const MAX_UNDO     = 30;
        let   originalState = null;        // snapshot taken when entering edit mode

        // Initialise filledWeeks from server data
        Object.entries(TASK_DATA).forEach(([tid, d]) => {{
          const s = new Set();
          for (let w = d.start_week; w < d.start_week + d.duration; w++) s.add(w);
          filledWeeks[tid] = s;
        }});

        // Auto-fix constraint violations in the loaded/saved state.
        // Runs once on page load; only triggers a cascade if a violation is found.
        // Uses hoisted function declarations (cascadeDownstream etc.) which are safe to call here.
        (function fixLoadedConstraints() {{
          const precondTids = _getPrecondTids();
          // 1. Check Preconditioning starts before prep chain end
          const precondViolated = precondTids.some(tid => {{
            const fw = filledWeeks[tid];
            return fw && fw.size > 0 && Math.min(...fw) < prepChainEnd;
          }});
          if (precondViolated) {{ cascadeDownstream(); return; }}
          // 2. Check other stress tasks start before preconditioning end
          const stressEndMap = _buildStressEndMap();
          const precondEndVal = precondTids.length
            ? Math.max(...precondTids.map(tid => stressEndMap[tid] || prepChainEnd))
            : prepChainEnd;
          const stressViolated = STRESS_TASKS.some(t => {{
            if (precondTids.includes(String(t.id))) return false;
            const fw = filledWeeks[String(t.id)];
            const minS = (TASK_DATA[String(t.id)] || {{}}).min_start || precondEndVal;
            return fw && fw.size > 0 && Math.min(...fw) < minS;
          }});
          if (stressViolated) {{ cascadeFromPrecond(); return; }}
          // 3. Check analysis / non-JEDEC / reporting tasks
          const anyChildViolated = Object.entries(TASK_DATA).some(([tid, data]) => {{
            if (!['Analysis', 'Non-JEDEC Test', 'Reporting'].includes(data.category)) return false;
            const fw = filledWeeks[tid];
            const minS = data.min_start || 1;
            return fw && fw.size > 0 && Math.min(...fw) < minS;
          }});
          if (anyChildViolated) cascadeAnalysisAll();
        }})();

        // ── n_mode / category helpers ────────────────────────────────────────────
        var CAT_N_DEFAULT = {{
          'Preparation': 'total', 'Reporting': 'total',
          'Stress': 'test',       'Analysis':  'test',
          'Non-JEDEC Test': 'custom',
        }};
        function toggleNcust() {{
          document.getElementById('e_ncust_div').style.display =
            document.getElementById('e_nmode').value === 'custom' ? '' : 'none';
        }}
        function toggleAddNcust(sel) {{
          document.getElementById('add_ncust_div').style.display =
            sel.value === 'custom' ? '' : 'none';
        }}
        function onEditCatChange() {{
          const cat = document.getElementById('e_cat').value;
          const def = CAT_N_DEFAULT[cat];
          if (def) document.getElementById('e_nmode').value = def;
          toggleNcust();
          const pd = document.getElementById('e_parent_div');
          const needsParent = cat === 'Analysis' || cat === 'Non-JEDEC Test';
          if (pd) pd.style.display = needsParent ? '' : 'none';
          if (needsParent) populateEditParentSel(null);
          // stress name dropdown
          const nd = document.getElementById('e_name_div');
          const sd = document.getElementById('e_stress_div');
          if (cat === 'Stress') {{
            populateEditStressSel(null, '');
            if (nd) nd.style.display = 'none';
            if (sd) sd.style.display = '';
          }} else {{
            if (nd) nd.style.display = '';
            if (sd) sd.style.display = 'none';
            document.getElementById('e_test_key').value = '';
          }}
        }}
        function onAddCatChange() {{
          const cat = document.getElementById('add_cat').value;
          const nmSel = document.querySelector('#addTaskPanel select[name="n_mode"]');
          if (nmSel) {{
            const def = CAT_N_DEFAULT[cat];
            if (def) nmSel.value = def;
            document.getElementById('add_ncust_div').style.display =
              nmSel.value === 'custom' ? '' : 'none';
          }}
          const pd = document.getElementById('add_parent_div');
          if (pd) pd.style.display = (cat === 'Analysis' || cat === 'Non-JEDEC Test') ? '' : 'none';
          // stress name dropdown
          const nd = document.getElementById('add_name_div');
          const sd = document.getElementById('add_stress_div');
          if (cat === 'Stress') {{
            // populate stress sel options
            const stressSel = document.getElementById('add_stress_sel');
            if (stressSel && stressSel.options.length <= 1) {{ // only placeholder
              Object.entries(PROJ_STRESS_TESTS).forEach(([k, name]) => {{
                const o = document.createElement('option');
                o.value = k; o.textContent = name;
                stressSel.insertBefore(o, stressSel.lastElementChild); // before "Other Test..."
              }});
            }}
            if (nd) nd.style.display = 'none';
            if (sd) sd.style.display = '';
            document.getElementById('add_name_text').value = '';
          }} else {{
            if (nd) nd.style.display = '';
            if (sd) sd.style.display = 'none';
            const testKeyInput = document.getElementById('add_test_key');
            if (testKeyInput) testKeyInput.value = '';
          }}
          // Auto-set start week based on dependency rules (display as ISO week)
          const swEl = document.getElementById('add_sw');
          if (swEl) {{
            if (cat === 'Preparation') {{
              swEl.value = relToIso(prepChainEnd);
            }} else if (cat === 'Stress') {{
              swEl.value = relToIso(prepChainEnd);
            }} else if (cat === 'Reporting') {{
              swEl.value = relToIso(reportingGate);
            }} else if (cat === 'Analysis') {{
              // Will be set when parent is selected; reset to blank for now
              swEl.value = '';
            }} else {{
              swEl.value = CURRENT_ISO_WEEK;
            }}
          }}
        }}
        function onAddStressSelChange() {{
          const sel = document.getElementById('add_stress_sel');
          const nameInput = document.getElementById('add_name_text');
          const testKeyInput = document.getElementById('add_test_key');
          if (sel.value === '_other' || !sel.value) {{
            nameInput.value = '';
            nameInput.placeholder = 'Custom test name';
            document.getElementById('add_name_div').style.display = '';
            if (testKeyInput) testKeyInput.value = '';
          }} else {{
            nameInput.value = PROJ_STRESS_TESTS[sel.value] || sel.value;
            document.getElementById('add_name_div').style.display = 'none';
            if (testKeyInput) testKeyInput.value = sel.value;
          }}
        }}

        function syncAddTaskName() {{
          const nameInput = document.getElementById('add_name_text');
          const taskNameVal = document.getElementById('add_task_name_val');
          if (taskNameVal) taskNameVal.value = nameInput.value;
        }}

        function onAddParentChange() {{
          const sel = document.getElementById('add_parent_sel');
          const swEl = document.getElementById('add_sw');
          if (!sel || !swEl) return;
          const pid = sel.value;
          if (pid) {{
            const stressEnd = STRESS_END_BY_ID[String(pid)] || prepChainEnd;
            swEl.value = relToIso(stressEnd);
          }} else {{
            swEl.value = '';
          }}
        }}
        function populateEditParentSel(selectedId) {{
          const sel = document.getElementById('e_parent_sel');
          if (!sel) return;
          sel.innerHTML = '<option value="">— select parent stress task —</option>';
          STRESS_TASKS.forEach(t => {{
            const o = document.createElement('option');
            o.value = t.id; o.textContent = t.name;
            if (selectedId && parseInt(t.id) === parseInt(selectedId)) o.selected = true;
            sel.appendChild(o);
          }});
        }}

        function populateEditStressSel(testKey, taskName) {{
          const sel = document.getElementById('e_stress_sel');
          if (!sel) return;
          sel.innerHTML = '<option value="">— select test —</option>';
          Object.entries(PROJ_STRESS_TESTS).forEach(([k, name]) => {{
            const o = document.createElement('option');
            o.value = k; o.textContent = name;
            if (k === testKey) o.selected = true;
            sel.appendChild(o);
          }});
          const other = document.createElement('option');
          other.value = '_other'; other.textContent = 'Other Test...';
          if (!testKey || !PROJ_STRESS_TESTS[testKey]) other.selected = true;
          sel.appendChild(other);
          // Sync name field
          onEditStressSelChange(taskName);
        }}

        function onEditStressSelChange(customName) {{
          const sel = document.getElementById('e_stress_sel');
          if (!sel) return;
          if (sel.value === '_other' || !sel.value) {{
            document.getElementById('e_name').value = typeof customName === 'string' ? customName : '';
            document.getElementById('e_name_div').style.display = '';
            document.getElementById('e_test_key').value = '';
          }} else {{
            document.getElementById('e_name').value = PROJ_STRESS_TESTS[sel.value] || sel.value;
            document.getElementById('e_name_div').style.display = 'none';
            document.getElementById('e_test_key').value = sel.value;
          }}
        }}

        // ── Edit modal ────────────────────────────────────────────────────────────
        function openEditModal(tid, name, cat, sw, dur, status, nmode, ncust, parentId, testKey) {{
          document.getElementById('e_test_key').value = testKey || '';
          if (cat === 'Stress') {{
            populateEditStressSel(testKey, name);
            document.getElementById('e_name_div').style.display = 'none';
            document.getElementById('e_stress_div').style.display = '';
          }} else {{
            document.getElementById('e_name').value = name;
            document.getElementById('e_name_div').style.display = '';
            document.getElementById('e_stress_div').style.display = 'none';
          }}
          document.getElementById('e_cat').value    = cat;
          document.getElementById('e_sw').value     = relToIso(sw);
          document.getElementById('e_dur').value    = dur;
          document.getElementById('e_status').value = status;
          document.getElementById('e_nmode').value  = nmode || 'auto';
          document.getElementById('e_ncust').value  =
            (ncust !== null && ncust !== undefined) ? ncust : '';
          toggleNcust();
          const pd = document.getElementById('e_parent_div');
          if (pd) pd.style.display = cat === 'Analysis' ? '' : 'none';
          if (cat === 'Analysis') populateEditParentSel(parentId);
          document.getElementById('editForm').action =
            '/projects/{pid}/tracker/task/' + tid + '/edit';
          bootstrap.Modal.getOrCreateInstance(document.getElementById('editModal')).show();
        }}

        // ── Undo stack helpers ────────────────────────────────────────────────────
        function _snapshotFW() {{
          const snap = {{}};
          Object.entries(filledWeeks).forEach(([tid, fw]) => snap[tid] = new Set(fw));
          return snap;
        }}
        function pushUndoState() {{
          if (!editActive) return;
          undoStack.push(_snapshotFW());
          if (undoStack.length > MAX_UNDO) undoStack.shift();
        }}
        function ganttUndo() {{
          if (undoStack.length === 0) return;
          const snap = undoStack.pop();
          Object.entries(snap).forEach(([tid, fw]) => filledWeeks[tid] = new Set(fw));
          delSel.clear(); ganttUpdateDelBtn(); ganttRenderAll();
        }}
        function ganttCancelEdit() {{
          if (originalState) {{
            Object.entries(originalState).forEach(([tid, fw]) => filledWeeks[tid] = new Set(fw));
            dirtyTids.clear();
          }}
          undoStack.length = 0;
          originalState = null;
          ganttDeactivateEdit();
        }}

        function ganttEnterEdit() {{
          if (!adminUnlocked) {{
            ganttAdminLogin();
            return;
          }}
          if (editActive) {{ ganttDeactivateEdit(); return; }}
          editActive = true;
          originalState = _snapshotFW();   // capture state for Cancel
          undoStack.length = 0;
          document.getElementById('ganttEditBtn').style.cssText =
            'border:1px solid #7c3aed;color:#6d28d9;background:#f5f3ff;font-size:.8rem';
          document.getElementById('ganttSaveBtn').style.display = '';
          document.getElementById('ganttCancelBtn').style.display = '';
          document.getElementById('editHint').style.display = '';
          document.querySelectorAll('#ganttSidebody tr').forEach(r => r.style.cursor = 'pointer');
        }}
        function ganttDeactivateEdit() {{
          editActive = false;
          if (selTid !== null) {{
            const prev = document.getElementById('tr-' + selTid);
            if (prev) prev.style.background = '';
          }}
          selTid = null; drag = null; delSel.clear();
          document.getElementById('ganttEditBtn').style.cssText =
            'border:1px solid #c4b5fd;color:#6d28d9;background:#fff;font-size:.8rem';
          document.getElementById('ganttSaveBtn').style.display  = 'none';
          document.getElementById('ganttCancelBtn').style.display = 'none';
          document.getElementById('ganttDelSelBtn').style.display = 'none';
          document.getElementById('editHint').style.display = 'none';
          document.querySelectorAll('#ganttSidebody tr').forEach(r => r.style.cursor = '');
          ganttRenderAll();
        }}

        // ── Row selection ─────────────────────────────────────────────────────────
        function ganttSelectRow(tid) {{
          if (!editActive) return;
          if (selTid !== null) {{
            const prev = document.getElementById('tr-' + selTid);
            if (prev) prev.style.background = '';
          }}
          delSel.clear(); ganttUpdateDelBtn();
          selTid = (String(tid) === String(selTid)) ? null : String(tid);
          if (selTid !== null) {{
            const row = document.getElementById('tr-' + selTid);
            if (row) row.style.background = '#fef3c7';
          }}
          ganttRenderAll();
        }}

        // ── Prep cascade ──────────────────────────────────────────────────────────
        function cascadePrep(fromIdx) {{
          for (let i = fromIdx; i < PREP_ORDER.length; i++) {{
            const tid     = String(PREP_ORDER[i]);
            const prevEnd = i === 0 ? 0 : (() => {{
              const pfw = filledWeeks[String(PREP_ORDER[i-1])] || new Set();
              return pfw.size ? Math.max(...pfw) : 0;
            }})();
            const newStart = prevEnd;
            const fw  = filledWeeks[tid] || new Set();
            const dur = fw.size
              ? Math.max(...fw) - Math.min(...fw) + 1
              : (TASK_DATA[tid] ? TASK_DATA[tid].duration : 1);
            const newFW = new Set();
            for (let w = newStart; w < newStart + dur; w++) newFW.add(w);
            filledWeeks[tid] = newFW;
            if (TASK_DATA[tid]) TASK_DATA[tid].min_start = newStart;
            dirtyTids.add(PREP_ORDER[i]);
          }}
          prepChainEnd = PREP_ORDER.length === 0 ? 0 : (() => {{
            const lfw = filledWeeks[String(PREP_ORDER[PREP_ORDER.length-1])] || new Set();
            return lfw.size ? Math.max(...lfw) : 0;
          }})();
          cascadeDownstream();
        }}

        // ── Downstream cascade: Stress → Analysis → Reporting ─────────────────────
        function _fwDur(tid) {{
          const fw = filledWeeks[String(tid)] || new Set();
          return fw.size
            ? Math.max(...fw) - Math.min(...fw) + 1
            : (TASK_DATA[String(tid)] ? TASK_DATA[String(tid)].duration : 1);
        }}
        function _moveTid(tid, newStart) {{
          const dur = _fwDur(tid);
          const newFW = new Set();
          for (let w = newStart; w < newStart + dur; w++) newFW.add(w);
          filledWeeks[String(tid)] = newFW;
          if (TASK_DATA[String(tid)]) TASK_DATA[String(tid)].min_start = newStart;
          dirtyTids.add(parseInt(tid));
        }}
        function _buildStressEndMap() {{
          const m = {{}};
          STRESS_TASKS.forEach(t => {{
            const fw = filledWeeks[String(t.id)] || new Set();
            m[String(t.id)] = fw.size ? Math.max(...fw) : prepChainEnd;
          }});
          return m;
        }}
        function _getPrecondTids() {{
          return STRESS_TASKS
            .filter(t => (TASK_DATA[String(t.id)] || {{}}).test_key === 'precond')
            .map(t => String(t.id));
        }}

        // Full downstream reset — called after any Prep change.
        function cascadeDownstream() {{
          const precondTids = _getPrecondTids();
          // 1. Preconditioning starts right after prep
          precondTids.forEach(tid => _moveTid(tid, prepChainEnd));
          // 2. All other stress tasks may start on the same week Preconditioning ends
          const sem = _buildStressEndMap();
          const precondEnd = precondTids.length
            ? Math.max(...precondTids.map(tid => sem[tid] || prepChainEnd))
            : prepChainEnd;
          STRESS_TASKS.forEach(t => {{
            if (precondTids.includes(String(t.id))) return;
            _moveTid(t.id, precondEnd);
          }});
          cascadeAnalysisAll();
        }}

        // Called when a Preconditioning bar changes — shifts other stress tasks,
        // then cascades analysis + reporting.
        function cascadeFromPrecond() {{
          const precondTids = _getPrecondTids();
          const sem = _buildStressEndMap();
          const precondEnd = precondTids.length
            ? Math.max(...precondTids.map(tid => sem[tid] || prepChainEnd))
            : prepChainEnd;
          STRESS_TASKS.forEach(t => {{
            if (precondTids.includes(String(t.id))) return;
            _moveTid(t.id, precondEnd);
          }});
          cascadeAnalysisAll();
        }}

        function cascadeAnalysisAll() {{
          const stressEnd = _buildStressEndMap();
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            const isAnalysis   = d.category === 'Analysis';
            const isPostQualNJ = d.category === 'Non-JEDEC Test' && !!d.parent_task_id;
            if (!isAnalysis && !isPostQualNJ) return;
            const parentEnd = stressEnd[String(d.parent_task_id)] ?? prepChainEnd;
            // Both Analysis and Non-JEDEC post-qual may start on the same week/day the parent ends
            const base = parentEnd;
            const fw = filledWeeks[String(tid)] || new Set();
            const curStart = fw.size ? Math.min(...fw) : base;
            const oldMinStart = d.min_start || base;
            const offset = Math.max(0, curStart - oldMinStart);
            _moveTid(tid, base + offset);
          }});
          cascadeReporting();
        }}

        // Live preview: temporarily shift analysis tasks while stress bar is being dragged
        function previewStressCascade(stressTid) {{
          if (!drag || String(drag.tid) !== String(stressTid)) return;
          const fw   = filledWeeks[String(stressTid)] || new Set();
          const data = TASK_DATA[String(stressTid)] || {{}};
          const minS = data.min_start || 1;
          const hi   = Math.max(drag.startW, drag.curW);
          const lo   = Math.min(drag.startW, drag.curW);
          let projectedEnd;
          if (drag.mode === 'fill') {{
            const existingMax = fw.size ? Math.max(...fw) : minS - 1;
            projectedEnd = Math.max(hi, existingMax);
          }} else {{
            // delete: shrink from edges
            const allFw = [...fw].filter(w => !(w >= lo && w <= hi)).sort((a,b) => a-b);
            projectedEnd = allFw.length ? allFw[allFw.length - 1] : minS;
          }}
          Object.entries(TASK_DATA).forEach(([aTid, d]) => {{
            const isAnalysis   = d.category === 'Analysis';
            const isPostQualNJ = d.category === 'Non-JEDEC Test' && !!d.parent_task_id;
            if (!isAnalysis && !isPostQualNJ) return;
            if (String(d.parent_task_id) !== String(stressTid)) return;
            const base = projectedEnd;
            const afw = filledWeeks[String(aTid)] || new Set();
            const dur = _fwDur(aTid);
            const curStart = afw.size ? Math.min(...afw) : base;
            const oldMinStart = d.min_start || base;
            const offset = Math.max(0, curStart - oldMinStart);
            const newStart = base + offset;
            const newFW = new Set();
            for (let w = newStart; w < newStart + dur; w++) newFW.add(w);
            filledWeeks[String(aTid)] = newFW;
          }});
        }}
        function cascadeReporting() {{
          const byParent = {{}};
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            if (d.category !== 'Analysis') return;
            const pid = String(d.parent_task_id || '');
            const fw  = filledWeeks[tid] || new Set();
            const end = fw.size ? Math.max(...fw) : 0;
            if (!byParent[pid]) byParent[pid] = [];
            byParent[pid].push(end);
          }});
          const groups = Object.values(byParent);
          const newGate = groups.length
            ? Math.min(...groups.map(g => Math.max(...g)))
            : prepChainEnd;
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            if (d.category !== 'Reporting') return;
            _moveTid(tid, newGate);
          }});
        }}

        // ── Day-mode cascade helpers ──────────────────────────────────────────────
        function _dayEnd(tid) {{
          const fd = filledDays[String(tid)] || new Set();
          return fd.size ? Math.max(...fd) : -1;
        }}
        function _moveTidDay(tid, newStartDay) {{
          const fd  = filledDays[String(tid)] || new Set();
          const dur = fd.size ? Math.max(...fd) - Math.min(...fd) + 1 : 7;
          const newFd = new Set();
          for (let d = newStartDay; d < newStartDay + dur; d++) newFd.add(d);
          filledDays[String(tid)] = newFd;
          filledWeeks[String(tid)] = filledDaysToWeeks(String(tid));
          dirtyTids.add(parseInt(tid));
        }}
        function _prepEndDay() {{
          let maxDay = -1;
          PREP_ORDER.forEach(tid => {{
            const e = _dayEnd(String(tid));
            if (e > maxDay) maxDay = e;
          }});
          return maxDay < 0 ? 0 : maxDay;
        }}
        function cascadeDownstreamDay() {{
          const precondTids = _getPrecondTids();
          const prepEnd = _prepEndDay();
          precondTids.forEach(tid => _moveTidDay(tid, prepEnd));
          const precondEnd = precondTids.length
            ? Math.max(...precondTids.map(tid => _dayEnd(tid)))
            : prepEnd;
          STRESS_TASKS.forEach(t => {{
            if (precondTids.includes(String(t.id))) return;
            _moveTidDay(t.id, precondEnd);
          }});
          cascadeAnalysisAllDay();
        }}
        function cascadeFromPrecondDay() {{
          const precondTids = _getPrecondTids();
          const precondEnd = precondTids.length
            ? Math.max(...precondTids.map(tid => _dayEnd(tid)))
            : _prepEndDay();
          STRESS_TASKS.forEach(t => {{
            if (precondTids.includes(String(t.id))) return;
            _moveTidDay(t.id, precondEnd);
          }});
          cascadeAnalysisAllDay();
        }}
        function cascadeAnalysisAllDay() {{
          const stressEndDay = {{}};
          STRESS_TASKS.forEach(t => {{
            stressEndDay[String(t.id)] = _dayEnd(String(t.id));
          }});
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            const isAnalysis   = d.category === 'Analysis';
            const isPostQualNJ = d.category === 'Non-JEDEC Test' && !!d.parent_task_id;
            if (!isAnalysis && !isPostQualNJ) return;
            const parentEnd = stressEndDay[String(d.parent_task_id)] ?? _prepEndDay();
            const fd = filledDays[String(tid)] || new Set();
            const curStart = fd.size ? Math.min(...fd) : parentEnd;
            const oldMin = d.min_start ? (d.min_start - 1) * 7 : parentEnd;
            const offset = Math.max(0, curStart - oldMin);
            _moveTidDay(tid, parentEnd + offset);
          }});
          cascadeReportingDay();
        }}
        function cascadeReportingDay() {{
          const byParent = {{}};
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            if (d.category !== 'Analysis') return;
            const pid = String(d.parent_task_id || '');
            const end = _dayEnd(String(tid));
            if (!byParent[pid]) byParent[pid] = [];
            byParent[pid].push(end);
          }});
          const groups = Object.values(byParent);
          const newGate = groups.length
            ? Math.min(...groups.map(g => Math.max(...g)))
            : _prepEndDay();
          Object.entries(TASK_DATA).forEach(([tid, d]) => {{
            if (d.category !== 'Reporting') return;
            _moveTidDay(tid, newGate);
          }});
        }}
        // After any week-mode cascade, sync filledDays for tasks that moved
        function syncAllDaysFromWeeks() {{
          Object.keys(TASK_DATA).forEach(tid => {{
            filledDays[tid] = filledWeeksToDays(tid);
          }});
        }}

        // ── Render ────────────────────────────────────────────────────────────────
        function ganttRenderRow(tid) {{
          if (ganttMode === 'day') {{ ganttRenderRowDay(tid); return; }}
          const row = document.querySelector('[data-gantt-row="' + tid + '"]');
          if (!row) return;
          const data  = TASK_DATA[String(tid)] || {{}};
          const fw    = filledWeeks[String(tid)] || new Set();
          const col   = data.color || '#9ca3af';
          const isSel = String(tid) === String(selTid);
          const minS  = data.min_start || 1;
          row.querySelectorAll('td').forEach(cell => {{
            const w = parseInt(cell.dataset.week);
            if (!w) return;
            const isFill   = fw.has(w);
            const isNow    = cell.dataset.now === '1';
            const inDelSel = isSel && delSel.has(w);
            let inFill = false, inDel = false;
            if (drag && String(drag.tid) === String(tid) && drag.mode) {{
              const lo = Math.min(drag.startW, drag.curW);
              const hi = Math.max(drag.startW, drag.curW);
              if (w >= lo && w <= hi) {{
                inFill = drag.mode === 'fill';
                inDel  = drag.mode === 'delete';
              }}
            }}
            let bg, cur;
            if      (inFill && w >= minS) {{ bg = '#93c5fd'; cur = 'crosshair'; }}
            else if (inDel || inDelSel)   {{ bg = '#fca5a5'; cur = 'pointer'; }}
            else if (isFill)              {{ bg = col; cur = editActive && isSel ? 'pointer' : 'default'; }}
            else if (w < minS && editActive && isSel) {{ bg = '#fee2e2'; cur = 'not-allowed'; }}
            else if (isNow)              {{ bg = '#fef9c3'; cur = editActive && isSel && w >= minS ? 'crosshair' : 'default'; }}
            else                          {{ bg = isSel && editActive ? '#ede9fe' : '#f9fafb';
                                            cur = editActive && isSel && w >= minS ? 'crosshair' : 'default'; }}
            cell.style.background = bg;
            cell.style.cursor     = cur;
          }});
        }}

        function ganttRenderRowDay(tid) {{
          const row = document.querySelector('[data-gantt-row="' + tid + '"]');
          if (!row) return;
          const data  = TASK_DATA[String(tid)] || {{}};
          const fd    = filledDays[String(tid)] || new Set();
          const col   = data.color || '#9ca3af';
          const isSel = String(tid) === String(selTid);
          row.querySelectorAll('td').forEach(cell => {{
            const day = parseInt(cell.dataset.day);
            if (isNaN(day)) return;
            const isFill  = fd.has(day);
            const isToday = (day === TODAY_DAY_IDX);
            const dow     = day % 7;
            const isWknd  = dow >= 5;
            const inDelSel = isSel && delSelDays.has(day);
            let inFill = false, inDel = false;
            if (drag && drag.dayMode && String(drag.tid) === String(tid) && drag.mode) {{
              const lo = Math.min(drag.startD, drag.curD);
              const hi = Math.max(drag.startD, drag.curD);
              if (day >= lo && day <= hi) {{
                inFill = drag.mode === 'fill';
                inDel  = drag.mode === 'delete';
              }}
            }}
            let bg, cur;
            if      (inFill)              {{ bg = '#93c5fd'; cur = 'crosshair'; }}
            else if (inDel || inDelSel)   {{ bg = '#fca5a5'; cur = 'pointer'; }}
            else if (isFill)              {{ bg = col; cur = editActive && isSel ? 'pointer' : 'default'; }}
            else if (isToday)             {{ bg = '#fef9c3'; cur = editActive && isSel ? 'crosshair' : 'default'; }}
            else                          {{ bg = isSel && editActive ? '#ede9fe' : isWknd ? '#ebebeb' : '#f9fafb';
                                            cur = editActive && isSel ? 'crosshair' : 'default'; }}
            cell.style.background = bg;
            cell.style.cursor     = cur;
          }});
        }}

        function ganttRenderAll() {{
          Object.keys(TASK_DATA).forEach(tid => ganttRenderRow(tid));
        }}

        // ── Drag ──────────────────────────────────────────────────────────────────
        const chartTable = document.getElementById('ganttChartTable');
        if (chartTable) {{
          chartTable.addEventListener('mousedown', e => {{
            if (!editActive) return;
            const cell = e.target.closest('[data-tid]');
            if (!cell) return;
            const cellTid = String(cell.dataset.tid);
            if (cellTid !== String(selTid)) {{
              if (selTid !== null) {{
                const prev = document.getElementById('tr-' + selTid);
                if (prev) prev.style.background = '';
              }}
              delSel.clear(); delSelDays.clear(); ganttUpdateDelBtn();
              selTid = cellTid;
              const row = document.getElementById('tr-' + selTid);
              if (row) row.style.background = '#fef3c7';
              ganttRenderAll();
            }}
            e.preventDefault();
            if (ganttMode === 'day') {{
              const day = parseInt(cell.dataset.day);
              drag = {{tid: selTid, startD: day, curD: day, mode: null, dayMode: true}};
            }} else {{
              const w = parseInt(cell.dataset.week);
              drag = {{tid: selTid, startW: w, curW: w, mode: null, dayMode: false}};
            }}
            ganttRenderRow(selTid);
          }});
          chartTable.addEventListener('mousemove', e => {{
            if (!drag) return;
            const cell = e.target.closest('[data-tid]');
            if (!cell) return;
            if (drag.dayMode) {{
              const day = parseInt(cell.dataset.day);
              if (drag.mode === null && day !== drag.startD)
                drag.mode = day > drag.startD ? 'fill' : 'delete';
              if (day !== drag.curD) {{
                drag.curD = day;
                ganttRenderRow(drag.tid);
              }}
            }} else {{
              const w = parseInt(cell.dataset.week);
              if (drag.mode === null && w !== drag.startW)
                drag.mode = w > drag.startW ? 'fill' : 'delete';
              if (w !== drag.curW) {{
                drag.curW = w;
                const dragData = TASK_DATA[String(drag.tid)] || {{}};
                if (dragData.category === 'Stress' && drag.mode !== null) {{
                  previewStressCascade(drag.tid);
                  ganttRenderAll();
                }} else {{
                  ganttRenderRow(drag.tid);
                }}
              }}
            }}
          }});
        }}
        // Snap all filled weeks for a task to be contiguous (no gaps)
        function _snapContiguous(tid) {{
          const fw = filledWeeks[String(tid)];
          if (!fw || fw.size < 2) return;
          const lo = Math.min(...fw), hi = Math.max(...fw);
          for (let w = lo; w <= hi; w++) fw.add(w);
        }}

        function ganttCommitDrag() {{
          if (!drag) return;

          // ── Day mode commit ────────────────────────────────────────────────────
          if (drag.dayMode) {{
            const tid  = String(drag.tid);
            const data = TASK_DATA[tid] || {{}};
            const fd   = filledDays[tid] || new Set();
            const lo   = Math.min(drag.startD, drag.curD);
            const hi   = Math.max(drag.startD, drag.curD);
            const mode = drag.mode || (fd.has(drag.startD) ? 'delete' : 'fill');

            if (!drag.mode && lo === hi && fd.has(lo) && delSelDays.has(lo)) {{
              delSelDays.delete(lo);
              ganttUpdateDelBtn(); ganttRenderRow(tid); drag = null; return;
            }}
            if (mode === 'fill') {{
              pushUndoState();
              for (let d = lo; d <= hi; d++) fd.add(d);
              // Snap contiguous
              if (fd.size >= 2) {{
                const dlo = Math.min(...fd), dhi = Math.max(...fd);
                for (let d = dlo; d <= dhi; d++) fd.add(d);
              }}
              filledDays[tid] = fd;
              filledWeeks[tid] = filledDaysToWeeks(tid);
              dirtyTids.add(parseInt(tid));
              if (data.category === 'Preparation') {{
                cascadeDownstreamDay();
              }} else if (data.category === 'Stress') {{
                if (_getPrecondTids().includes(tid)) cascadeFromPrecondDay();
                else cascadeAnalysisAllDay();
              }} else if (data.category === 'Analysis') {{
                cascadeReportingDay();
              }}
              ganttRenderAll();
            }} else {{
              for (let d = lo; d <= hi; d++) {{ if (fd.has(d)) delSelDays.add(d); }}
              ganttUpdateDelBtn(); ganttRenderRow(tid);
            }}
            drag = null;
            return;
          }}

          // ── Week mode commit ───────────────────────────────────────────────────
          const tid  = String(drag.tid);
          const data = TASK_DATA[tid] || {{}};
          const fw   = filledWeeks[tid] || new Set();
          const minS = data.min_start || 1;
          const lo   = Math.min(drag.startW, drag.curW);
          const hi   = Math.max(drag.startW, drag.curW);
          const mode = drag.mode || (fw.has(drag.startW) ? 'delete' : 'fill');

          // Single-click on an already-red (delSel) cell → toggle it off
          if (!drag.mode && lo === hi && fw.has(lo) && delSel.has(lo)) {{
            delSel.delete(lo);
            ganttUpdateDelBtn();
            ganttRenderRow(tid);
            drag = null;
            return;
          }}

          if (mode === 'fill') {{
            pushUndoState();
            if (data.locked_start) {{
              // Prep: extend end to hi, start locked at minS
              const newEnd = Math.max(hi, minS);
              const newFW = new Set();
              for (let w = minS; w <= newEnd; w++) newFW.add(w);
              filledWeeks[tid] = newFW;
            }} else {{
              for (let w = Math.max(lo, minS); w <= hi; w++) fw.add(w);
              // Snap contiguous: bridge any gap between existing cells and new fill
              _snapContiguous(tid);
            }}
            dirtyTids.add(parseInt(tid));
            const idx = data.prep_idx !== undefined ? data.prep_idx : -1;
            if (data.category === 'Preparation' && idx >= 0) {{
              cascadePrep(idx + 1);
            }} else if (data.category === 'Stress') {{
              if (_getPrecondTids().includes(tid)) {{
                cascadeFromPrecond();
              }} else {{
                cascadeAnalysisAll();
              }}
            }} else if (data.category === 'Analysis') {{
              cascadeReporting();
            }}
            syncAllDaysFromWeeks();
            ganttRenderAll();
          }} else {{
            for (let w = lo; w <= hi; w++) {{ if (fw.has(w)) delSel.add(w); }}
            ganttUpdateDelBtn();
            ganttRenderRow(tid);
          }}
          drag = null;
        }}
        document.addEventListener('mouseup', () => {{ ganttCommitDrag(); }});

        // ── Delete ────────────────────────────────────────────────────────────────
        function ganttDeleteSel() {{
          if (!selTid) return;
          pushUndoState();
          const tid  = String(selTid);
          const data = TASK_DATA[tid] || {{}};
          const fw   = filledWeeks[tid] || new Set();
          const minS = data.min_start || 1;
          if (data.locked_start) {{
            const marked = [...delSel].filter(w => fw.has(w));
            if (marked.length) {{
              const newEnd = Math.min(...marked) - 1;
              const newFW = new Set();
              if (newEnd >= minS) for (let w = minS; w <= newEnd; w++) newFW.add(w);
              filledWeeks[tid] = newFW;
              delSel.clear(); ganttUpdateDelBtn();
              const idx = data.prep_idx !== undefined ? data.prep_idx : -1;
              if (idx >= 0) cascadePrep(idx + 1);
              ganttRenderAll();
            }} else {{
              delSel.clear(); ganttUpdateDelBtn();
            }}
          }} else {{
            // Edge-trim only: shrink bar from start or end to maintain contiguity
            const allFw = [...fw].sort((a,b) => a-b);
            if (allFw.length > 0) {{
              let newMin = allFw[0], newMax = allFw[allFw.length-1];
              while (delSel.has(newMin) && newMin <= newMax) newMin++;
              while (delSel.has(newMax) && newMax >= newMin) newMax--;
              const newFW = new Set();
              if (newMin <= newMax) for (let w = newMin; w <= newMax; w++) newFW.add(w);
              filledWeeks[tid] = newFW;
            }}
            delSel.clear(); ganttUpdateDelBtn();
            if (data.category === 'Stress') {{
              if (_getPrecondTids().includes(tid)) {{
                cascadeFromPrecond();
              }} else {{
                cascadeAnalysisAll();
              }}
              syncAllDaysFromWeeks(); ganttRenderAll();
            }} else if (data.category === 'Analysis') {{
              cascadeReporting(); syncAllDaysFromWeeks(); ganttRenderAll();
            }} else {{
              syncAllDaysFromWeeks(); ganttRenderRow(tid);
            }}
          }}
          dirtyTids.add(parseInt(tid));
        }}
        function ganttUpdateDelBtn() {{
          const btn = document.getElementById('ganttDelSelBtn');
          if (btn) btn.style.display = delSel.size > 0 ? '' : 'none';
        }}

        // ── Keyboard shortcuts ────────────────────────────────────────────────────
        document.addEventListener('keydown', e => {{
          if ((e.metaKey || e.ctrlKey) && e.key === 'z' && !e.shiftKey) {{
            e.preventDefault();
            if (editActive) {{
              ganttUndo();  // client-side undo while editing
            }} else {{
              const uf = document.getElementById('undoForm');
              if (uf) uf.submit();  // server-side undo when not in edit mode
            }}
          }}
          const tag = document.activeElement ? document.activeElement.tagName : '';
          const inInput = tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT';
          if ((e.key === 'Delete' || e.key === 'Backspace') && editActive && !inInput) {{
            if (ganttMode === 'day' && delSelDays.size > 0) {{ e.preventDefault(); ganttDeleteSelDays(); }}
            else if (ganttMode === 'week' && delSel.size > 0) {{ e.preventDefault(); ganttDeleteSel(); }}
          }}
          if (e.key === 'Enter' && editActive && drag && drag.mode === 'fill' && !inInput) {{
            e.preventDefault(); ganttCommitDrag();
          }}
        }});

        // ── Day-mode delete ───────────────────────────────────────────────────────
        function ganttDeleteSelDays() {{
          if (!selTid) return;
          pushUndoState();
          const tid = String(selTid);
          const fd  = filledDays[tid] || new Set();
          const allDays = [...fd].sort((a,b) => a-b);
          if (allDays.length > 0) {{
            let lo = allDays[0], hi = allDays[allDays.length-1];
            while (delSelDays.has(lo) && lo <= hi) lo++;
            while (delSelDays.has(hi) && hi >= lo) hi--;
            const newFd = new Set();
            if (lo <= hi) for (let d = lo; d <= hi; d++) newFd.add(d);
            filledDays[tid] = newFd;
            filledWeeks[tid] = filledDaysToWeeks(tid);
          }}
          delSelDays.clear(); ganttUpdateDelBtn();
          dirtyTids.add(parseInt(tid));
          ganttRenderAll();
        }}

        // ── Save ──────────────────────────────────────────────────────────────────
        function ganttSave() {{
          if (ganttMode === 'day' && delSelDays.size > 0) ganttDeleteSelDays();
          else if (ganttMode === 'week' && delSel.size > 0) ganttDeleteSel();
          const changes = [];
          dirtyTids.forEach(tid => {{
            // In day mode, convert filledDays → weeks before saving
            let fw;
            if (ganttMode === 'day') {{
              fw = filledDaysToWeeks(String(tid));
              filledWeeks[String(tid)] = fw;
            }} else {{
              fw = filledWeeks[String(tid)];
            }}
            if (!fw || fw.size === 0) return;
            const sw = Math.min(...fw), ew = Math.max(...fw);
            changes.push({{id: tid, start_week: sw, duration: ew - sw + 1}});
          }});
          if (changes.length === 0) {{ ganttDeactivateEdit(); return; }}
          document.getElementById('ganttBulkChanges').value = JSON.stringify(changes);
          document.getElementById('ganttBulkForm').submit();
        }}

        // ── Drag-to-reorder sidebar ───────────────────────────────────────────────
        (function() {{
          const sideBody  = document.getElementById('ganttSidebody');
          const chartBody = document.getElementById('ganttChartBody');
          if (!sideBody || typeof Sortable === 'undefined') return;
          Sortable.create(sideBody, {{
            handle: '.gantt-drag-handle', animation: 150, ghostClass: 'sortable-ghost',
            onEnd: function() {{
              const rows = sideBody.querySelectorAll('tr[id^="tr-"]');
              const orderedIds = [...rows].map(r => r.id.replace('tr-', ''));
              orderedIds.forEach(tid => {{
                const cr = chartBody.querySelector('[data-gantt-row="' + tid + '"]');
                if (cr) chartBody.appendChild(cr);
              }});
              fetch('/projects/{pid}/tracker', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                body: 'action=reorder&order=' + encodeURIComponent(JSON.stringify(orderedIds))
              }}).catch(() => {{}});
            }}
          }});
        }})();

        // ── Live progress bars ────────────────────────────────────────────────────
        (function() {{
          function updateTestBars() {{
            document.querySelectorAll('.test-progress-bar').forEach(bar => {{
              var started = bar.dataset.started, dur = parseFloat(bar.dataset.duration);
              if (!started || !dur) return;
              var elapsed = (Date.now() - new Date(started + 'Z').getTime()) / 3600000;
              var pct = Math.min(100, elapsed / dur * 100);
              bar.style.width = pct.toFixed(1) + '%';
              bar.style.background = pct >= 100 ? '#16a34a' : '#f97316';
              var txt = bar.parentElement.nextElementSibling;
              if (txt && txt.classList.contains('test-progress-txt'))
                txt.textContent = elapsed.toFixed(1) + 'h / ' + dur.toFixed(0) + 'h';
            }});
          }}
          updateTestBars();
          setInterval(updateTestBars, 60000);
        }})();
        </script>
        """


class ProjectTaskHandler(Base):
    """Handles edit and delete of individual GANTT tasks."""

    def post(self, pid, tid, action):
        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404); return
        tid = int(tid)

        # Fetch current task state before mutating (for undo)
        all_tasks = _db.list_gantt_tasks(p["id"])
        prev = next((t for t in all_tasks if t["id"] == tid), None)

        if action == "delete":
            if prev:
                _db.push_gantt_history(p["id"], "delete_task", prev)
            _db.delete_gantt_task(tid, p["id"])
        elif action == "edit":
            if prev:
                _db.push_gantt_history(p["id"], "edit_task", prev)
            _nm = self.get_argument("n_mode", "auto").strip() or "auto"
            _nc_raw = self.get_argument("n_custom", "").strip()
            _nc = int(_nc_raw) if _nc_raw.isdigit() else None
            _par_raw = self.get_argument("parent_task_id", "").strip()
            _par = int(_par_raw) if _par_raw.isdigit() else None
            _edit_anchor  = _get_gantt_anchor(p["id"])
            _iso_sw_raw   = self.get_argument("start_week", "").strip()
            _iso_sw       = int(_iso_sw_raw) if _iso_sw_raw.isdigit() else _edit_anchor.isocalendar()[1]
            _rel_sw       = _iso_week_to_relative(_edit_anchor, _iso_sw)
            _db.update_gantt_task(
                tid, p["id"],
                task_name  = self.get_argument("task_name",  ""),
                category   = self.get_argument("category",   ""),
                start_week = _rel_sw,
                duration   = int(self.get_argument("duration",    "1")),
                status     = self.get_argument("status", "not_started"),
                n_mode     = _nm,
                n_custom   = _nc,
                parent_task_id = _par,
            )
        self.redirect(f"/projects/{p['id']}/tracker")


class ProjectTrackerXlsxHandler(Base):
    """Return the GANTT schedule as a downloadable Excel workbook with visual chart."""

    def get(self, pid):
        import io as _io
        from datetime import date as _date, timedelta as _td
        try:
            import openpyxl
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.set_status(500)
            self.finish("openpyxl not installed"); return

        p = _db.get_project(int(pid))
        if not p:
            self.send_error(404); return

        tasks = _db.list_gantt_tasks(p["id"])
        meta  = _db.get_meta(p["id"])

        start_str = (meta.get("gantt_start_date") or "").strip()
        try:
            anchor = _date.fromisoformat(start_str)
        except ValueError:
            anchor = _date.today()
            anchor -= _td(days=anchor.weekday())

        def week_to_date(wk: int) -> _date:
            return anchor + _td(weeks=wk - 1)

        status_labels = {
            "not_started": "Not Started",
            "in_progress":  "In Progress",
            "complete":     "Complete",
            "blocked":      "Blocked",
            "na":           "N/A",
        }

        # Category → fill color (matches web GANTT sidebar palette)
        _cat_palette = ["3B82F6", "8B5CF6", "EC4899", "F97316", "14B8A6", "64748B"]
        _cats_seen = list(dict.fromkeys(
            (t.get("category") or "").strip() for t in tasks
            if (t.get("category") or "").strip()
        ))
        cat_hex = {c: _cat_palette[i % len(_cat_palette)]
                   for i, c in enumerate(_cats_seen)}

        def cat_fill(cat: str) -> PatternFill:
            h = cat_hex.get(cat, "64748B")
            return PatternFill("solid", fgColor="FF" + h)

        def cat_font(cat: str) -> Font:
            return Font(name="Arial", size=9, bold=True, color="FFFFFFFF")

        thin = Side(style="thin", color="FFD1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Compute week range
        max_week = max((t["start_week"] + t["duration"] - 1 for t in tasks), default=1)
        n_weeks  = max_week + 2  # small buffer

        # ── Build workbook ───────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GANTT Chart"

        # ── Info columns: A=#  B=Task Name  C=Category  D=Status ───────────
        INFO_COLS = 4   # columns before week grid starts
        WEEK_START_COL = INFO_COLS + 1  # = column 5 (E)

        # Row 1: month banners (merged across weeks sharing a month)
        # Row 2: week headers (Wk N  +  date)
        HDR_ROW1 = 1
        HDR_ROW2 = 2
        DATA_START_ROW = 3

        # ── Header row 1: month groupings ───────────────────────────────────
        hdr1_fill = PatternFill("solid", fgColor="FF1E3A5F")
        hdr1_font = Font(name="Arial", size=8, bold=True, color="FFFFFFFF")
        hdr2_fill = PatternFill("solid", fgColor="FF2D4E7E")
        hdr2_font = Font(name="Arial", size=7, bold=True, color="FFFFFFFF")
        info_hdr_fill = PatternFill("solid", fgColor="FF1E3A5F")
        info_hdr_font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center")

        # Info header labels
        for col, label in enumerate(["#", "Task Name", "Category", "Status"], 1):
            c = ws.cell(row=HDR_ROW1, column=col, value=label)
            c.fill, c.font, c.alignment, c.border = \
                info_hdr_fill, info_hdr_font, center, hdr_border
            # span both header rows
        ws.merge_cells(start_row=HDR_ROW1, start_column=1,
                       end_row=HDR_ROW2,   end_column=1)
        ws.merge_cells(start_row=HDR_ROW1, start_column=2,
                       end_row=HDR_ROW2,   end_column=2)
        ws.merge_cells(start_row=HDR_ROW1, start_column=3,
                       end_row=HDR_ROW2,   end_column=3)
        ws.merge_cells(start_row=HDR_ROW1, start_column=4,
                       end_row=HDR_ROW2,   end_column=4)
        # Re-apply style to anchor cells only — merged secondary cells are
        # MergedCell objects in openpyxl and do not accept style attributes.
        for col in range(1, 5):
            c = ws.cell(row=HDR_ROW1, column=col)
            c.fill, c.font, c.alignment, c.border = \
                info_hdr_fill, info_hdr_font, center, hdr_border

        # Month grouping in row 1 + week numbers in row 2
        month_start_col = None
        cur_month       = None
        for wk in range(1, n_weeks + 1):
            col = WEEK_START_COL + (wk - 1)
            d   = week_to_date(wk)
            mo  = (d.year, d.month)
            # Row 2: week number
            c2 = ws.cell(row=HDR_ROW2, column=col,
                         value=f"W{d.isocalendar()[1]}")
            c2.fill, c2.font = hdr2_fill, hdr2_font
            c2.alignment     = Alignment(horizontal="center", vertical="center",
                                          text_rotation=90)
            c2.border        = hdr_border
            # Row 1: month label — merge on transition
            if mo != cur_month:
                if cur_month is not None and month_start_col is not None:
                    end_col = col - 1
                    if end_col > month_start_col:
                        ws.merge_cells(start_row=HDR_ROW1, start_column=month_start_col,
                                       end_row=HDR_ROW1,   end_column=end_col)
                    ws.cell(row=HDR_ROW1, column=month_start_col).alignment = center
                cur_month       = mo
                month_start_col = col
                label = d.strftime("%b %Y")
                c1 = ws.cell(row=HDR_ROW1, column=col, value=label)
                c1.fill, c1.font, c1.alignment, c1.border = \
                    hdr1_fill, hdr1_font, center, hdr_border
            else:
                c1 = ws.cell(row=HDR_ROW1, column=col)
                c1.fill, c1.border = hdr1_fill, hdr_border
        # close last month merge
        if month_start_col is not None:
            end_col = WEEK_START_COL + n_weeks - 1
            if end_col > month_start_col:
                ws.merge_cells(start_row=HDR_ROW1, start_column=month_start_col,
                               end_row=HDR_ROW1,   end_column=end_col)
            ws.cell(row=HDR_ROW1, column=month_start_col).alignment = center

        # ── Data rows ───────────────────────────────────────────────────────
        alt_fill  = PatternFill("solid", fgColor="FFF9FAFB")
        base_fill = PatternFill("solid", fgColor="FFFFFFFF")
        empty_fill = PatternFill("solid", fgColor="FFF3F4F6")
        task_font = Font(name="Arial", size=9)
        num_font  = Font(name="Arial", size=9, color="FF6B7280")

        for i, t in enumerate(tasks):
            row   = DATA_START_ROW + i
            sw    = t["start_week"]
            ew    = sw + t["duration"] - 1
            cat   = (t.get("category") or "").strip()
            stat  = status_labels.get(t.get("status", "not_started"),
                                       t.get("status", ""))
            row_bg = alt_fill if i % 2 else base_fill

            # Col A: row number
            ca = ws.cell(row=row, column=1, value=i + 1)
            ca.font, ca.alignment, ca.fill, ca.border = \
                num_font, center, row_bg, border

            # Col B: task name
            cb = ws.cell(row=row, column=2, value=t["task_name"])
            cb.font, cb.alignment, cb.fill, cb.border = \
                task_font, left, row_bg, border

            # Col C: category (colored badge)
            cc = ws.cell(row=row, column=3, value=cat)
            cc.fill, cc.font, cc.alignment, cc.border = \
                cat_fill(cat), cat_font(cat), center, border

            # Col D: status
            cd = ws.cell(row=row, column=4, value=stat)
            cd.font, cd.alignment, cd.fill, cd.border = \
                task_font, center, row_bg, border

            # Week cells
            bar_fill = cat_fill(cat)
            for wk in range(1, n_weeks + 1):
                col = WEEK_START_COL + (wk - 1)
                c = ws.cell(row=row, column=col)
                c.border = border
                if sw <= wk <= ew:
                    c.fill = bar_fill
                else:
                    c.fill = empty_fill
                c.alignment = center

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 13
        for wk in range(1, n_weeks + 1):
            ws.column_dimensions[get_column_letter(WEEK_START_COL + wk - 1)].width = 3.2

        # ── Row heights ──────────────────────────────────────────────────────
        ws.row_dimensions[HDR_ROW1].height = 14
        ws.row_dimensions[HDR_ROW2].height = 42   # rotated text needs height
        for i in range(len(tasks)):
            ws.row_dimensions[DATA_START_ROW + i].height = 18

        # ── Freeze panes at first data cell after info cols ──────────────────
        ws.freeze_panes = ws.cell(row=DATA_START_ROW,
                                   column=WEEK_START_COL)

        # ── Legend sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Legend")
        ws2["A1"] = p["name"]
        ws2["A1"].font = Font(name="Arial", size=12, bold=True)
        ws2["A2"] = f"Anchor date: {anchor.strftime('%B %d, %Y')}"
        ws2["A2"].font = Font(name="Arial", size=9, color="FF6B7280")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 18
        row_l = 4
        ws2.cell(row=row_l, column=1, value="Category").font = \
            Font(name="Arial", size=9, bold=True)
        ws2.cell(row=row_l, column=2, value="Color").font = \
            Font(name="Arial", size=9, bold=True)
        for cat in _cats_seen:
            row_l += 1
            ws2.cell(row=row_l, column=1, value=cat).font = Font(name="Arial", size=9)
            lc = ws2.cell(row=row_l, column=2)
            lc.fill = cat_fill(cat)
            lc.font = cat_font(cat)
            lc.value = cat
            lc.alignment = Alignment(horizontal="center", vertical="center")

        # ── Serialize ────────────────────────────────────────────────────────
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in p["name"])
        filename  = f"gantt_{safe_name}.xlsx"
        self.set_header("Content-Type",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")
        self.set_header("Content-Disposition",
                        f'attachment; filename="{filename}"')
        self.finish(buf.read())


# ── App routing & entry point ─────────────────────────────────────────────────

def make_app():
    specs_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "specs")
    return tornado.web.Application(
        [
            (r"/",                                   IndexHandler),
            (r"/part-type",                          PartTypeHandler),
            (r"/lookup",                             LookupHandler),
            (r"/sample-size",                        SampleSizeHandler),
            (r"/pass-fail",                          PassFailHandler),
            (r"/report",                             ReportHandler),
            (r"/report/pdf",                         ReportPdfHandler),
            # ── Project management ──────────────────────────────────────────
            (r"/projects",                           ProjectListHandler),
            (r"/projects/new",                       ProjectNewHandler),
            (r"/projects/(\d+)",                     ProjectDetailHandler),
            (r"/projects/(\d+)/delete",              ProjectDeleteHandler),
            (r"/projects/(\d+)/meta",                ProjectMetaHandler),
            (r"/projects/(\d+)/sample-size",         ProjectSampleSizeHandler),
            (r"/projects/(\d+)/pass-fail",           ProjectPassFailHandler),
            (r"/projects/(\d+)/report",              ProjectReportHandler),
            (r"/projects/(\d+)/report/pdf",          ProjectReportPdfHandler),
            (r"/projects/(\d+)/csam",                ProjectCsamHandler),
            (r"/projects/(\d+)/csam/(\d+)",          ProjectCsamImageHandler),
            (r"/projects/(\d+)/csam/(\d+)/thumb",    ProjectCsamThumbHandler),
            (r"/projects/(\d+)/csam/(\d+)/delete",   ProjectCsamDeleteHandler),
            (r"/projects/(\d+)/tracker",             ProjectTrackerHandler),
            (r"/projects/(\d+)/tracker/xlsx",        ProjectTrackerXlsxHandler),
            (r"/projects/(\d+)/tracker/task/(\d+)/(edit|delete)", ProjectTaskHandler),
            # ── Static spec PDFs ────────────────────────────────────────────
            (r"/specs/(.*)",                         tornado.web.StaticFileHandler,
                                                     {"path": specs_path}),
        ],
        cookie_secret=COOKIE_SECRET,
        debug=False,
    )


if __name__ == "__main__":
    _db.init_db()          # create / migrate SQLite schema on first run
    PORT = int(os.environ.get("PORT", 5001))
    app = make_app()
    app.listen(PORT, address="0.0.0.0")
    url = f"http://localhost:{PORT}"
    print(f"\n  Package Reliability Qualification Suite")
    print(f"  ──────────────────────────────────────────")
    print(f"  Server running at {url}")
    print(f"  Press Ctrl+C to stop.\n")
    if PORT == 5001:
        webbrowser.open(url)
    tornado.ioloop.IOLoop.current().start()
